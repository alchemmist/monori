"""
Deterministic generator for web/e2e/fixtures/template-workbook.xlsx — the
cut-down "live template" workbook the workbook e2e spec migrates through the
real UI (a miniature of the user's YNAB-like spreadsheet: Russian T-Bank
transaction headers, keyword side table, a year sheet with a budget grid).

Regenerate with:  cd server && uv run python ../scripts/make-e2e-workbook.py

The script self-checks the artifact against the real template importer: the
cached grid figures must reconcile exactly (no synthetic adjustment rows) and
the counts must match what the spec hardcodes.
"""

import datetime
import pathlib
import sys
from io import BytesIO
from typing import TYPE_CHECKING, Protocol

from openpyxl import Workbook

if TYPE_CHECKING:
    from collections.abc import Callable

    from openpyxl.worksheet.worksheet import Worksheet

    from app.workbook.models import ParsedWorkbook


class _SpecModule(Protocol):
    SHEET_TRANSACTIONS: str


REPO = pathlib.Path(__file__).resolve().parent.parent


def _load_workbook_modules() -> tuple[
    _SpecModule, dict[str, tuple[str, ...]], Callable[[bytes], ParsedWorkbook]
]:
    sys.path.insert(0, str(REPO / "server"))
    from app.workbook import spec
    from app.workbook.parser import TX_ALIASES, parse_workbook

    return spec, TX_ALIASES, parse_workbook


spec, TX_ALIASES, parse_workbook = _load_workbook_modules()

OUT = REPO / "web" / "e2e" / "fixtures" / "template-workbook.xlsx"

TX_HEADER = [
    TX_ALIASES[f][0]
    for f in ("date", "card", "status", "amount", "currency", "bank_category", "mcc", "description")
]
WorkbookCell = datetime.datetime | float | int | str | None


def tx(
    date: datetime.datetime,
    amount: float,
    category: str,
    *,
    card: str = "*1111",
    desc: str = "",
    kw: tuple[str, str] | None = None,
) -> list[WorkbookCell]:
    row: list[WorkbookCell] = [None] * 12
    row[0] = date
    row[1] = card
    row[2] = "OK"
    row[3] = amount
    row[4] = "RUB"
    row[5] = "Super"
    row[6] = "5411"
    row[7] = desc
    row[9] = category
    if kw:
        row[10], row[11] = kw
    return row


def build() -> Workbook:
    wb = Workbook()
    active = wb.active
    assert active is not None
    wb.remove(active)

    ws: Worksheet = wb.create_sheet(spec.SHEET_TRANSACTIONS)
    ws.append(TX_HEADER)
    rows = [
        tx(
            datetime.datetime(2026, 1, 5),
            50000.0,
            "Salary",
            card="*2222",
            desc="PAYROLL JAN",
            kw=("Groceries", "lenta|okey"),
        ),
        tx(
            datetime.datetime(2026, 1, 15),
            -3000.0,
            "Groceries",
            desc="LENTA-101",
            kw=("Cafe", "coffee|cafe"),
        ),
        tx(
            datetime.datetime(2026, 2, 10),
            -4500.0,
            "Groceries",
            desc="OKEY 55",
            kw=("Salary", "payroll"),
        ),
        tx(datetime.datetime(2026, 2, 14), -1200.0, "Cafe", desc="COFFEE POINT"),
        tx(datetime.datetime(2026, 2, 20), -800.0, "", desc="MISC SHOP"),
    ]
    for row in rows:
        ws.append(row)

    # year grid: Jan+Feb 2026, cached figures consistent with the rows above
    # so the importer reconciles without synthetic adjustments
    ys: Worksheet = wb.create_sheet("2026")
    header_row = 8
    bases = [2, 6]
    ys.cell(row=1, column=bases[0], value="ЯНВ 2026")
    ys.cell(row=header_row, column=1, value="Категория")
    for b in bases:
        ys.cell(row=header_row, column=b, value="Бюджет")
        ys.cell(row=header_row, column=b + 1, value="Расход")
        ys.cell(row=header_row, column=b + 2, value="Баланс")
    grid = [
        ("▼Daily", None),
        ("Groceries", {1: (5000, 3000, 2000), 2: (5000, 4500, 2500)}),
        ("Cafe", {1: (None, 0, 0), 2: (1200, 1200, 0)}),
        ("▲Inflow", None),
        ("Salary", None),
    ]
    r = header_row + 1
    for label, vals in grid:
        ys.cell(row=r, column=1, value=label)
        if vals:
            for mi, mnum in enumerate((1, 2)):
                if mnum in vals:
                    b = bases[mi]
                    budget, outflow, balance = vals[mnum]
                    if budget is not None:
                        ys.cell(row=r, column=b, value=budget)
                    if outflow is not None:
                        ys.cell(row=r, column=b + 1, value=outflow)
                    if balance is not None:
                        ys.cell(row=r, column=b + 2, value=balance)
        r += 1
    ys.cell(row=2, column=bases[0] + 2, value="Income for January")
    ys.cell(row=2, column=bases[0] + 1, value=50000)
    for mi, avail in ((0, 45000), (1, 38800)):
        b = bases[mi]
        ys.cell(row=6, column=b + 1, value="Available")
        ys.cell(row=5, column=b + 1, value=avail)

    return wb


def main() -> None:
    wb = build()
    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    parsed = parse_workbook(data)
    synthetic_names = {"Salary", "Groceries", "Cafe", "Opening balance"}
    synthetic = [t for t in parsed.transactions if t.description in synthetic_names]
    markers = sorted({t.marker for t in parsed.transactions if t.marker})
    counts = {
        "groups": len(parsed.groups),
        "categories": len(parsed.categories),
        "transactions": len(parsed.transactions),
        "budget_cells": len(parsed.budgets),
    }
    print(f"parsed: {counts}, markers={markers}")
    # the miniature must parse spotlessly — a warning means the fixture itself
    # drifted from what the importer expects, so fail instead of narrating
    assert not parsed.warnings, parsed.warnings
    assert not synthetic, f"cached grid does not reconcile, synthetic rows: {synthetic}"
    assert not parsed.errors, parsed.errors

    # counts alone would not catch a column-index slip in tx() — pin the
    # content of one row per domain as well
    by_desc = {t.description: t for t in parsed.transactions}
    lenta = by_desc["LENTA-101"]
    assert lenta.amount == -300000, lenta
    assert lenta.monori_category == "Groceries", lenta
    assert lenta.marker == "*1111", lenta
    assert by_desc["PAYROLL JAN"].amount == 5000000, by_desc["PAYROLL JAN"]
    assert by_desc["MISC SHOP"].monori_category == "", by_desc["MISC SHOP"]
    keywords = {c.name: c.keywords for c in parsed.categories}
    assert keywords["Groceries"] == "lenta|okey", keywords
    cell = next(b for b in parsed.budgets if b.category == "Groceries" and b.month == 1)
    assert (cell.year, cell.amount) == (2026, 500000), cell
    # 3 categories, all from the grid: the workbook already has an income group,
    # so the reader has somewhere to put rebuilt income and invents nothing
    assert counts == {"groups": 2, "categories": 3, "transactions": 5, "budget_cells": 3}, counts
    assert markers == ["*1111", "*2222"], markers

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_bytes(data)
    print(f"wrote {OUT} ({len(data)} bytes)")


if __name__ == "__main__":
    main()

"""
Explain a migration mismatch: replay what monori will show for a workbook and
compare it with the numbers the sheet itself cached.

    uv run --project server python scripts/diagnose_workbook.py book.xlsx
    uv run --project server python scripts/diagnose_workbook.py book.xlsx 2025-07

Without a month it prints the running Available monori's budget engine rebuilds,
month by month, so the first one that drifts from the sheet is visible. With a
month it dumps that month's grid per category, sheet value beside rebuilt value.
Nothing is written anywhere; the workbook is only read.
"""

import pathlib
import sys

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent.parent / "server"))

from app.workbook.parser import (  # noqa: E402
    YEAR_RE,
    _find_layout,
    _parse_year_sheet,
    parse_workbook,
)
from openpyxl import load_workbook  # noqa: E402

# web/src/App.jsx starts the budget engine here and never earlier
CLIENT_FIRST_YEAR = 2020


def rub(kop):
    return "             ." if kop is None else f"{kop / 100:>14,.2f}"


def read_sheets(path):
    """The year grids as the sheet cached them, keyed by year."""
    wb = load_workbook(path, data_only=True)
    try:
        sheets = {}
        for name in wb.sheetnames:
            match = YEAR_RE.match(name)
            if not match:
                continue
            layout = _find_layout(wb[name])
            if layout is None:
                continue
            sheets[name] = _parse_year_sheet(wb[name], int(match.group(1)), layout)
        return sheets
    finally:
        wb.close()


def replay(parsed):
    """budget.js, on the rows and budgets this workbook produces."""
    kind_of_group = {g["name"]: g["kind"] for g in parsed["groups"]}
    kind = {c["name"]: kind_of_group.get(c["group"], "expense") for c in parsed["categories"]}
    tx: dict[tuple, int] = {}
    for row in parsed["transactions"]:
        cat = row["monori_category"]
        if not cat:
            continue
        key = (cat, int(row["date"][:4]), int(row["date"][5:7]))
        tx[key] = tx.get(key, 0) + row["amount"]
    budget: dict[tuple, int] = {}
    for cell in parsed["budgets"]:
        key = (cell["category"], cell["year"], cell["month"])
        budget[key] = budget.get(key, 0) + cell["amount"]
    return kind, tx, budget


def main(path, only_month=None):
    parsed = parse_workbook(pathlib.Path(path).read_bytes())
    sheets = read_sheets(path)
    kind, tx, budget = replay(parsed)

    print("== sheets ==")
    for name, sheet in sorted(sheets.items()):
        print(f"  {name:<14} months {sheet['months']}  seed={rub(sheet['seed']).strip()}")

    print("\n== groups ==")
    for g in parsed["groups"]:
        names = [c["name"] for c in parsed["categories"] if c["group"] == g["name"]]
        print(f"  {g['kind']:<8} {g['name']}  ({len(names)}) {', '.join(names[:12])}")

    print("\n== warnings ==")
    for w in parsed["warnings"]:
        print(f"  - {w}")
    if parsed["errors"]:
        print(f"  ({len(parsed['errors'])} unparseable rows)")

    dropped = [b for b in parsed["budgets"] if kind.get(b["category"]) == "income"]
    if dropped:
        print("\n== budget cells on income-side categories ==")
        print("   monori's grid budgets expense categories only — these never show up")
        for b in sorted(dropped, key=lambda b: (b["year"], b["month"], b["category"])):
            print(f"  {b['year']}-{b['month']:02d} {b['category']:<30}{rub(b['amount'])}")

    expense_cats = [c["name"] for c in parsed["categories"] if kind[c["name"]] != "income"]
    income_cats = [c["name"] for c in parsed["categories"] if kind[c["name"]] == "income"]
    by_year = {}
    for name, sheet in sheets.items():
        if name.endswith("_archive") and sheet["year"] in by_year:
            continue
        by_year[sheet["year"]] = sheet
    last_year = max(
        [*(b["year"] for b in parsed["budgets"]), *(y for _, y, _ in tx)] or [CLIENT_FIRST_YEAR]
    )

    if only_month:
        year, month = (int(p) for p in only_month.split("-"))
        sheet = by_year.get(year)
        cats = sheet["cats"] if sheet else {}
        print(f"\n== {only_month} per category: sheet vs monori ==")
        print(f"{'category':<32}{'sheet bud':>15}{'ours':>15}{'sheet out':>15}{'ours':>15}")
        for c in parsed["categories"]:
            name = c["name"]
            entry = cats.get(name, {})
            print(
                f"{name:<32}{rub(entry.get('budgets', {}).get(month))}"
                f"{rub(budget.get((name, year, month)))}"
                f"{rub(entry.get('outflows', {}).get(month))}"
                f"{rub(tx.get((name, year, month)))}"
            )
        missing = sorted(set(cats) - {c["name"] for c in parsed["categories"]})
        if missing:
            print(f"\n  rows in the sheet with no category in monori: {', '.join(missing)}")
        return

    print("\n== Available: what monori rebuilds vs what the sheet cached ==")
    head = f"{'month':<9}{'income':>15}{'budgeted':>15}{'overspent':>15}{'ours':>15}{'sheet':>15}"
    print(head)
    balances: dict[str, int] = {}
    avail = 0
    prev_overspent = 0
    for year in range(CLIENT_FIRST_YEAR, last_year + 1):
        for m in range(1, 13):
            income = sum(tx.get((c, year, m), 0) for c in income_cats)
            budgeted = sum(budget.get((c, year, m), 0) for c in expense_cats)
            overspent = 0
            for c in expense_cats:
                bal = (
                    max(balances.get(c, 0), 0)
                    + budget.get((c, year, m), 0)
                    + tx.get((c, year, m), 0)
                )
                balances[c] = bal
                overspent += min(bal, 0)
            avail = avail + prev_overspent + income - budgeted
            prev_overspent = overspent
            sheet = by_year.get(year)
            cached = sheet["available"].get(m) if sheet else None
            if income or budgeted or overspent or cached is not None:
                gap = "" if cached is None else f"   gap {rub(cached - avail).strip()}"
                print(
                    f"{year}-{m:02d}  {rub(income)}{rub(budgeted)}"
                    f"{rub(overspent)}{rub(avail)}{rub(cached)}{gap}"
                )


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    main(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)

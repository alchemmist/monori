"""
End-to-end migration parity: import a workbook the way the server does, then
check that the budget monori shows equals the one the spreadsheet cached.

    uv run --project server python scripts/verify_migration.py book.xlsx
    uv run --project server python scripts/verify_migration.py book.xlsx --category Vacation

Runs the real path — ``parse_workbook`` into ``apply_workbook`` against a
throwaway database — reads the snapshot back out, replays the client's budget
engine on it, and diffs every category/month cell (budgeted, outflows, balance)
plus the running Available against the numbers the sheet computed for itself.
Also reports duplicate rows and anything imported uncategorized. Exits non-zero
when a cell disagrees, so it can guard a change to the importer. ``--category``
traces one category month by month instead, for reading a single disagreement.

Nothing outside the temporary database is written; the workbook is only read.
"""

import argparse
import pathlib
import sqlite3
import sys
import tempfile
from dataclasses import dataclass

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent.parent / "server"))

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.db import connect
from app.workbook.apply import apply_workbook
from app.workbook.models import ParsedWorkbook, WorkbookApplyResult
from app.workbook.parser import (
    YEAR_RE,
    _find_layout,
    _parse_year_sheet,
    _s,
    _summary_value,
    account_slot,
    parse_workbook,
)

MONEY = 100

MAX_MONTH = 12
MONTHS_IN_YEAR = 12
HEADER_RECONCILIATION_TOLERANCE = 5
SNAPSHOT_RECONCILIATION_TOLERANCE = 2

@dataclass(frozen=True, slots=True)
class DuplicateRow:
    """Detected duplicate workbook row used in import idempotence checks."""

    accounts: str
    date: str
    amount: int
    description: str
    count: int


@dataclass(frozen=True, slots=True)
class LedgerSnapshot:
    """In-memory snapshot of workbook-derived ledger data."""

    categories: dict[int, tuple[str, str]]
    transactions: dict[tuple[str, int, int], int]
    budgets: dict[tuple[str, int, int], int]
    uncategorized: int


@dataclass(frozen=True, slots=True)
class Header:
    """Year-header summary values used as reconciliation anchors."""

    overspent: int | None
    budgeted: int | None
    income: int | None
    carried: int | None
    available: int | None


@dataclass(frozen=True, slots=True)
class YearCat:
    """Per-category monthly aggregates for a given year."""

    budgets: dict[int, int]
    outflows: dict[int, int]
    balances: dict[int, int]


@dataclass(frozen=True, slots=True)
class YearGrid:
    """Parsed year-wide budget grid keyed by category."""

    months: list[int]
    cats: dict[str, YearCat]


@dataclass(frozen=True, slots=True)
class ReplayCell:
    """Replay validation value for one cell."""

    budgeted: int
    outflows: int
    balance: int


@dataclass(frozen=True, slots=True)
class ReplayMonth:
    """Aggregated monthly totals for reconciliation against UI state."""

    income: int
    budgeted: int
    overspent: int
    available: int
    cells: dict[str, ReplayCell]


def rub(kop: int | None) -> str:
    """Rub for this module."""
    return "." if kop is None else f"{kop / MONEY:,.2f}"


def sheet_grids(path: str) -> tuple[dict[int, YearGrid], dict[tuple[int, int], Header]]:
    """Year grids keyed by year, plus the header figures, as the sheet cached them."""
    wb = load_workbook(path, data_only=True)
    try:
        grids: dict[int, YearGrid] = {}
        headers: dict[tuple[int, int], Header] = {}
        for name in wb.sheetnames:
            match = YEAR_RE.match(name)
            if not match:
                continue
            ws: Worksheet = wb[name]
            layout = _find_layout(ws)
            if layout is None:
                continue
            year = int(match.group(1))
            if match.group(2) and year in grids:
                continue  # a plain sheet is the working copy and wins
            parsed_year = _parse_year_sheet(ws, year, layout)
            grids[year] = YearGrid(
                months=parsed_year.months,
                cats={
                    category: YearCat(row.budgets, row.outflows, row.balances)
                    for category, row in parsed_year.cats.items()
                },
            )
            for i, base in enumerate(layout.bases):
                month = layout.start_month + i
                if month > MAX_MONTH:
                    break
                headers[(year, month)] = Header(
                    overspent=_summary_value(ws, base, ("Overspent in", "Перерасход")),
                    budgeted=_summary_value(ws, base, ("Budgeted in", "Заложено")),
                    income=_summary_value(ws, base, ("Income for", "Поступления в")),
                    carried=_summary_value(ws, base, ("Not budgeted", "Не заложено")),
                    available=_available(ws, base),
                )
        return grids, headers
    finally:
        wb.close()


def _available(ws: Worksheet, base: int) -> int | None:
    for r in (5, 6):
        if _s(ws.cell(r + 1, base + 1)).startswith(("Available", "Доступный")):
            from app.workbook.parser import _kop

            return _kop(ws.cell(r, base + 1))
    return None


def import_into_db(
    path: str, db_path: str
) -> tuple[sqlite3.Connection, int, ParsedWorkbook, WorkbookApplyResult]:
    """The server's own path: parse, map every slot onto a matching account, apply."""
    parsed = parse_workbook(pathlib.Path(path).read_bytes())
    c = connect(db_path)
    c.execute(
        "INSERT INTO users (email, email_canonical, password_hash, created_at)"
        " VALUES ('parity@local', 'parity@local', 'x', '2020-01-01')"
    )
    uid = c.execute("SELECT id FROM users").fetchone()["id"]
    mapping: dict[str, int] = {}
    for row in parsed.transactions:
        key = account_slot(row)
        if key in mapping:
            continue
        currency = key.split(":", 1)[0]
        cur = c.execute(
            "INSERT INTO accounts (user_id, name, currency) VALUES (?, ?, ?)",
            (uid, key, currency),
        )
        lastrowid = cur.lastrowid
        if lastrowid is None:
            message = "inserted account did not return a row id"
            raise RuntimeError(message)
        mapping[key] = lastrowid
    result = apply_workbook(c, uid, parsed, mapping, "overwrite")
    c.commit()
    return c, uid, parsed, result


def snapshot(c: sqlite3.Connection, uid: int) -> LedgerSnapshot:
    """Snapshot for this module."""
    kinds = {
        r["id"]: r["kind"]
        for r in c.execute("SELECT id, kind FROM category_groups WHERE user_id=?", (uid,))
    }
    cats = {
        r["id"]: (r["name"], kinds.get(r["group_id"], "expense"))
        for r in c.execute(
            "SELECT c.id, c.name, c.group_id FROM categories c"
            " JOIN category_groups g ON g.id = c.group_id WHERE g.user_id=?",
            (uid,),
        )
    }
    tx: dict[tuple[str, int, int], int] = {}
    uncategorized = 0
    for r in c.execute(
        "SELECT t.date, t.amount, t.category_id FROM transactions t"
        " JOIN accounts a ON a.id = t.account_id WHERE a.user_id=? AND t.transfer_id IS NULL",
        (uid,),
    ):
        if r["category_id"] is None:
            uncategorized += 1
            continue
        name = cats[r["category_id"]][0]
        key = (name, int(r["date"][:4]), int(r["date"][5:7]))
        tx[key] = tx.get(key, 0) + r["amount"]
    budgets: dict[tuple[str, int, int], int] = {}
    for r in c.execute(
        "SELECT b.year, b.month, b.amount, c.name FROM budgets b"
        " JOIN categories c ON c.id = b.category_id"
        " JOIN category_groups g ON g.id = c.group_id WHERE g.user_id=?",
        (uid,),
    ):
        budgets[(r["name"], r["year"], r["month"])] = r["amount"]
    return LedgerSnapshot(cats, tx, budgets, uncategorized)


def duplicates(c: sqlite3.Connection, uid: int) -> tuple[list[DuplicateRow], list[DuplicateRow]]:
    """
    Rows a human would read as the same entry twice — on one account, or the
    same day/amount/description spread over several accounts, which is how a
    feed delivered through two doors looks once routing has split the copies.
    """
    on_one: list[DuplicateRow] = [
        DuplicateRow(
            str(r["name"]),
            str(r["date"]),
            int(r["amount"]),
            str(r["description"]),
            int(r["n"]),
        )
        for r in c.execute(
            "SELECT a.name, t.date, t.amount, t.description, COUNT(*) n"
            " FROM transactions t JOIN accounts a ON a.id = t.account_id"
            " WHERE a.user_id=? GROUP BY t.account_id, t.date, t.amount, t.description"
            " HAVING n > 1 ORDER BY n DESC, t.date",
            (uid,),
        )
    ]
    across: list[DuplicateRow] = [
        DuplicateRow(
            str(r["names"]),
            str(r["day"]),
            int(r["amount"]),
            str(r["description"]),
            int(r["n"]),
        )
        for r in c.execute(
            "SELECT group_concat(DISTINCT a.name) names, substr(t.date, 1, 10) day,"
            " t.amount, t.description, COUNT(*) n"
            " FROM transactions t JOIN accounts a ON a.id = t.account_id"
            " WHERE a.user_id=? AND t.amount != 0"
            " GROUP BY day, t.amount, t.description"
            " HAVING n > 1 AND COUNT(DISTINCT t.account_id) > 1"
            " ORDER BY n DESC, day",
            (uid,),
        )
    ]
    return on_one, across


def _header_adds_up(head: Header) -> bool | None:
    """
    Whether the month's own header cells make its Available: carried over +
    overspent + income - budgeted. The template's formula has been edited by
    hand over the years and in places dropped a term, so a month can disagree
    with itself; monori applies the envelope rule to every month alike.
    """
    if (
        head.available is None
        or head.carried is None
        or head.overspent is None
        or head.income is None
        or head.budgeted is None
    ):
        return None
    carried = head.carried
    overspent = head.overspent
    income = head.income
    budgeted = head.budgeted
    available = head.available
    calculated = carried + overspent + income + budgeted
    return abs(calculated - available) <= HEADER_RECONCILIATION_TOLERANCE


def _carried_overspend(
    grids: dict[int, YearGrid], head: Header, year: int, month: int
) -> bool | None:
    """
    Whether the month's "Overspent in <previous>" cell equals what the previous
    month's own Balance column actually adds up to. A cell whose formula was
    repointed and never recalculated keeps a figure from a different month.
    """
    cached = head.overspent
    if cached is None:
        return None
    prev_year, prev_month = (year - 1, 12) if month == 1 else (year, month - 1)
    grid = grids.get(prev_year)
    if grid is None or prev_month not in grid.months:
        return None
    total = sum(min(entry.balances.get(prev_month, 0), 0) for entry in grid.cats.values())
    return abs(total - cached) <= SNAPSHOT_RECONCILIATION_TOLERANCE


def _self_consistent(grids: dict[int, YearGrid], name: str, year: int, month: int) -> bool | None:
    """
    Whether the sheet's own three numbers for this cell agree with each other:
    carried balance + budgeted + outflows == balance. When they do not, the cell
    is a stale cache in the spreadsheet and monori follows the balance, which is
    the figure the sheet displays and carries forward.
    """
    entry = grids[year].cats.get(name)
    if entry is None:
        return None
    prev_year, prev_month = (year - 1, 12) if month == 1 else (year, month - 1)
    prev_grid = grids.get(prev_year)
    if prev_grid is None:
        prev = None
    else:
        prev_cat = prev_grid.cats.get(name)
        prev = None if prev_cat is None else prev_cat.balances.get(prev_month)
    budgeted = entry.budgets.get(month)
    outflows = entry.outflows.get(month)
    balance = entry.balances.get(month)
    if None in (budgeted, outflows, balance):
        return None
    assert budgeted is not None
    assert outflows is not None
    assert balance is not None
    prev_value = 0 if prev is None else prev
    calculated = max(prev_value, 0) + budgeted + outflows
    return abs(calculated - balance) <= SNAPSHOT_RECONCILIATION_TOLERANCE


def replay(
    cats: dict[int, tuple[str, str]],
    tx: dict[tuple[str, int, int], int],
    budgets: dict[tuple[str, int, int], int],
    first_year: int,
    last_year: int,
) -> dict[tuple[int, int], ReplayMonth]:
    """web/src/engine/budget.js, on what the database actually holds."""
    expense = sorted({n for n, k in cats.values() if k != "income"})
    income_cats = sorted({n for n, k in cats.values() if k == "income"})
    balances: dict[str, int] = {}
    avail = 0
    prev_overspent = 0
    out: dict[tuple[int, int], ReplayMonth] = {}
    for year in range(first_year, last_year + 1):
        for m in range(1, MONTHS_IN_YEAR + 1):
            income = sum(tx.get((n, year, m), 0) for n in income_cats)
            budgeted = sum(budgets.get((n, year, m), 0) for n in expense)
            overspent = 0
            cells: dict[str, ReplayCell] = {}
            for n in expense:
                bal = (
                    max(balances.get(n, 0), 0)
                    + budgets.get((n, year, m), 0)
                    + tx.get((n, year, m), 0)
                )
                balances[n] = bal
                if bal < 0:
                    overspent += bal
                cells[n] = ReplayCell(
                    budgeted=budgets.get((n, year, m), 0),
                    outflows=tx.get((n, year, m), 0),
                    balance=bal,
                )
            avail = avail + prev_overspent + income - budgeted
            prev_overspent = overspent
            out[(year, m)] = ReplayMonth(income, budgeted, overspent, avail, cells)
    return out


def trace(
    grids: dict[int, YearGrid], replayed: dict[tuple[int, int], ReplayMonth], name: str
) -> None:
    """One category, month by month, the sheet's three numbers beside ours."""
    print(f"=== {name} ===")
    cols = (
        "month",
        "sheet bud",
        "sheet out",
        "sheet bal",
        "our bud",
        "our out",
        "our bal",
    )
    print(f"{cols[0]:<9}" + "".join(f"{h:>15}" for h in cols[1:]))
    for year in sorted(grids):
        grid = grids[year]
        entry = grid.cats.get(name, YearCat({}, {}, {}))
        for m in grid.months:
            month = replayed.get((year, m))
            ours = None if month is None else month.cells.get(name)
            if ours is None:
                continue
            sheet = [entry.budgets.get(m), entry.outflows.get(m), entry.balances.get(m)]
            mine = [ours.budgeted, ours.outflows, ours.balance]
            if not any(sheet) and not any(mine):
                continue
            print(f"{year}-{m:02d}  " + "".join(f"{rub(v):>15}" for v in sheet + mine))


def main() -> int:
    """Run this module as a CLI entrypoint and return its exit code."""
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--report", default="", help="write the full mismatch list here")
    ap.add_argument("--limit", type=int, default=25, help="mismatches to print per section")
    ap.add_argument("--category", default="", help="trace one category month by month instead")
    args = ap.parse_args()

    grids, headers = sheet_grids(args.workbook)
    with tempfile.TemporaryDirectory() as tmp:
        db_path = str(pathlib.Path(tmp) / "parity.db")
        c, uid, parsed, result = import_into_db(args.workbook, db_path)
        ledger = snapshot(c, uid)
        dupes, cross_dupes = duplicates(c, uid)
        c.close()

    years = sorted(grids)
    first_year = min(
        [
            *(year for _, year, _ in ledger.transactions),
            *(year for _, year, _ in ledger.budgets),
            *years,
        ]
        or [years[0] if years else 0]
    )
    replayed = replay(
        ledger.categories, ledger.transactions, ledger.budgets, first_year, max(years)
    )

    if args.category:
        trace(grids, replayed, args.category)
        return 0

    lines = []
    bad_cells = []
    for year in years:
        grid = grids[year]
        for name, entry in grid.cats.items():
            for m in grid.months:
                ours = replayed[(year, m)].cells.get(name)
                if ours is None:
                    continue  # income rows have no envelope; checked via the totals
                for field, want, ours_value in (
                    ("budgeted", entry.budgets.get(m), ours.budgeted),
                    ("outflows", entry.outflows.get(m), ours.outflows),
                    ("balance", entry.balances.get(m), ours.balance),
                ):
                    if want is None:
                        continue
                    if abs(want - ours_value) > SNAPSHOT_RECONCILIATION_TOLERANCE:
                        bad_cells.append(
                            (
                                year,
                                m,
                                name,
                                field,
                                want,
                                ours_value,
                                _self_consistent(grids, name, year, m),
                            )
                        )

    # Available is a running total, so a gap opened once is repeated by every
    # later month. Only the months where it *changes* say anything.
    bad_months = []
    carried_gap = 0
    for (year, m), head in sorted(headers.items()):
        avail_ours = replayed.get((year, m))
        want = head.available
        if avail_ours is None or want is None:
            continue
        gap = want - avail_ours.available
        if abs(gap - carried_gap) > HEADER_RECONCILIATION_TOLERANCE:
            stale = _header_adds_up(head)
            if stale is not False and _carried_overspend(grids, head, year, m) is False:
                stale = False
            bad_months.append((year, m, want, avail_ours.available, gap - carried_gap, stale))
        carried_gap = gap

    lines.append(f"transactions imported : {result.inserted}")
    lines.append(f"skipped as duplicate  : {result.skipped}")
    lines.append(f"budget cells written  : {result.budgets_written}")
    lines.append(f"budget cells skipped  : {result.budgets_skipped}")
    lines.append(f"rows left uncategorized: {ledger.uncategorized}")
    lines.append(
        f"duplicate row groups  : {len(dupes)} on one account, {len(cross_dupes)} across accounts"
    )
    lines.append(f"grid cells wrong      : {len(bad_cells)}")
    lines.append(f"months where the Available gap changes: {len(bad_months)}")
    lines.append("")
    lines.extend(f"  warning: {w}" for w in [*parsed.warnings, *result.warnings])
    if dupes:
        lines.append("\n-- duplicate rows (account, date, amount, description, count) --")
        lines += [
            f"  {row.accounts} {row.date[:10]} {rub(row.amount):>14} "
            f"{row.description[:40]!r} x{row.count}"
            for row in dupes
        ]
    if cross_dupes:
        lines.append(
            "\n-- same day/amount/description on several accounts (for review:"
            " the sheet itself records these on distinct cards) --"
        )
        lines += [
            f"  {row.accounts} {row.date} {rub(row.amount):>14} "
            f"{row.description[:40]!r} x{row.count}"
            for row in cross_dupes
        ]
    if bad_cells:
        lines.append("\n-- grid cells: sheet vs monori --")
        lines.append("   'sheet stale' = the sheet's own budgeted+outflows do not make its balance")
        lines += [
            f"  {y}-{m:02d} {n[:28]:<28} {f:<9} sheet {rub(w):>14}  ours {rub(g):>14}"
            f"  d {rub(w - g):>12}  {'sheet stale' if ok is False else ''}"
            for y, m, n, f, w, g, ok in bad_cells
        ]
    if bad_months:
        lines.append("\n-- Available: months where the gap to the sheet changes --")
        lines.append("   'sheet stale' = the sheet's own header cells do not make its own total")
        lines += [
            f"  {y}-{m:02d} sheet {rub(w):>16}  ours {rub(g):>16}  gap opened {rub(d):>14}"
            f"  {'sheet stale' if ok is False else ''}"
            for y, m, w, g, d, ok in bad_months
        ]

    text = "\n".join(lines)
    if args.report:
        pathlib.Path(args.report).write_text(text + "\n")
    head_text, *rest = text.split("\n\n", 1)
    print(head_text)
    if rest:
        body = rest[0].split("\n")
        print("\n".join(body[: args.limit]))
        if len(body) > args.limit:
            print(
                f"  … {len(body) - args.limit} more lines"
                + (f" in {args.report}" if args.report else "")
            )
    return 1 if bad_cells or bad_months or dupes or ledger.uncategorized else 0


if __name__ == "__main__":
    sys.exit(main())

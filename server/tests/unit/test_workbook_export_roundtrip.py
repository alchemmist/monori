import datetime
from dataclasses import dataclass
from io import BytesIO

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from monori.server.app.workbook.parser import (
    WorkbookError,
    _amount,
    _parse_dt,
    _unquote,
    parse_workbook,
)


def _workbook_datetime(*parts: int) -> datetime.datetime:
    values = list(parts) + [0] * (6 - len(parts))
    return datetime.datetime(
        values[0],
        values[1],
        values[2],
        values[3],
        values[4],
        values[5],
        tzinfo=datetime.UTC,
    ).replace(tzinfo=None)


def _active(wb: Workbook) -> Worksheet:
    ws = wb.active
    assert ws is not None
    return ws


def _workbook() -> tuple[Workbook, Worksheet, Worksheet]:
    wb = Workbook()
    ws = _active(wb)
    ws.title = "Categories"
    ws.append(["Sort Order", "Category Group", "Category", "Keywords"])
    tx = wb.create_sheet("Transactions")
    tx.append(
        [
            "Operation date",
            "Payment date",
            "Operation date",
            "Card",
            "Status",
            "Operation amount",
            "Transaction currency",
            "Amount",
            "Payment currency",
            "Cashback",
            "Category",
            "MCC",
            "Description",
            "Monori Category",
            "Account",
            "Comment",
        ],
    )
    return wb, ws, tx


def _bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_year_header(ws: Worksheet, months: int = 12) -> None:
    """Handle The two header rows the exporter writes above a year grid."""
    names = [
        "January",
        "February",
        "March",
        "April",
        "May",
        "June",
        "July",
        "August",
        "September",
        "October",
        "November",
        "December",
    ][:months]
    for index, name in enumerate(names):
        base = 2 + index * 3
        ws.cell(1, base, name)
        ws.cell(2, base, "Budgeted")
        ws.cell(2, base + 1, "Outflows")
        ws.cell(2, base + 2, "Balance")


@dataclass(frozen=True)
class TransactionFixture:
    op: str = "05.01.2026 10:00:00"
    status: str = "OK"
    amount: float = -125.5
    description: str = "Lenta"
    card: str = ""
    account: str = ""
    monori: str = ""
    comment: str = ""

    def append_to(self, ws: Worksheet) -> None:
        row = ws.max_row + 1
        ws.cell(row, 1, self.op)
        ws.cell(row, 2, "05.01.2026")
        ws.cell(row, 3, "05.01.2026")
        ws.cell(row, 4, self.card)
        ws.cell(row, 5, self.status)
        ws.cell(row, 6, self.amount)
        ws.cell(row, 7, "RUB")
        ws.cell(row, 8, "-125.50 ₽")
        ws.cell(row, 9, "RUB")
        ws.cell(row, 11, "Super")
        ws.cell(row, 12, "5411")
        ws.cell(row, 13, self.description)
        ws.cell(row, 14, self.monori)
        ws.cell(row, 15, self.account)
        ws.cell(row, 16, self.comment)


def test_unquote_reverses_only_the_formula_escape() -> None:
    assert _unquote("'=SUM(A1)") == "=SUM(A1)"
    assert _unquote("'+A1") == "+A1"
    assert _unquote("'@cmd") == "@cmd"
    assert _unquote("plain") == "plain"
    assert _unquote("mid'dle") == "mid'dle"
    assert _unquote("'legit apostrophe") == "'legit apostrophe"


def test_parse_dt_variants() -> None:
    ws = Workbook().active
    assert ws is not None
    assert _parse_dt(ws.cell(1, 1, _workbook_datetime(2026, 1, 5, 10))) == _workbook_datetime(
        2026,
        1,
        5,
        10,
    )
    assert _parse_dt(ws.cell(1, 2, datetime.date(2026, 1, 5))) == _workbook_datetime(2026, 1, 5)
    assert _parse_dt(ws.cell(1, 3, "05.01.2026 10:00:00")) == _workbook_datetime(2026, 1, 5, 10)
    assert _parse_dt(ws.cell(1, 4, "2026-01-05T10:00:00")) == _workbook_datetime(2026, 1, 5, 10)
    assert _parse_dt(ws.cell(1, 5, "")) is None
    assert _parse_dt(ws.cell(1, 6, "garbage")) is None
    assert _parse_dt(ws.cell(1, 7)) is None


def test_amount_variants() -> None:
    ws = Workbook().active
    assert ws is not None
    assert _amount(ws.cell(1, 1, -125.5)) == -12550
    assert _amount(ws.cell(1, 2, 500)) == 50000
    assert _amount(ws.cell(1, 3, "-1 500,00")) == -150000
    assert _amount(ws.cell(1, 4, "")) is None
    assert _amount(ws.cell(1, 5)) is None
    assert _amount(ws.cell(1, 6, "abc")) is None


def test_rejects_garbage_bytes() -> None:
    with pytest.raises(WorkbookError) as e:
        parse_workbook(b"nope")
    assert "not a readable .xlsx workbook" in str(e.value)


def test_transactions_are_the_only_required_sheet() -> None:
    """
    The category structure is read from a sheet of its own when there is one and.

    inferred from the year grids when there isn't, so only the rows themselves.
    are indispensable.
    """
    wb = Workbook()
    ws = _active(wb)
    ws.title = "Whatever"
    buf = BytesIO()
    wb.save(buf)
    with pytest.raises(WorkbookError) as e:
        parse_workbook(buf.getvalue())
    assert str(e.value) == "missing required sheet: Transactions"

    wb, _, tx = _workbook()
    TransactionFixture(monori="Groceries").append_to(tx)
    parsed = parse_workbook(_bytes(wb))
    assert [t.monori_category for t in parsed.transactions] == ["Groceries"]


def test_missing_transactions_sheet() -> None:
    wb = Workbook()
    ws = _active(wb)
    ws.title = "Categories"
    buf = BytesIO()
    wb.save(buf)
    with pytest.raises(WorkbookError) as e:
        parse_workbook(buf.getvalue())
    assert str(e.value) == "missing required sheet: Transactions"


def test_missing_required_transaction_columns() -> None:
    wb = Workbook()
    ws = _active(wb)
    ws.title = "Categories"
    tx = wb.create_sheet("Transactions")
    tx.append(["Operation date", "Status"])
    buf = BytesIO()
    wb.save(buf)
    with pytest.raises(WorkbookError) as e:
        parse_workbook(buf.getvalue())
    msg = str(e.value)
    assert msg.startswith("Transactions sheet is missing required columns:")
    assert "Operation amount" in msg


def test_categories_main_and_group_tables() -> None:
    wb, categories, _ = _workbook()
    categories.append([1, "▼Daily", "Groceries", "lenta|okey"])
    categories.append([2, "▲Inflow", "Salary", ""])
    categories.append([])
    categories.append(["Category Group", "Sort Order", "Type"])
    categories.append(["▼Daily", 1, "OUT"])
    categories.append(["▲Inflow", 2, "IN"])
    parsed = parse_workbook(_bytes(wb))
    assert [(group.name, group.sort, group.kind) for group in parsed.groups] == [
        ("Daily", 1, "expense"),
        ("Inflow", 2, "income"),
    ]
    cats = {c.name: c for c in parsed.categories}
    assert cats["Groceries"].group == "Daily"
    assert cats["Groceries"].keywords == "lenta|okey"
    assert cats["Salary"].group == "Inflow"
    assert cats["Salary"].group_kind == "income"
    assert parsed.warnings == []


def test_categories_unrecognized_row_warns() -> None:
    wb, categories, _ = _workbook()
    categories.append([1, "▼Daily", "Groceries", ""])
    categories.append(["junk", "row"])
    parsed = parse_workbook(_bytes(wb))
    assert any(w.startswith("Categories: unrecognized row skipped:") for w in parsed.warnings)


def test_category_sheet_saying_nothing_defers_to_the_grids() -> None:
    """
    The live spreadsheet has a sheet called Categories too, laid out nothing like.

    ours. Rather than report every row of it, the reader treats a sheet it cannot.
    read as absent and takes the structure from the year grids.
    """
    wb, categories, _ = _workbook()
    categories.append(["junk", "row"])
    categories.append(["more", "junk"])
    parsed = parse_workbook(_bytes(wb))
    assert parsed.categories == []
    assert parsed.groups == []
    assert parsed.warnings == [
        "Categories: no category rows recognized (2 rows skipped),"
        " structure taken from the year grids",
    ]


def test_groups_derived_when_group_table_missing() -> None:
    wb, categories, _ = _workbook()
    categories.append([3, "▼Daily", "Groceries", ""])
    categories.append([3, "▼Daily", "Cafes", ""])
    categories.append([5, "▲Inflow", "Salary", ""])
    parsed = parse_workbook(_bytes(wb))
    assert "Categories: group table missing, groups derived from category rows" in parsed.warnings
    assert [(group.name, group.sort, group.kind) for group in parsed.groups] == [
        ("Daily", 3, "expense"),
        ("Inflow", 5, "income"),
    ]


def test_transactions_parse_and_markers() -> None:
    wb, _, tx = _workbook()
    TransactionFixture(card="*2947", monori="Groceries", comment="note").append_to(tx)
    TransactionFixture(op="06.01.2026 09:30:00", account="Card", description="X").append_to(tx)
    tx.cell(tx.max_row, 6, "-1 500,00")
    TransactionFixture(status="FAILED").append_to(tx)
    tx.append([])
    parsed = parse_workbook(_bytes(wb))
    rows = parsed.transactions
    assert len(rows) == 2
    first, second = rows
    assert first.date == "2026-01-05T10:00:00"
    assert first.amount == -12550
    assert first.marker == "*2947"
    assert first.monori_category == "Groceries"
    assert first.comment == "note"
    assert first.bank_category == "Super"
    assert first.mcc == "5411"
    assert second.date == "2026-01-06T09:30:00"
    assert second.amount == -150000
    assert second.marker == "Card"
    assert parsed.warnings == ["Transactions: 1 non-OK rows skipped"]
    assert parsed.errors == []


def test_transaction_only_category_uses_supplied_group() -> None:
    wb, _, transactions = _workbook()
    transactions.insert_cols(14)
    transactions.cell(1, 14, "Monori Category Group")
    transactions.cell(1, 15, "Monori Category")
    transactions.cell(2, 1, "05.01.2026 10:00:00")
    transactions.cell(2, 5, "OK")
    transactions.cell(2, 6, -125.5)
    transactions.cell(2, 7, "RUB")
    transactions.cell(2, 13, "Rent payment")
    transactions.cell(2, 14, "Housing")
    transactions.cell(2, 15, "Rent")

    parsed = parse_workbook(_bytes(wb))

    assert [(category.group, category.name) for category in parsed.categories] == [
        ("Housing", "Rent")
    ]
    assert parsed.transactions[0].monori_category_group == "Housing"


def test_legacy_ambiguous_category_requires_group_identity() -> None:
    wb, categories, tx = _workbook()
    categories.append([1, "▼Home", "Other", ""])
    categories.append([2, "▼Travel", "Other", ""])
    TransactionFixture(monori="Other").append_to(tx)

    parsed = parse_workbook(_bytes(wb))

    assert [(category.group, category.name) for category in parsed.categories] == [
        ("Home", "Other"),
        ("Travel", "Other"),
    ]
    assert any("ambiguous legacy category 'Other'" in error.error for error in parsed.errors)


def test_transactions_unparseable_rows_reported_with_row_number() -> None:
    wb, _, tx = _workbook()
    TransactionFixture(op="garbage").append_to(tx)
    TransactionFixture().append_to(tx)
    tx.cell(tx.max_row, 6, "zzz")
    parsed = parse_workbook(_bytes(wb))
    assert parsed.transactions == []
    assert [(error.row, error.error) for error in parsed.errors] == [
        (2, "unparseable date or amount"),
        (3, "unparseable date or amount"),
    ]


def test_year_sheet_budgets_and_unknown_labels() -> None:
    wb, categories, _ = _workbook()
    categories.append([1, "▼Daily", "Groceries", ""])
    categories.append(["Category Group", "Sort Order", "Type"])
    categories.append(["▼Daily", 1, "OUT"])
    year = wb.create_sheet("2026")
    _write_year_header(year)
    year.append(["Month Summary", 200, 125.5, 74.5])
    year.append(["▼Daily"])
    year.append(["Groceries", 200, 125.5, 74.5, 300])
    year.append(["Mystery", 1, 2, 3])
    notes = wb.create_sheet("Notes")
    notes.append(["hi"])
    parsed = parse_workbook(_bytes(wb))
    cells = parsed.budgets
    assert {(c.category, c.year, c.month, c.amount) for c in cells} == {
        ("Groceries", 2026, 1, 20000),
        ("Groceries", 2026, 2, 30000),
    }
    assert "2026: unknown row label skipped: Mystery" in parsed.warnings
    assert "unknown sheet ignored: Notes" in parsed.warnings


def test_year_sheet_only_four_digit_names() -> None:
    wb, _, _ = _workbook()
    year = wb.create_sheet("20266")
    year.append(["Groceries", 1])
    parsed = parse_workbook(_bytes(wb))
    assert parsed.budgets == []
    assert "unknown sheet ignored: 20266" in parsed.warnings


def test_dashdata_sheet_is_known_and_silent() -> None:
    wb, _, _ = _workbook()
    dash = wb.create_sheet("DashData")
    dash.append(["Month", "Income"])
    parsed = parse_workbook(_bytes(wb))
    assert parsed.warnings == []

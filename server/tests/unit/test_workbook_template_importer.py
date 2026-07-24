import datetime
from io import BytesIO

import pytest
from openpyxl import Workbook

from app.workbook import spec
from app.workbook.importer import parse_workbook
from app.workbook.template_importer import (
    MONTH_TOKENS,
    RU_HEADERS,
    TemplateError,
    _find_layout,
    _kop,
    _last_day,
    _month_range,
    _month_token,
    _parse_dt,
    _parse_keywords,
    _parse_transactions,
    _parse_year_sheet,
    _s,
    _sheet_sections,
    _stamp,
    _summary_value,
    _synthetic,
    is_template_transactions_header,
    looks_like_template,
    parse_template_workbook,
)

TX_HEADER = list(RU_HEADERS.values())


def _write_year(
    ws,
    *,
    months,
    rows,
    header_row=5,
    start_token="ЯНВ 2025",
    income=None,
    available=None,
    seed=None,
):
    bases = [2 + 4 * i for i in range(len(months))]
    ws.cell(row=1, column=bases[0], value=start_token)
    ws.cell(row=header_row, column=1, value="Категория")
    for b in bases:
        ws.cell(row=header_row, column=b, value="Бюджет")
        ws.cell(row=header_row, column=b + 1, value="Расход")
        ws.cell(row=header_row, column=b + 2, value="Баланс")
    r = header_row + 1
    for label, vals in rows:
        ws.cell(row=r, column=1, value=label)
        if vals:
            for mi, mnum in enumerate(months):
                if mnum in vals:
                    b = bases[mi]
                    budget, outflow, balance = vals[mnum]
                    if budget is not None:
                        ws.cell(row=r, column=b, value=budget)
                    if outflow is not None:
                        ws.cell(row=r, column=b + 1, value=outflow)
                    if balance is not None:
                        ws.cell(row=r, column=b + 2, value=balance)
        r += 1
    if seed is not None:
        ws.cell(row=1, column=bases[0] + 2, value="Not budgeted in Dec")
        ws.cell(row=1, column=bases[0] + 1, value=seed)
    if income:
        for mi, mnum in enumerate(months):
            if mnum in income:
                b = bases[mi]
                ws.cell(row=2, column=b + 2, value="Income for month")
                ws.cell(row=2, column=b + 1, value=income[mnum])
    if available:
        for mi, mnum in enumerate(months):
            if mnum in available:
                b = bases[mi]
                ws.cell(row=6, column=b + 1, value="Available")
                ws.cell(row=5, column=b + 1, value=available[mnum])


def _tx_sheet(wb, tx_rows):
    ws = wb.create_sheet(spec.SHEET_TRANSACTIONS)
    ws.append(TX_HEADER)
    for row in tx_rows:
        ws.append(row)
    return ws


def _tx(date, amount, category, *, card="*1111", status="OK", currency="RUB", desc="", kw=None):
    row = [None] * 12
    row[0] = date
    row[1] = card
    row[2] = status
    row[3] = amount
    row[4] = currency
    row[5] = "Super"
    row[6] = "5411"
    row[7] = desc
    row[9] = category
    if kw:
        row[10], row[11] = kw
    return row


def _save(wb):
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- pure helpers ---------------------------------------------------------


def test_s_strips_and_handles_none():
    assert _s(None) == ""
    assert _s("  hi  ") == "hi"
    assert _s(5) == "5"


def test_kop_rejects_bool_and_non_numbers():
    assert _kop(True) is None
    assert _kop(False) is None
    assert _kop(None) is None
    assert _kop("abc") is None
    assert _kop("   ") is None
    assert _kop(12.5) == 1250
    assert _kop(-3) == -300


def test_kop_parses_formatted_strings():
    assert _kop("12") == 1200
    assert _kop("-4 172,00") == -417200
    assert _kop("1 234,5") == 123450
    assert _kop("2 000") == 200000


def test_last_day_handles_december_and_others():
    assert _last_day(2024, 1) == datetime.date(2024, 1, 31)
    assert _last_day(2024, 2) == datetime.date(2024, 2, 29)
    assert _last_day(2025, 2) == datetime.date(2025, 2, 28)
    assert _last_day(2024, 12) == datetime.date(2024, 12, 31)


def test_stamp_is_noon_on_last_day():
    assert _stamp(2025, 1) == "2025-01-31T12:00:00"
    assert _stamp(2024, 12) == "2024-12-31T12:00:00"


def test_month_token_matches_ru_and_en_and_rejects():
    assert _month_token("ЯНВ 2025") == 1
    assert _month_token("мая") == 5
    assert _month_token("DEC 2024") == 12
    assert _month_token("garbage") is None
    assert _month_token(None) is None
    assert MONTH_TOKENS["ИЮЛ"] == 7


def test_parse_dt_variants():
    assert _parse_dt(datetime.datetime(2025, 1, 5, 10)) == datetime.datetime(2025, 1, 5, 10)
    assert _parse_dt(datetime.date(2025, 1, 5)) == datetime.datetime(2025, 1, 5)
    assert _parse_dt("2025-01-05") is None
    assert _parse_dt(None) is None


def test_month_range_wraps_across_years():
    assert list(_month_range((2024, 11), (2025, 2))) == [
        (2024, 11),
        (2024, 12),
        (2025, 1),
        (2025, 2),
    ]
    assert list(_month_range((2025, 3), (2025, 1))) == []


def test_synthetic_shape_and_hash():
    a = _synthetic(2025, 1, 20000, "Groceries", "Migration adjustment: Groceries")
    assert a["date"] == "2025-01-31T12:00:00"
    assert a["amount"] == 20000
    assert a["monori_category"] == "Groceries"
    assert a["description"] == "Migration adjustment: Groceries"
    assert a["marker"] == ""
    assert a["bank_category"] == "" and a["mcc"] == "" and a["comment"] == ""
    assert a["hash"]


# --- detection ------------------------------------------------------------


def test_is_template_header_detects_russian_status():
    assert is_template_transactions_header(["Статус", "Сумма операции"]) is True
    assert is_template_transactions_header(["Status", "Статус"]) is False
    assert is_template_transactions_header(["Дата операции"]) is False


def test_looks_like_template_true_and_false():
    wb = Workbook()
    wb.remove(wb.active)
    _tx_sheet(wb, [])
    assert looks_like_template(_save(wb)) is True

    export = Workbook()
    export.active.title = spec.SHEET_TRANSACTIONS
    export.active.append(["Дата операции", "Status", "Сумма операции"])
    assert looks_like_template(_save(export)) is False

    assert looks_like_template(b"not a workbook") is False

    no_tx = Workbook()
    no_tx.active.title = "Something"
    assert looks_like_template(_save(no_tx)) is False


# --- layout & sections ----------------------------------------------------


def _one_year_wb(**kw):
    wb = Workbook()
    ws = wb.active
    ws.title = "2025"
    _write_year(ws, **kw)
    return wb, ws


def test_find_layout_reads_bases_offsets_label_and_month():
    _, ws = _one_year_wb(months=[1, 2], rows=[("▼Daily", None)])
    layout = _find_layout(ws)
    assert layout["header_row"] == 5
    assert layout["bases"] == [2, 6]
    assert layout["out_off"] == 1
    assert layout["bal_off"] == 2
    assert layout["label_col"] == 1
    assert layout["start_month"] == 1


def test_find_layout_none_without_two_budget_headers():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=5, column=2, value="Бюджет")
    ws.cell(row=5, column=3, value="Расход")
    assert _find_layout(ws) is None


def test_find_layout_none_without_outflow_or_balance():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=5, column=2, value="Бюджет")
    ws.cell(row=5, column=6, value="Бюджет")
    assert _find_layout(ws) is None


def test_find_layout_label_and_month_fallbacks():
    wb = Workbook()
    ws = wb.active
    for c, v in ((2, "Budgeted"), (3, "Outflows"), (4, "Balance"), (6, "Budgeted"), (8, "Balance")):
        ws.cell(row=5, column=c, value=v)
    ws.cell(row=5, column=7, value="Outflows")
    layout = _find_layout(ws)
    assert layout["label_col"] == 3
    assert layout["start_month"] == 1


def test_sheet_sections_glyph_groups_and_gap_rule():
    wb = Workbook()
    ws = wb.active
    layout = {"header_row": 1, "label_col": 1}
    ws.cell(row=2, column=1, value="▼Daily")
    ws.cell(row=3, column=1, value="Groceries")
    ws.cell(row=4, column=1, value="Cafes")
    ws.cell(row=5, column=1, value="Month Summary")
    ws.cell(row=7, column=1, value="Rent")
    ws.cell(row=8, column=1, value="▲Inflow")
    ws.cell(row=9, column=1, value="Salary")
    sections = _sheet_sections(ws, layout)
    daily = sections[0]
    assert daily["name"] == "Daily" and daily["kind"] == "expense"
    assert [name for _, name in daily["rows"]] == ["Groceries", "Cafes"]
    gap_group = sections[1]
    assert gap_group["name"] == "Rent" and gap_group["rows"] == []
    inflow = sections[2]
    assert inflow["name"] == "Inflow" and inflow["kind"] == "income"
    assert [name for _, name in inflow["rows"]] == ["Salary"]


def test_summary_value_matches_prefix_only():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=3, column=4, value="Income for January")
    ws.cell(row=3, column=3, value=6000)
    assert _summary_value(ws, 2, ("Income for", "Поступления в")) == 600000
    assert _summary_value(ws, 2, ("Nope",)) is None


def test_parse_year_sheet_reads_grid_income_available_seed():
    wb = Workbook()
    ws = wb.active
    _write_year(
        ws,
        header_row=8,
        months=[1, 2],
        rows=[
            ("▼Daily", None),
            ("Groceries", {1: (1000, 300, 900), 2: (1000, 500, 1400)}),
        ],
        income={1: 6000},
        available={1: 5100, 2: 4000},
        seed=1234,
    )
    layout = _find_layout(ws)
    parsed = _parse_year_sheet(ws, 2025, layout)
    groceries = parsed["cats"]["Groceries"]
    assert groceries["group"] == "Daily"
    assert groceries["budgets"] == {1: 100000, 2: 100000}
    assert groceries["outflows"] == {1: 30000, 2: 50000}
    assert groceries["balances"] == {1: 90000, 2: 140000}
    assert parsed["income"] == {1: 600000}
    assert parsed["available"] == {1: 510000, 2: 400000}
    assert parsed["seed"] == 123400
    assert parsed["months"] == [1, 2]


# --- transactions & keywords ---------------------------------------------


def _tx_only_ws(tx_rows):
    wb = Workbook()
    wb.remove(wb.active)
    return _tx_sheet(wb, tx_rows)


def test_parse_transactions_empty_and_missing_columns():
    from openpyxl import load_workbook

    empty = Workbook()
    empty.remove(empty.active)
    empty.create_sheet(spec.SHEET_TRANSACTIONS)
    ro = load_workbook(BytesIO(_save(empty)), read_only=True, data_only=True)
    with pytest.raises(TemplateError, match="Transactions sheet is empty"):
        _parse_transactions(ro[spec.SHEET_TRANSACTIONS], [], [])

    bad = Workbook()
    ws = bad.create_sheet(spec.SHEET_TRANSACTIONS)
    ws.append(["Дата операции", "Статус"])
    with pytest.raises(TemplateError, match="missing required columns"):
        _parse_transactions(ws, [], [])


def test_parse_transactions_dedup_status_currency_and_category():
    d = datetime.datetime(2025, 1, 15, 10)
    ws = _tx_only_ws(
        [
            _tx(d, -300.0, "Groceries", desc="Lenta"),
            _tx(d, -300.0, "Groceries", desc="Lenta"),  # duplicate row
            _tx(datetime.datetime(2025, 1, 16), -50.0, "Cafes", status="FAILED"),
            _tx(datetime.datetime(2025, 1, 17), -20.0, "Travel", currency="USD", desc="Abroad"),
            _tx("bad", -1.0, "X"),  # unparseable date
            [None] * 12,  # blank
        ]
    )
    warnings, errors = [], []
    rows = _parse_transactions(ws, warnings, errors)
    assert [r["monori_category"] for r in rows] == ["Groceries", "Travel"]
    first = rows[0]
    assert first["date"] == "2025-01-15T10:00:00"
    assert first["amount"] == -30000
    assert first["marker"] == "*1111"
    assert first["bank_category"] == "Super"
    assert first["mcc"] == "5411"
    assert errors == [{"row": 6, "error": "unparseable date or amount"}]
    assert "Transactions: 1 duplicated rows collapsed" in warnings
    assert "Transactions: 1 non-OK rows skipped" in warnings
    assert "Transactions: 1 non-RUB rows imported with their face value" in warnings


def test_parse_transactions_pay_amount_fallback_and_blankish_rows():
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(spec.SHEET_TRANSACTIONS)
    ws.append(TX_HEADER + ["Сумма платежа", "Валюта платежа"])

    manual = [None] * 13
    manual[0] = datetime.datetime(2025, 3, 1, 9)
    manual[2] = "OK"
    manual[4] = "RUB"
    manual[7] = "Salary"
    manual[8] = 47337.0
    manual[11] = "Income"
    ws.append(manual)

    formatted = [None] * 13
    formatted[0] = datetime.datetime(2025, 3, 2)
    formatted[2] = "OK"
    formatted[7] = "Points"
    formatted[8] = "-4 172,00"
    formatted[9] = "RUB"
    formatted[11] = "Groceries"
    ws.append(formatted)

    foreign = [None] * 13
    foreign[0] = datetime.datetime(2025, 3, 3)
    foreign[2] = "OK"
    foreign[7] = "Abroad"
    foreign[8] = 10.0
    foreign[9] = "USD"
    foreign[11] = "Travel"
    ws.append(foreign)

    blankish = [None] * 13
    blankish[11] = 0
    ws.append(blankish)

    dated_only = [None] * 13
    dated_only[0] = datetime.datetime(2025, 3, 4)
    dated_only[7] = "no amount anywhere"
    ws.append(dated_only)

    warnings, errors = [], []
    rows = _parse_transactions(ws, warnings, errors)
    assert [r["amount"] for r in rows] == [4733700, -417200, 1000]
    assert rows[0]["monori_category"] == "Income"
    assert rows[1]["monori_category"] == "Groceries"
    assert errors == [{"row": 6, "error": "unparseable date or amount"}]
    assert "Transactions: 1 non-RUB rows imported with their face value" in warnings


def test_parse_keywords_reads_side_table():
    ws = _tx_only_ws(
        [
            _tx(datetime.datetime(2025, 1, 1), -10.0, "Groceries", kw=("Groceries", "lenta|okey")),
            _tx(datetime.datetime(2025, 1, 2), -10.0, "Cafes", kw=("Cafes", "starbucks")),
            _tx(datetime.datetime(2025, 1, 3), -10.0, "X", kw=("Skip", "a")),
        ]
    )
    idx = {name: i for i, name in enumerate(TX_HEADER)}
    kws = _parse_keywords(ws, idx)
    assert kws["Groceries"] == "lenta|okey"
    assert kws["Cafes"] == "starbucks"
    assert "Skip" not in kws  # single-char keyword rejected


# --- end to end -----------------------------------------------------------


def _live_year_wb():
    wb = Workbook()
    wb.remove(wb.active)
    _tx_sheet(
        wb,
        [
            _tx(datetime.datetime(2025, 1, 15), -300.0, "Groceries", desc="Lenta"),
            _tx(datetime.datetime(2025, 2, 10), -500.0, "Groceries", desc="Okey"),
            _tx(datetime.datetime(2025, 1, 5), 5000.0, "Income", card="*2222", desc="Payroll"),
            _tx(datetime.datetime(2025, 1, 20), -100.0, "", desc="Uncategorized"),
        ],
    )
    ws = wb.create_sheet("2025")
    _write_year(
        ws,
        header_row=8,
        months=[1, 2],
        start_token="ЯНВ 2025",
        rows=[
            ("▼Daily", None),
            ("Groceries", {1: (1000, 300, 900), 2: (1000, 500, 1400)}),
            ("▲Inflow", None),
            ("Salary", None),
        ],
        income={1: 6000},
        available={1: 5100, 2: 4000},
    )
    junk = wb.create_sheet("2019")
    junk.cell(row=1, column=1, value="not a year grid")
    return wb


def test_live_year_reconciles_to_cached_balances_and_available():
    parsed = parse_template_workbook(_save(_live_year_wb()))

    assert parsed["groups"] == [
        {"name": "Inflow", "sort": 0, "kind": "income"},
        {"name": "Daily", "sort": 1, "kind": "expense"},
    ]
    assert [c["name"] for c in parsed["categories"]] == ["Income", "Groceries", "Salary"]

    budgets = {(b["category"], b["year"], b["month"]): b["amount"] for b in parsed["budgets"]}
    assert budgets == {("Groceries", 2025, 1): 100000, ("Groceries", 2025, 2): 100000}

    synth = {t["description"]: t for t in parsed["transactions"] if "Migration" in t["description"]}
    assert synth["Migration adjustment: Groceries"]["amount"] == 20000
    assert synth["Migration adjustment: Groceries"]["date"] == "2025-01-31T12:00:00"
    assert synth["Migration adjustment: income"]["amount"] == 100000
    assert len(parsed["transactions"]) == 6  # 4 real (one uncategorized) + 2 synthetic

    assert (
        "reconciliation: 2 adjustment transactions align live months with the sheet"
        in parsed["warnings"]
    )
    assert "verify: available 2025-01 differs by 100.00" in parsed["warnings"]
    assert "2019: unrecognized year sheet layout, ignored" in parsed["warnings"]
    assert parsed["errors"] == []


def test_parse_workbook_dispatches_to_template_parser():
    parsed = parse_workbook(_save(_live_year_wb()))
    assert [c["name"] for c in parsed["categories"]] == ["Income", "Groceries", "Salary"]


def test_archive_history_and_seam_carry():
    wb = Workbook()
    wb.remove(wb.active)
    _tx_sheet(wb, [])

    _write_year(
        wb.create_sheet("2024_archive"),
        months=[1, 2],
        start_token="ЯНВ 2024",
        rows=[("▼Daily", None), ("Groceries", {1: (None, None, 500)})],
        income={1: 100},
    )
    _write_year(
        wb.create_sheet("2024"),
        months=[1, 2],
        start_token="ЯНВ 2024",
        rows=[("▼Daily", None), ("Groceries", {2: (None, None, 800)})],
    )
    _write_year(
        wb.create_sheet("2025"),
        months=[1, 2],
        start_token="ЯНВ 2025",
        rows=[("▼Daily", None), ("Groceries", {1: (None, None, 800)})],
    )

    parsed = parse_template_workbook(_save(wb))
    synth = {t["description"]: t for t in parsed["transactions"]}
    assert synth["Migration history: income"]["amount"] == 10000
    assert synth["Migration history: Groceries"]["amount"] == 50000
    assert synth["Migration history: Groceries"]["date"] == "2024-01-31T12:00:00"
    assert synth["Migration carry: Groceries"]["amount"] == 30000
    assert synth["Migration carry: Groceries"]["date"] == "2024-12-31T12:00:00"
    assert len(parsed["transactions"]) == 3

    assert "history: 2 synthetic transactions rebuilt from archive sheets" in parsed["warnings"]
    assert "seam: 1 carry corrections at 2024-12" in parsed["warnings"]


def test_outflow_fallback_when_balance_cell_missing():
    wb = Workbook()
    wb.remove(wb.active)
    _tx_sheet(wb, [])
    _write_year(
        wb.create_sheet("2025"),
        months=[1, 2],
        start_token="ЯНВ 2025",
        rows=[("▼Daily", None), ("Groceries", {1: (None, None, 100), 2: (None, 200, None)})],
    )
    parsed = parse_template_workbook(_save(wb))
    synth = {t["description"]: t for t in parsed["transactions"]}
    # Jan aligns balance to 100.00; Feb has no balance cell, so the outflow drives
    # the target: projected(100) - have(0) + outflow(200) = 300 -> +200.00 delta.
    assert synth["Migration adjustment: Groceries"]["amount"] == 20000


def test_dead_category_and_available_seed_at_seam():
    wb = Workbook()
    wb.remove(wb.active)
    _tx_sheet(wb, [])
    _write_year(
        wb.create_sheet("2024_archive"),
        months=[1, 2],
        start_token="ЯНВ 2024",
        rows=[
            ("▼Daily", None),
            ("Groceries", {1: (None, None, 500)}),
            ("OldPhone", {1: (None, None, 300)}),
        ],
    )
    _write_year(
        wb.create_sheet("2024"),
        months=[1, 2],
        start_token="ЯНВ 2024",
        rows=[("▼Daily", None), ("Groceries", {2: (None, None, 800)})],
    )
    _write_year(
        wb.create_sheet("2025"),
        months=[1, 2],
        start_token="ЯНВ 2025",
        rows=[("▼Daily", None), ("Groceries", {1: (None, None, 800)})],
        seed=200,
    )
    parsed = parse_template_workbook(_save(wb))
    synth = {t["description"]: t for t in parsed["transactions"]}
    assert synth["Migration history: OldPhone"]["amount"] == 30000
    assert synth["Migration carry: OldPhone"]["amount"] == -30000  # dead category zeroed
    assert synth["Migration: available seed"]["amount"] == 20000
    assert synth["Migration: available seed"]["date"] == "2024-12-31T12:00:00"
    assert "history: 2 synthetic transactions rebuilt from archive sheets" in parsed["warnings"]
    assert "seam: 3 carry corrections at 2024-12" in parsed["warnings"]


def test_available_seed_excludes_seam_overspend():
    """
    The template's "Not budgeted in Dec" seed is the December available BEFORE
    overspend; the sheet adds "Overspent in Dec" separately in January. The
    seed correction must therefore target avail alone, not avail + overspent.
    """
    wb = Workbook()
    wb.remove(wb.active)
    _tx_sheet(wb, [])
    _write_year(
        wb.create_sheet("2024_archive"),
        months=[1, 2],
        start_token="ЯНВ 2024",
        rows=[("▼Daily", None), ("Groceries", {1: (None, None, 500)})],
        income={1: 100},
        header_row=8,
    )
    _write_year(
        wb.create_sheet("2024"),
        months=[1, 2],
        start_token="ЯНВ 2024",
        rows=[("▼Daily", None), ("Groceries", {2: (None, None, -100)})],
    )
    _write_year(
        wb.create_sheet("2025"),
        months=[1, 2],
        start_token="ЯНВ 2025",
        rows=[("▼Daily", None), ("Groceries", None)],
        seed=200,
        available={1: 100},
        header_row=8,
    )
    parsed = parse_template_workbook(_save(wb))
    synth = {t["description"]: t for t in parsed["transactions"]}
    assert synth["Migration carry: Groceries"]["amount"] == -60000
    assert synth["Migration: available seed"]["amount"] == 10000
    assert not any(w.startswith("verify:") for w in parsed["warnings"])


def test_no_live_year_sheets_raises():
    wb = Workbook()
    wb.remove(wb.active)
    _tx_sheet(wb, [])
    _write_year(
        wb.create_sheet("2024_archive"),
        months=[1, 2],
        start_token="ЯНВ 2024",
        rows=[("▼Daily", None), ("Groceries", {1: (None, None, 500)})],
    )
    with pytest.raises(TemplateError, match="no live year sheets found"):
        parse_template_workbook(_save(wb))


def test_missing_transactions_sheet_raises():
    wb = Workbook()
    wb.active.title = "2025"
    with pytest.raises(TemplateError, match="missing required sheet: Transactions"):
        parse_template_workbook(_save(wb))


def test_parse_template_rejects_garbage_bytes():
    with pytest.raises(TemplateError, match="not a readable .xlsx workbook"):
        parse_template_workbook(b"nope")

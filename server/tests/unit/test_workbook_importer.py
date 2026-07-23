import datetime
from io import BytesIO

import pytest
from openpyxl import Workbook

from app.workbook.importer import (
    WorkbookError,
    _parse_amount_cell,
    _parse_dt_cell,
    _unquote,
    parse_workbook,
)


def _book(categories=None, transactions=None, extra_sheets=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Categories"
    ws.append(["Sort Order", "Category Group", "Category", "Keywords"])
    for row in categories or []:
        ws.append(row)
    tx = wb.create_sheet("Transactions")
    tx.append(
        [
            "Дата операции",
            "Date",
            "Дата платежа",
            "Номер карты",
            "Status",
            "Сумма операции",
            "Валюта операции",
            "Amount",
            "Валюта платежа",
            "Кэшбэк",
            "Категория",
            "MCC",
            "Description",
            "Monori Category",
            "Account",
            "Comment",
        ]
    )
    for row in transactions or []:
        tx.append(row)
    for name, rows in (extra_sheets or {}).items():
        s = wb.create_sheet(name)
        for row in rows:
            s.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _tx_row(
    op="05.01.2026 10:00:00",
    status="OK",
    amount=-125.5,
    description="Lenta",
    card="",
    account="",
    monori="",
    comment="",
):
    return [
        op,
        "05.01.2026",
        "05.01.2026",
        card,
        status,
        amount,
        "RUB",
        "-125.50 ₽",
        "RUB",
        "",
        "Super",
        "5411",
        description,
        monori,
        account,
        comment,
    ]


def test_unquote_strips_only_leading_apostrophe():
    assert _unquote("'=SUM(A1)") == "=SUM(A1)"
    assert _unquote("plain") == "plain"
    assert _unquote("mid'dle") == "mid'dle"


def test_parse_dt_cell_variants():
    assert _parse_dt_cell(datetime.datetime(2026, 1, 5, 10, 0)) == datetime.datetime(
        2026, 1, 5, 10, 0
    )
    assert _parse_dt_cell(datetime.date(2026, 1, 5)) == datetime.datetime(2026, 1, 5)
    assert _parse_dt_cell("05.01.2026 10:00:00") == datetime.datetime(2026, 1, 5, 10)
    assert _parse_dt_cell("2026-01-05T10:00:00") == datetime.datetime(2026, 1, 5, 10)
    assert _parse_dt_cell("") is None
    assert _parse_dt_cell("garbage") is None
    assert _parse_dt_cell(None) is None


def test_parse_amount_cell_variants():
    assert _parse_amount_cell(-125.5) == -12550
    assert _parse_amount_cell(500) == 50000
    assert _parse_amount_cell("-1 500,00") == -150000
    assert _parse_amount_cell("") is None
    assert _parse_amount_cell(None) is None
    assert _parse_amount_cell("abc") is None


def test_rejects_garbage_bytes():
    with pytest.raises(WorkbookError) as e:
        parse_workbook(b"nope")
    assert "not a readable .xlsx workbook" in str(e.value)


def test_missing_sheets_are_reported():
    wb = Workbook()
    wb.active.title = "Whatever"
    buf = BytesIO()
    wb.save(buf)
    with pytest.raises(WorkbookError) as e:
        parse_workbook(buf.getvalue())
    assert str(e.value) == "missing required sheet: Categories"


def test_missing_transactions_sheet():
    wb = Workbook()
    wb.active.title = "Categories"
    buf = BytesIO()
    wb.save(buf)
    with pytest.raises(WorkbookError) as e:
        parse_workbook(buf.getvalue())
    assert str(e.value) == "missing required sheet: Transactions"


def test_missing_required_transaction_columns():
    wb = Workbook()
    wb.active.title = "Categories"
    tx = wb.create_sheet("Transactions")
    tx.append(["Дата операции", "Status"])
    buf = BytesIO()
    wb.save(buf)
    with pytest.raises(WorkbookError) as e:
        parse_workbook(buf.getvalue())
    msg = str(e.value)
    assert msg.startswith("Transactions sheet is missing required columns:")
    assert "Сумма операции" in msg and "Description" in msg


def test_categories_main_and_group_tables():
    data = _book(
        categories=[
            [1, "▼Daily", "Groceries", "lenta|okey"],
            [2, "▲Inflow", "Salary", ""],
            [],
            ["Category Group", "Sort Order", "Type"],
            ["▼Daily", 1, "OUT"],
            ["▲Inflow", 2, "IN"],
        ]
    )
    parsed = parse_workbook(data)
    assert parsed["groups"] == [
        {"name": "Daily", "sort": 1, "kind": "expense"},
        {"name": "Inflow", "sort": 2, "kind": "income"},
    ]
    cats = {c["name"]: c for c in parsed["categories"]}
    assert cats["Groceries"]["group"] == "Daily"
    assert cats["Groceries"]["keywords"] == "lenta|okey"
    assert cats["Salary"]["group"] == "Inflow"
    assert cats["Salary"]["group_kind"] == "income"
    assert parsed["warnings"] == []


def test_categories_unrecognized_row_warns():
    data = _book(categories=[["junk", "row", None]])
    parsed = parse_workbook(data)
    assert any(w.startswith("Categories: unrecognized row skipped:") for w in parsed["warnings"])


def test_groups_derived_when_group_table_missing():
    data = _book(
        categories=[
            [3, "▼Daily", "Groceries", ""],
            [3, "▼Daily", "Cafes", ""],
            [5, "▲Inflow", "Salary", ""],
        ]
    )
    parsed = parse_workbook(data)
    assert (
        "Categories: group table missing, groups derived from category rows" in (parsed["warnings"])
    )
    assert parsed["groups"] == [
        {"name": "Daily", "sort": 3, "kind": "expense"},
        {"name": "Inflow", "sort": 5, "kind": "income"},
    ]


def test_transactions_parse_and_markers():
    data = _book(
        transactions=[
            _tx_row(card="*2947", monori="Groceries", comment="note"),
            _tx_row(op="06.01.2026 09:30:00", amount="-1 500,00", account="Card", description="X"),
            _tx_row(status="FAILED"),
            [None] * 16,
        ]
    )
    parsed = parse_workbook(data)
    rows = parsed["transactions"]
    assert len(rows) == 2
    first, second = rows
    assert first["date"] == "2026-01-05T10:00:00"
    assert first["amount"] == -12550
    assert first["marker"] == "*2947"
    assert first["monori_category"] == "Groceries"
    assert first["comment"] == "note"
    assert first["bank_category"] == "Super"
    assert first["mcc"] == "5411"
    assert second["date"] == "2026-01-06T09:30:00"
    assert second["amount"] == -150000
    assert second["marker"] == "Card"
    assert parsed["warnings"] == ["Transactions: 1 non-OK rows skipped"]
    assert parsed["errors"] == []


def test_transactions_unparseable_rows_reported_with_row_number():
    data = _book(transactions=[_tx_row(op="garbage"), _tx_row(amount="zzz")])
    parsed = parse_workbook(data)
    assert parsed["transactions"] == []
    assert parsed["errors"] == [
        {"row": 2, "error": "unparseable date or amount"},
        {"row": 3, "error": "unparseable date or amount"},
    ]


def test_year_sheet_budgets_and_unknown_labels():
    year_rows = [
        ["Month Summary", 200, 125.5, 74.5],
        ["▼Daily", None, None, None],
        ["Groceries", 200, 125.5, 74.5, 300],
        ["Mystery", 1, 2, 3],
    ]
    data = _book(
        categories=[
            [1, "▼Daily", "Groceries", ""],
            ["Category Group", "Sort Order", "Type"],
            ["▼Daily", 1, "OUT"],
        ],
        extra_sheets={"2026": [[], []] + year_rows, "Notes": [["hi"]]},
    )
    parsed = parse_workbook(data)
    cells = parsed["budgets"]
    assert {(c["category"], c["year"], c["month"], c["amount"]) for c in cells} == {
        ("Groceries", 2026, 1, 20000),
        ("Groceries", 2026, 2, 30000),
    }
    assert "2026: unknown row label skipped: Mystery" in parsed["warnings"]
    assert "unknown sheet ignored: Notes" in parsed["warnings"]


def test_year_sheet_only_four_digit_names():
    data = _book(extra_sheets={"20266": [["Groceries", 1]]})
    parsed = parse_workbook(data)
    assert parsed["budgets"] == []
    assert "unknown sheet ignored: 20266" in parsed["warnings"]


def test_dashdata_sheet_is_known_and_silent():
    data = _book(extra_sheets={"DashData": [["Month", "Income"]]})
    parsed = parse_workbook(data)
    assert parsed["warnings"] == []

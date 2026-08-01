import datetime
from dataclasses import dataclass
from io import BytesIO
from typing import TYPE_CHECKING

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.workbook import spec
from app.workbook.parser import (
    MONTH_ABBREVS,
    TX_ALIASES,
    LayoutRow,
    WorkbookError,
    _find_layout,
    _kop,
    _label_col,
    _last_day,
    _month_num,
    _month_range,
    _parse_dt,
    _parse_keywords,
    _parse_transactions,
    _parse_year_sheet,
    _s,
    _sheet_sections,
    _stamp,
    _summary_value,
    _synthetic,
    _tx_header_index,
    parse_workbook,
)

if TYPE_CHECKING:
    from app.workbook.models import WorkbookParseError as WorkbookParseErrorRow

TX_HEADER: list[str] = [
    TX_ALIASES[f][0]
    for f in ("date", "card", "status", "amount", "currency", "bank_category", "mcc", "description")
]


def _workbook_datetime(*parts: int) -> datetime.datetime:
    return datetime.datetime(*parts, tzinfo=datetime.UTC).replace(tzinfo=None)


def _active(wb: Workbook) -> Worksheet:
    ws = wb.active
    assert ws is not None
    return ws


def _new_wb() -> Workbook:
    wb = Workbook()
    active = _active(wb)
    wb.remove(active)
    return wb


def _drop_active_sheet(wb: Workbook) -> None:
    active = _active(wb)
    wb.remove(active)


def _write_year_rows(
    ws: Worksheet,
    months: list[int],
    rows: list[tuple[str, dict[int, tuple[int | None, int | None, int | None]] | None]],
    bases: list[int],
    first_row: int,
) -> None:
    for row, (label, values) in enumerate(rows, start=first_row):
        ws.cell(row=row, column=1, value=label)
        if not values:
            continue
        for month_index, month in enumerate(months):
            if month not in values:
                continue
            base = bases[month_index]
            budget, outflow, balance = values[month]
            for column, value in enumerate((budget, outflow, balance), start=base):
                if value is not None:
                    ws.cell(row=row, column=column, value=value)


@dataclass(frozen=True)
class YearOptions:
    header_row: int = 5
    start_token: str = "JAN 2025"
    income: dict[int, int] | None = None
    available: dict[int, int] | None = None
    seed: int | None = None
    seeds: dict[int, int] | None = None
    labels: str = "en"
    available_row: int = 6


def _write_year(
    ws: Worksheet,
    *,
    months: list[int],
    rows: list[tuple[str, dict[int, tuple[int | None, int | None, int | None]] | None]],
    options: YearOptions | None = None,
) -> None:
    if options is None:
        options = YearOptions()
    carried_label, income_label, available_label = {
        "en": ("Not budgeted in Dec", "Income for month", "Available"),
        "ru": ("Not budgeted in Dec", "Income for month", "Available"),
    }.get(
        options.labels,
        ("Not budgeted in Dec", "Income for month", "Available"),
    )
    bases = [2 + 4 * i for i in range(len(months))]
    ws.cell(row=1, column=bases[0], value=options.start_token)
    ws.cell(row=options.header_row, column=1, value="Category")
    for b in bases:
        ws.cell(row=options.header_row, column=b, value="Budgeted")
        ws.cell(row=options.header_row, column=b + 1, value="Outflows")
        ws.cell(row=options.header_row, column=b + 2, value="Balance")
    _write_year_rows(ws, months, rows, bases, options.header_row + 1)
    seed_values = {
        **({} if options.seed is None else {months[0]: options.seed}),
        **(options.seeds or {}),
    }
    for mnum, value in seed_values.items():
        b = bases[months.index(mnum)]
        ws.cell(row=1, column=b + 2, value=carried_label)
        ws.cell(row=1, column=b + 1, value=value)
    if options.income:
        for mi, mnum in enumerate(months):
            if mnum in options.income:
                b = bases[mi]
                ws.cell(row=2, column=b + 2, value=income_label)
                ws.cell(row=2, column=b + 1, value=options.income[mnum])
    if options.available:
        for mi, mnum in enumerate(months):
            if mnum in options.available:
                b = bases[mi]
                ws.cell(row=options.available_row, column=b + 1, value=available_label)
                ws.cell(
                    row=options.available_row - 1,
                    column=b + 1,
                    value=options.available[mnum],
                )


@dataclass(frozen=True)
class TransactionFixture:
    date: datetime.datetime | str
    amount: float
    category: str
    card: str = "*1111"
    status: str = "OK"
    currency: str = "RUB"
    description: str = ""
    keyword: tuple[str, str] | None = None

    def append_to(self, ws: Worksheet) -> None:
        row = ws.max_row + 1
        ws.cell(row, 1, self.date)
        ws.cell(row, 2, self.card)
        ws.cell(row, 3, self.status)
        ws.cell(row, 4, self.amount)
        ws.cell(row, 5, self.currency)
        ws.cell(row, 6, "Super")
        ws.cell(row, 7, "5411")
        ws.cell(row, 8, self.description)
        ws.cell(row, 10, self.category)
        if self.keyword is not None:
            ws.cell(row, 11, self.keyword[0])
            ws.cell(row, 12, self.keyword[1])


def _tx_sheet(wb: Workbook, tx_rows: list[TransactionFixture | None]) -> Worksheet:
    ws = wb.create_sheet(spec.SHEET_TRANSACTIONS)
    assert isinstance(ws, Worksheet)
    ws.append(TX_HEADER)
    for row in tx_rows:
        if row is None:
            ws.append([])
        else:
            row.append_to(ws)
    return ws


def _save(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_s_strips_and_handles_none() -> None:
    ws = Workbook().active
    assert ws is not None
    assert _s(ws.cell(1, 1)) == ""
    assert _s(ws.cell(1, 2, "  hi  ")) == "hi"
    assert _s(ws.cell(1, 3, 5)) == "5"


def test_kop_rejects_bool_and_non_numbers() -> None:
    ws = Workbook().active
    assert ws is not None
    assert _kop(ws.cell(row=1, column=1, value=True)) is None
    assert _kop(ws.cell(row=1, column=2, value=False)) is None
    assert _kop(ws.cell(1, 3)) is None
    assert _kop(ws.cell(1, 4, "abc")) is None
    assert _kop(ws.cell(1, 5, "   ")) is None
    assert _kop(ws.cell(1, 6, 12.5)) == 1250
    assert _kop(ws.cell(1, 7, -3)) == -300


def test_kop_parses_formatted_strings() -> None:
    ws = Workbook().active
    assert ws is not None
    assert _kop(ws.cell(1, 1, "12")) == 1200
    assert _kop(ws.cell(1, 2, "-4 172,00")) == -417200
    assert _kop(ws.cell(1, 3, "1 234,5")) == 123450
    assert _kop(ws.cell(1, 4, "2 000")) == 200000


def test_last_day_handles_december_and_others() -> None:
    assert _last_day(2024, 1) == datetime.date(2024, 1, 31)
    assert _last_day(2024, 2) == datetime.date(2024, 2, 29)
    assert _last_day(2025, 2) == datetime.date(2025, 2, 28)
    assert _last_day(2024, 12) == datetime.date(2024, 12, 31)


def test_stamp_is_noon_on_last_day() -> None:
    assert _stamp(2025, 1) == "2025-01-31T12:00:00"
    assert _stamp(2024, 12) == "2024-12-31T12:00:00"


def test_month_num_matches_ru_and_en_and_rejects() -> None:
    ws = Workbook().active
    assert ws is not None
    assert _month_num(ws.cell(1, 1, "JAN 2025")) == 1
    assert _month_num(ws.cell(1, 2, "MAY")) == 5
    assert _month_num(ws.cell(1, 3, "DEC 2024")) == 12
    assert _month_num(ws.cell(1, 4, "garbage")) is None
    assert _month_num(ws.cell(1, 5)) is None
    assert MONTH_ABBREVS["JUL"] == 7


def test_parse_dt_variants() -> None:
    ws = Workbook().active
    assert ws is not None
    assert _parse_dt(ws.cell(1, 1, _workbook_datetime(2025, 1, 5, 10))) == _workbook_datetime(
        2025,
        1,
        5,
        10,
    )
    assert _parse_dt(ws.cell(1, 2, datetime.date(2025, 1, 5))) == _workbook_datetime(2025, 1, 5)
    assert _parse_dt(ws.cell(1, 3, "2025-01-05")) == _workbook_datetime(2025, 1, 5)
    assert _parse_dt(ws.cell(1, 4, "05.01.2025 10:00:00")) == _workbook_datetime(2025, 1, 5, 10)
    assert _parse_dt(ws.cell(1, 5, "garbage")) is None
    assert _parse_dt(ws.cell(1, 6)) is None


def test_month_range_wraps_across_years() -> None:
    assert list(_month_range((2024, 11), (2025, 2))) == [
        (2024, 11),
        (2024, 12),
        (2025, 1),
        (2025, 2),
    ]
    assert list(_month_range((2025, 3), (2025, 1))) == []


def test_synthetic_shape() -> None:
    a = _synthetic(2025, 1, 20000, "Groceries", "Groceries")
    assert a.date == "2025-01-31T12:00:00"
    assert a.amount == 20000
    assert a.monori_category == "Groceries"
    assert a.description == "Groceries"
    assert a.marker == ""
    assert a.bank_category == ""
    assert a.mcc == ""
    assert a.comment == ""
    assert not hasattr(a, "hash")


def _one_year_wb(
    *,
    months: list[int],
    rows: list[tuple[str, dict[int, tuple[int | None, int | None, int | None]] | None]],
) -> tuple[Workbook, Worksheet]:
    wb = _new_wb()
    ws = wb.create_sheet("2025")
    ws.title = "2025"
    _write_year(ws, months=months, rows=rows)
    return wb, ws


def test_find_layout_reads_bases_offsets_label_and_month() -> None:
    _, ws = _one_year_wb(months=[1, 2], rows=[("▼Daily", None)])
    layout = _find_layout(ws)
    assert layout is not None
    assert layout.header_row == 5
    assert layout.bases == [2, 6]
    assert layout.out_off == 1
    assert layout.bal_off == 2
    assert layout.label_col == 1
    assert layout.start_month == 1


def test_find_layout_none_without_two_budget_headers() -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.cell(row=5, column=2, value="Budgeted")
    ws.cell(row=5, column=3, value="Outflows")
    assert _find_layout(ws) is None


def test_find_layout_none_without_outflow_or_balance() -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.cell(row=5, column=2, value="Budgeted")
    ws.cell(row=5, column=6, value="Budgeted")
    assert _find_layout(ws) is None


def test_find_layout_label_and_month_fallbacks() -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    for c, v in ((2, "Budgeted"), (3, "Outflows"), (4, "Balance"), (6, "Budgeted"), (8, "Balance")):
        ws.cell(row=5, column=c, value=v)
    ws.cell(row=5, column=7, value="Outflows")
    ws.cell(row=6, column=1, value="▼Daily")
    ws.cell(row=7, column=1, value="Groceries")
    layout = _find_layout(ws)
    assert layout is not None
    assert layout.label_col == 1
    assert layout.start_month == 1


def test_sheet_sections_glyph_groups_and_gap_rule() -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    layout = LayoutRow(1, [], 0, 0, 1, 1)
    ws.cell(row=2, column=1, value="▼Daily")
    ws.cell(row=3, column=1, value="Groceries")
    ws.cell(row=4, column=1, value="Cafes")
    ws.cell(row=5, column=1, value="Month Summary")
    ws.cell(row=7, column=1, value="Rent")
    ws.cell(row=8, column=1, value="▲Inflow")
    ws.cell(row=9, column=1, value="Salary")
    sections = _sheet_sections(ws, layout)
    daily = sections[0]
    assert daily.name == "Daily"
    assert daily.kind == "expense"
    assert [name for _, name in daily.rows] == ["Groceries", "Cafes"]
    gap_group = sections[1]
    assert gap_group.name == "Rent"
    assert gap_group.rows == []
    inflow = sections[2]
    assert inflow.name == "Inflow"
    assert inflow.kind == "income"
    assert [name for _, name in inflow.rows] == ["Salary"]


def test_summary_value_matches_prefix_only() -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.cell(row=3, column=4, value="Income for January")
    ws.cell(row=3, column=3, value=6000)
    assert _summary_value(ws, 2, ("Income for",)) == 600000
    assert _summary_value(ws, 2, ("Nope",)) is None


def test_parse_year_sheet_reads_grid_income_available_seed() -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    _write_year(
        ws,
        months=[1, 2],
        rows=[
            ("▼Daily", None),
            ("Groceries", {1: (1000, 300, 900), 2: (1000, 500, 1400)}),
        ],
        options=YearOptions(
            header_row=8,
            income={1: 6000},
            available={1: 5100, 2: 4000},
            seed=1234,
        ),
    )
    layout = _find_layout(ws)
    assert layout is not None
    assert layout is not None
    parsed = _parse_year_sheet(ws, 2025, layout)
    groceries = parsed.cats["Groceries"]
    assert groceries.group == "Daily"
    assert groceries.budgets == {1: 100000, 2: 100000}
    assert groceries.outflows == {1: 30000, 2: 50000}
    assert groceries.balances == {1: 90000, 2: 140000}
    assert parsed.income == {1: 600000}
    assert parsed.available == {1: 510000, 2: 400000}
    assert parsed.seed == 123400
    assert parsed.months == [1, 2]


def _tx_only_ws(tx_rows: list[TransactionFixture | None]) -> Worksheet:
    wb = _new_wb()
    return _tx_sheet(wb, tx_rows)


def test_parse_transactions_empty_and_missing_columns() -> None:
    empty = _new_wb()
    empty.create_sheet(spec.SHEET_TRANSACTIONS)
    ro = load_workbook(BytesIO(_save(empty)), read_only=True, data_only=True)
    with pytest.raises(WorkbookError, match="Transactions sheet is empty"):
        _parse_transactions(ro[spec.SHEET_TRANSACTIONS], [], [])

    bad = _new_wb()
    ws = bad.create_sheet(spec.SHEET_TRANSACTIONS)
    ws.append(["Operation date", "Status"])
    with pytest.raises(WorkbookError, match="missing required columns"):
        _parse_transactions(ws, [], [])


def test_parse_transactions_dedup_status_currency_and_category() -> None:
    d = _workbook_datetime(2025, 1, 15, 10)
    ws = _tx_only_ws(
        [
            TransactionFixture(d, -300.0, "Groceries", description="Lenta"),
            TransactionFixture(d, -300.0, "Groceries", description="Lenta"),
            TransactionFixture(_workbook_datetime(2025, 1, 16), -50.0, "Cafes", status="FAILED"),
            TransactionFixture(
                _workbook_datetime(2025, 1, 17),
                -20.0,
                "Travel",
                currency="USD",
                description="Abroad",
            ),
            TransactionFixture("bad", -1.0, "X"),
            None,
        ],
    )
    warnings: list[str] = []
    errors: list[WorkbookParseErrorRow] = []
    rows = _parse_transactions(ws, warnings, errors)
    assert [r.monori_category for r in rows] == ["Groceries", "Travel"]
    first = rows[0]
    assert first.date == "2025-01-15T10:00:00"
    assert first.amount == -30000
    assert first.marker == "*1111"
    assert first.bank_category == "Super"
    assert first.mcc == "5411"
    assert [(error.row, error.error) for error in errors] == [(6, "unparseable date or amount")]
    assert (
        "Transactions: 1 rows identical in date, amount, description and card — kept once"
        in warnings
    )
    assert "Transactions: 1 non-OK rows skipped" in warnings
    assert [r.currency for r in rows] == ["RUB", "USD"]
    assert "Transactions: 1 rows in USD — they need an account held in USD to land on" in (warnings)


def test_split_operation_keeps_both_parts_with_their_own_amounts() -> None:
    """
    One card operation split across categories repeats the operation's full.

    amount on every part and carries each part's real share in "Payment amount" —.
    sometimes as a formula the user typed by hand. Reading the operation amount
    made the parts look identical and collapsed all but the first, dropping the
    money and the category the user assigned to it.
    """
    wb = _new_wb()
    ws = wb.create_sheet(spec.SHEET_TRANSACTIONS)
    ws.append([*TX_HEADER, "Payment amount", "Payment currency"])

    for row, category in ((2, "Clothes"), (3, "Wedding")):
        ws.cell(row, 1, _workbook_datetime(2026, 5, 31, 14, 30))
        ws.cell(row, 2, "*0548")
        ws.cell(row, 3, "OK")
        ws.cell(row, 4, -48480.0)
        ws.cell(row, 5, "RUB")
        ws.cell(row, 8, "Brandshop")
        ws.cell(row, 12, category)
    ws.cell(2, 9, "=-48480+16990")
    ws.cell(3, 9, -16990.0)

    warnings: list[str] = []
    errors: list[WorkbookParseErrorRow] = []
    rows = _parse_transactions(ws, warnings, errors)
    assert [(r.amount, r.monori_category) for r in rows] == [
        (-3149000, "Clothes"),
        (-1699000, "Wedding"),
    ]
    assert sum(r.amount for r in rows) == -4848000
    assert errors == []
    assert not any("duplicated" in w for w in warnings)


def test_parse_transactions_pay_amount_fallback_and_blankish_rows() -> None:
    wb = _new_wb()
    ws = wb.create_sheet(spec.SHEET_TRANSACTIONS)
    ws.append([*TX_HEADER, "Payment amount", "Payment currency"])

    ws.cell(2, 1, _workbook_datetime(2025, 3, 1, 9))
    ws.cell(2, 3, "OK")
    ws.cell(2, 5, "RUB")
    ws.cell(2, 8, "Salary")
    ws.cell(2, 9, 47337.0)
    ws.cell(2, 12, "Income")

    ws.cell(3, 1, _workbook_datetime(2025, 3, 2))
    ws.cell(3, 3, "OK")
    ws.cell(3, 8, "Points")
    ws.cell(3, 9, "-4 172,00")
    ws.cell(3, 10, "RUB")
    ws.cell(3, 12, "Groceries")

    ws.cell(4, 1, _workbook_datetime(2025, 3, 3))
    ws.cell(4, 3, "OK")
    ws.cell(4, 8, "Abroad")
    ws.cell(4, 9, 10.0)
    ws.cell(4, 10, "USD")
    ws.cell(4, 12, "Travel")

    ws.cell(5, 12, 0)
    ws.cell(6, 1, _workbook_datetime(2025, 3, 4))
    ws.cell(6, 8, "no amount anywhere")

    warnings: list[str] = []
    errors: list[WorkbookParseErrorRow] = []
    rows = _parse_transactions(ws, warnings, errors)
    assert [r.amount for r in rows] == [4733700, -417200, 1000]
    assert rows[0].monori_category == "Income"
    assert rows[1].monori_category == "Groceries"
    assert [(error.row, error.error) for error in errors] == [(6, "unparseable date or amount")]
    assert [r.currency for r in rows] == ["RUB", "RUB", "USD"]
    assert "Transactions: 1 rows in USD — they need an account held in USD to land on" in (warnings)


def test_parse_keywords_reads_side_table() -> None:
    ws = _tx_only_ws(
        [
            TransactionFixture(
                _workbook_datetime(2025, 1, 1),
                -10.0,
                "Groceries",
                keyword=("Groceries", "lenta|okey"),
            ),
            TransactionFixture(
                _workbook_datetime(2025, 1, 2), -10.0, "Cafes", keyword=("Cafes", "starbucks")
            ),
            TransactionFixture(_workbook_datetime(2025, 1, 3), -10.0, "X", keyword=("Skip", "a")),
        ],
    )
    idx: dict[str, int] = {name: i for i, name in enumerate(TX_HEADER)}
    kws = _parse_keywords(ws, idx)
    assert kws["Groceries"] == "lenta|okey"
    assert kws["Cafes"] == "starbucks"
    assert "Skip" not in kws


def _live_year_wb() -> Workbook:
    wb = Workbook()
    _drop_active_sheet(wb)
    _tx_sheet(
        wb,
        [
            TransactionFixture(
                _workbook_datetime(2025, 1, 15), -300.0, "Groceries", description="Lenta"
            ),
            TransactionFixture(
                _workbook_datetime(2025, 2, 10), -500.0, "Groceries", description="Okey"
            ),
            TransactionFixture(
                _workbook_datetime(2025, 1, 5),
                5000.0,
                "Income",
                card="*2222",
                description="Payroll",
            ),
            TransactionFixture(
                _workbook_datetime(2025, 1, 20), -100.0, "", description="Uncategorized"
            ),
        ],
    )
    ws = wb.create_sheet("2025")
    _write_year(
        ws,
        months=[1, 2],
        rows=[
            ("▼Daily", None),
            ("Groceries", {1: (1000, 300, 900), 2: (1000, 500, 1400)}),
            ("▲Inflow", None),
            ("Salary", None),
        ],
        options=YearOptions(
            header_row=8,
            start_token="JAN 2025",
            income={1: 6000},
            available={1: 5100, 2: 4000},
        ),
    )
    junk = wb.create_sheet("2019")
    junk.cell(row=1, column=1, value="not a year grid")
    return wb


def test_live_year_reconciles_rows_to_cached_totals() -> None:
    parsed = parse_workbook(_save(_live_year_wb()))

    assert [(group.name, group.sort, group.kind) for group in parsed.groups] == [
        ("Daily", 0, "expense"),
        ("Inflow", 1, "income"),
    ]
    assert [category.name for category in parsed.categories] == ["Groceries", "Salary", "Income"]

    budgets = {
        (budget.category, budget.year, budget.month): budget.amount for budget in parsed.budgets
    }
    assert budgets == {("Groceries", 2025, 1): 100000, ("Groceries", 2025, 2): 100000}

    synth = [tx for tx in parsed.transactions if tx.date.endswith("T12:00:00")]
    assert len(parsed.transactions) == 6
    assert {(tx.description, tx.amount) for tx in synth} == {
        ("Salary", 100000),
        ("Groceries", 20000),
    }

    assert any(
        w == "reconciliation: 2 adjustment transactions align live months with the sheet"
        for w in parsed.warnings
    )
    assert any(
        w.startswith("verify: the sheet's own Available differs") and "100.00 (2025-01)" in w
        for w in parsed.warnings
    )
    assert "2019: unrecognized year sheet layout, ignored" in parsed.warnings
    assert parsed.errors == []


def test_archive_history_and_seam_carry() -> None:
    wb = Workbook()
    _drop_active_sheet(wb)
    _tx_sheet(wb, [])

    _write_year(
        wb.create_sheet("2024_archive"),
        months=[1, 2],
        rows=[("▼Daily", None), ("Groceries", {1: (None, None, 500)})],
        options=YearOptions(
            start_token="JAN 2024",
            income={1: 100},
        ),
    )
    _write_year(
        wb.create_sheet("2024"),
        months=[1, 2],
        rows=[("▼Daily", None), ("Groceries", {2: (None, None, 800)})],
        options=YearOptions(
            start_token="JAN 2024",
        ),
    )
    _write_year(
        wb.create_sheet("2025"),
        months=[1, 2],
        rows=[("▼Daily", None), ("Groceries", {1: (None, None, 800)})],
        options=YearOptions(
            start_token="JAN 2025",
        ),
    )

    parsed = parse_workbook(_save(wb))
    synth = {(tx.description, tx.date): tx for tx in parsed.transactions}
    assert synth[("Income", "2024-01-31T12:00:00")].amount == 10000
    assert synth[("Groceries", "2024-01-31T12:00:00")].amount == 50000
    assert synth[("Groceries", "2024-12-31T12:00:00")].amount == 30000
    assert len(parsed.transactions) == 3

    assert any(w.startswith("history: 2 transactions stand in for") for w in parsed.warnings)
    assert "seam: 1 carry corrections at 2024-12" in parsed.warnings


def test_outflow_fallback_when_balance_cell_missing() -> None:
    wb = Workbook()
    _drop_active_sheet(wb)
    _tx_sheet(wb, [])
    _write_year(
        wb.create_sheet("2025"),
        months=[1, 2],
        rows=[("▼Daily", None), ("Groceries", {1: (None, None, 100), 2: (None, 200, None)})],
        options=YearOptions(
            start_token="JAN 2025",
        ),
    )
    parsed = parse_workbook(_save(wb))

    assert len(parsed.transactions) == 2
    synth = {(tx.description, tx.date): tx for tx in parsed.transactions}
    assert len(synth) == 2

    assert synth[("Groceries", "2025-02-28T12:00:00")].amount == 20000


def test_dead_category_and_available_seed_at_seam() -> None:
    wb = Workbook()
    _drop_active_sheet(wb)
    _tx_sheet(wb, [])
    _write_year(
        wb.create_sheet("2024_archive"),
        months=[1, 2],
        rows=[
            ("▼Daily", None),
            ("Groceries", {1: (None, None, 500)}),
            ("OldPhone", {1: (None, None, 300)}),
        ],
        options=YearOptions(
            start_token="JAN 2024",
        ),
    )
    _write_year(
        wb.create_sheet("2024"),
        months=[1, 2],
        rows=[("▼Daily", None), ("Groceries", {2: (None, None, 800)})],
        options=YearOptions(
            start_token="JAN 2024",
        ),
    )
    _write_year(
        wb.create_sheet("2025"),
        months=[1, 2],
        rows=[("▼Daily", None), ("Groceries", {1: (None, None, 800)})],
        options=YearOptions(
            start_token="JAN 2025",
            seed=200,
        ),
    )
    parsed = parse_workbook(_save(wb))
    synth = {(tx.description, tx.date): tx for tx in parsed.transactions}
    assert len(synth) == len(parsed.transactions)
    assert synth[("OldPhone", "2024-01-31T12:00:00")].amount == 30000
    assert synth[("OldPhone", "2024-12-31T12:00:00")].amount == -30000
    assert synth[("Income", "2024-12-31T12:00:00")].amount == 20000
    assert any(w.startswith("history: 2 transactions stand in for") for w in parsed.warnings)
    assert "seam: 3 carry corrections at 2024-12" in parsed.warnings


def test_available_seed_excludes_seam_overspend() -> None:
    """
    The template's "Not budgeted in Dec" seed is the December available BEFORE.

    overspend; the sheet adds "Overspent in Dec" separately in January. The.
    seed correction must therefore target avail alone, not avail + overspent.
    """
    wb = Workbook()
    _drop_active_sheet(wb)
    _tx_sheet(wb, [])
    _write_year(
        wb.create_sheet("2024_archive"),
        months=[1, 2],
        rows=[("▼Daily", None), ("Groceries", {1: (None, None, 500)})],
        options=YearOptions(
            start_token="JAN 2024",
            income={1: 100},
            header_row=8,
        ),
    )
    _write_year(
        wb.create_sheet("2024"),
        months=[1, 2],
        rows=[("▼Daily", None), ("Groceries", {2: (None, None, -100)})],
        options=YearOptions(
            start_token="JAN 2024",
        ),
    )
    _write_year(
        wb.create_sheet("2025"),
        months=[1, 2],
        rows=[("▼Daily", None), ("Groceries", None)],
        options=YearOptions(
            start_token="JAN 2025",
            seed=200,
            available={1: 100},
            header_row=8,
        ),
    )
    parsed = parse_workbook(_save(wb))
    synth = {(tx.description, tx.date): tx for tx in parsed.transactions}
    assert synth[("Groceries", "2024-12-31T12:00:00")].amount == -60000
    assert synth[("Income", "2024-12-31T12:00:00")].amount == 10000
    assert not any(w.startswith("verify:") for w in parsed.warnings)


def test_russian_header_labels_are_read_like_the_english_ones() -> None:
    """
    The live spreadsheet labels its month headers in Russian. Every figure the.

    reconciliation leans on — income, the carried-over remainder, the running.
    Available — is found by those labels, so a sheet in Russian has to yield
    exactly what the same sheet in English does.
    """
    wb = Workbook()
    _drop_active_sheet(wb)
    _tx_sheet(
        wb,
        [
            TransactionFixture(
                _workbook_datetime(2025, 1, 15), -300.0, "Groceries", description="Lenta"
            )
        ],
    )
    _write_year(
        wb.create_sheet("2025"),
        months=[1, 2],
        rows=[("▼Daily", None), ("Groceries", {1: (1000, -300, 700)})],
        options=YearOptions(
            income={1: 5000},
            available={1: 5900},
            seeds={1: 1900},
            labels="ru",
            header_row=8,
        ),
    )
    parsed = parse_workbook(_save(wb))
    opening = next(tx for tx in parsed.transactions if tx.description == "Opening balance")
    assert opening.amount == 190000
    assert not any(w.startswith("verify:") for w in parsed.warnings)
    income = next(tx for tx in parsed.transactions if tx.description == "Income")
    assert income.amount == 500000


def test_available_label_is_found_wherever_the_summary_block_puts_it() -> None:
    wb = Workbook()
    _drop_active_sheet(wb)
    _tx_sheet(
        wb,
        [
            TransactionFixture(
                _workbook_datetime(2025, 1, 15), -300.0, "Groceries", description="Lenta"
            )
        ],
    )
    _write_year(
        wb.create_sheet("2025"),
        months=[1, 2],
        rows=[("▼Daily", None), ("Groceries", {1: (1000, -300, 700)})],
        options=YearOptions(
            income={1: 5000},
            available={1: 99999},
            available_row=7,
            header_row=8,
        ),
    )
    parsed = parse_workbook(_save(wb))
    assert any(w.startswith("verify:") for w in parsed.warnings)


def test_summary_labels_below_the_header_block_are_not_read() -> None:
    """
    The summary block is the first six rows; row seven is where the grid's own.

    Budgeted/Outflows/Balance header sits on the live sheet. Reading one row.
    further would take a column heading for a figure.
    """
    wb = Workbook()
    _drop_active_sheet(wb)
    _tx_sheet(wb, [])
    ws = wb.create_sheet("2025")
    _write_year(
        ws,
        months=[1, 2],
        rows=[("▼Daily", None), ("Groceries", {1: (1000, 0, 1000)})],
        options=YearOptions(
            header_row=8,
        ),
    )
    ws.cell(row=7, column=4, value="Income for month")
    ws.cell(row=7, column=3, value=777)
    parsed = parse_workbook(_save(wb))
    assert not any(tx.monori_category == "Income" for tx in parsed.transactions)


def test_a_sheet_running_to_december_keeps_its_last_month() -> None:
    _, ws = _one_year_wb(
        months=list(range(1, 13)),
        rows=[("▼Daily", None), ("Groceries", {12: (1000, -300, 700)})],
    )
    layout = _find_layout(ws)
    assert layout is not None
    parsed = _parse_year_sheet(ws, 2025, layout)
    assert parsed.months[-1] == 12
    assert parsed.cats["Groceries"].balances[12] == 70000


def test_history_and_adjustment_split_follows_the_rows_not_the_sheet_name() -> None:
    """
    A workbook can keep years of history on ordinary year sheets and never write.

    `_archive` once. What makes a correction a stand-in is that the month has no.
    rows of its own, so counting by sheet name reports hundreds of history rows
    as live adjustments.
    """
    wb = Workbook()
    _drop_active_sheet(wb)
    _tx_sheet(
        wb,
        [
            TransactionFixture(
                _workbook_datetime(2025, 1, 15), -300.0, "Groceries", description="Lenta"
            )
        ],
    )
    _write_year(
        wb.create_sheet("2024"),
        months=[1, 2],
        rows=[("▼Daily", None), ("Groceries", {1: (1000, -700, 300), 2: (0, 0, 300)})],
        options=YearOptions(
            start_token="JAN 2024",
            header_row=8,
        ),
    )
    _write_year(
        wb.create_sheet("2025"),
        months=[1, 2],
        rows=[("▼Daily", None), ("Groceries", {1: (1000, -900, 400)})],
        options=YearOptions(
            start_token="JAN 2025",
            header_row=8,
        ),
    )
    parsed = parse_workbook(_save(wb))

    assert any(w.startswith("history: 1 transactions stand in for") for w in parsed.warnings)
    assert (
        "reconciliation: 1 adjustment transactions align live months with the sheet"
        in parsed.warnings
    )


def test_opening_balance_is_taken_from_the_first_month_with_rows() -> None:
    """
    A spreadsheet is started with money already in hand, and the only place that.

    money exists is the header cell of its first real month: what was left.
    unbudgeted the month before. The earlier blocks of that year are empty
    scaffolding, so reading the seed off the January block finds nothing.
    """
    wb = Workbook()
    _drop_active_sheet(wb)
    _tx_sheet(
        wb,
        [
            TransactionFixture(
                _workbook_datetime(2025, 7, 15), -300.0, "Groceries", description="Lenta"
            )
        ],
    )
    _write_year(
        wb.create_sheet("2025"),
        months=[6, 7],
        rows=[("▼Daily", None), ("Groceries", {7: (1000, -300, 700)})],
        options=YearOptions(
            start_token="JUN 2025",
            income={7: 5000},
            available={7: 5900},
            seeds={6: 0, 7: 1900},
            header_row=8,
        ),
    )
    parsed = parse_workbook(_save(wb))
    synth = {(tx.description, tx.date): tx for tx in parsed.transactions}
    assert synth[("Opening balance", "2025-06-30T12:00:00")].amount == 190000
    assert any(w.startswith("opening balance: 1,900.00") for w in parsed.warnings)

    assert not any(w.startswith("verify:") for w in parsed.warnings)


def test_opening_balance_predating_the_sheet_is_dated_before_it() -> None:
    """
    When the very first month of the earliest sheet already has rows, the money.

    it started with belongs to the December before — a month no sheet covers.
    The row still has to exist, or Available starts short by that amount.
    """
    wb = Workbook()
    _drop_active_sheet(wb)
    _tx_sheet(
        wb,
        [
            TransactionFixture(
                _workbook_datetime(2025, 1, 15), -300.0, "Groceries", description="Lenta"
            )
        ],
    )
    _write_year(
        wb.create_sheet("2025"),
        months=[1, 2],
        rows=[("▼Daily", None), ("Groceries", {1: (1000, -300, 700)})],
        options=YearOptions(
            income={1: 5000},
            available={1: 5900},
            seeds={1: 1900},
            header_row=8,
        ),
    )
    parsed = parse_workbook(_save(wb))
    opening = next(tx for tx in parsed.transactions if tx.description == "Opening balance")
    assert opening.date == "2024-12-31T12:00:00"
    assert opening.amount == 190000
    assert not any(w.startswith("verify:") for w in parsed.warnings)


def test_activity_span_reads_two_digit_months_whole() -> None:
    """
    A ledger that only starts in October is where a one-character month slice.

    stops being harmless: it would read 11 as 1 and seed the opening balance.
    ten months early, off the wrong header cell.
    """
    wb = Workbook()
    _drop_active_sheet(wb)
    _tx_sheet(
        wb,
        [
            TransactionFixture(
                _workbook_datetime(2025, 11, 15), -300.0, "Groceries", description="Lenta"
            )
        ],
    )
    _write_year(
        wb.create_sheet("2025"),
        months=[10, 11],
        rows=[("▼Daily", None), ("Groceries", {11: (1000, -300, 700)})],
        options=YearOptions(
            start_token="OCT 2025",
            income={11: 5000},
            seeds={10: 400, 11: 1900},
            header_row=8,
        ),
    )
    parsed = parse_workbook(_save(wb))
    opening = next(tx for tx in parsed.transactions if tx.description == "Opening balance")
    assert opening.date == "2025-10-31T12:00:00"
    assert opening.amount == 190000


def test_opening_balance_left_alone_when_the_sheet_starts_from_nothing() -> None:
    wb = Workbook()
    _drop_active_sheet(wb)
    _tx_sheet(
        wb,
        [
            TransactionFixture(
                _workbook_datetime(2025, 7, 15), -300.0, "Groceries", description="Lenta"
            )
        ],
    )
    _write_year(
        wb.create_sheet("2025"),
        months=[6, 7],
        rows=[("▼Daily", None), ("Groceries", {7: (1000, -300, 700)})],
        options=YearOptions(
            start_token="JUN 2025",
            income={7: 5000},
            seeds={6: 0, 7: 0},
            header_row=8,
        ),
    )
    parsed = parse_workbook(_save(wb))
    assert not any(tx.description == "Opening balance" for tx in parsed.transactions)
    assert not any(w.startswith("opening balance:") for w in parsed.warnings)


def test_available_ignores_budget_cells_on_income_categories() -> None:
    """
    The budget grid only spends down expense envelopes, so a Budgeted cell on an.

    income row buys nothing and must not be subtracted from Available — which is.
    what the client computes and therefore what the verify check has to assume.
    """
    wb = Workbook()
    _drop_active_sheet(wb)
    _tx_sheet(wb, [])
    _write_year(
        wb.create_sheet("2025"),
        months=[1],
        rows=[
            ("▲Inflow", None),
            ("Salary", {1: (700, None, None)}),
            ("▼Daily", None),
            ("Groceries", {1: (1000, -300, 700)}),
        ],
        options=YearOptions(
            income={1: 5000},
            available={1: 4000},
            header_row=8,
        ),
    )
    parsed = parse_workbook(_save(wb))
    assert not any(w.startswith("verify:") for w in parsed.warnings)


def test_missing_transactions_sheet_raises() -> None:
    wb = Workbook()
    ws = _active(wb)
    ws.title = "2025"
    with pytest.raises(WorkbookError, match="missing required sheet: Transactions"):
        parse_workbook(_save(wb))


def test_trailing_zero_cached_months_get_no_synthetic_rows() -> None:
    """
    An empty trailing block keeps cached zero balances from its formulas; the.

    reconciliation must not fabricate rows there to zero out the carry.
    """
    wb = Workbook()
    _drop_active_sheet(wb)
    _tx_sheet(
        wb,
        [
            TransactionFixture(
                _workbook_datetime(2025, 1, 15), -300.0, "Groceries", description="Lenta"
            )
        ],
    )
    _write_year(
        wb.create_sheet("2025"),
        months=[1, 2, 3],
        rows=[("▼Daily", None), ("Groceries", {1: (1000, 300, 700), 2: (0, 0, 0), 3: (0, 0, 0)})],
        options=YearOptions(
            income={1: 5000},
            header_row=8,
        ),
    )
    parsed = parse_workbook(_save(wb))
    assert all(tx.date < "2025-02" for tx in parsed.transactions)


def test_month_with_blank_category_gets_an_explicit_grid_correction() -> None:
    """
    The historical row stays uncategorized, while the grid's category total is.

    represented by a separate correction. That preserves the source row exactly.
    and lets the imported budget balance equal the spreadsheet.
    """
    wb = Workbook()
    _drop_active_sheet(wb)
    _tx_sheet(
        wb,
        [TransactionFixture(_workbook_datetime(2025, 1, 26), -450.0, "", description="Y.M*lb")],
    )
    _write_year(
        wb.create_sheet("2025"),
        months=[1, 2],
        rows=[("▼Daily", None), ("LifeLink", {1: (0, 450, -450)})],
        options=YearOptions(
            income={1: 0},
            header_row=8,
        ),
    )
    parsed = parse_workbook(_save(wb))
    assert [(tx.date, tx.amount, tx.monori_category) for tx in parsed.transactions] == [
        ("2025-01-26T00:00:00", -45000, ""),
        ("2025-01-31T12:00:00", -45000, "LifeLink"),
    ]


def test_uncategorized_trailing_tx_does_not_extend_reconciliation() -> None:
    """
    Uncategorized transactions are ignored by the reconciliation sums, so a.

    trailing month whose only activity is one must not be reconciled — its.
    zero-cached cells would fabricate a synthetic row against the carry.
    """
    wb = Workbook()
    _drop_active_sheet(wb)
    _tx_sheet(
        wb,
        [
            TransactionFixture(
                _workbook_datetime(2025, 1, 15), -300.0, "Groceries", description="Lenta"
            ),
            TransactionFixture(_workbook_datetime(2025, 2, 10), -50.0, "", description="Mystery"),
        ],
    )
    _write_year(
        wb.create_sheet("2025"),
        months=[1, 2],
        rows=[("▼Daily", None), ("Groceries", {1: (1000, 300, 700), 2: (0, 0, 0)})],
        options=YearOptions(
            income={1: 5000},
            header_row=8,
        ),
    )
    parsed = parse_workbook(_save(wb))
    synth = [tx for tx in parsed.transactions if not tx.marker]
    assert all(tx.date < "2025-02" for tx in synth)


def test_prepared_next_year_sheet_adds_no_future_rows() -> None:
    wb = _live_year_wb()
    _write_year(
        wb.create_sheet("2026"),
        months=[1, 2],
        rows=[("▼Daily", None), ("Groceries", {1: (0, 0, 0), 2: (0, 0, 0)})],
        options=YearOptions(
            start_token="JAN 2026",
            header_row=8,
        ),
    )
    parsed = parse_workbook(_save(wb))
    assert max(tx.date for tx in parsed.transactions) < "2025-03"


def test_parse_template_rejects_garbage_bytes() -> None:
    with pytest.raises(WorkbookError, match="not a readable .xlsx workbook"):  # noqa: RUF043
        parse_workbook(b"nope")


def test_live_layout_locates_category_and_keywords_by_content() -> None:
    """
    The live template's keyword table starts at row 1 (its cells pollute the.

    header index) and the category column follows the bank headers with no.
    gap column — both must be found by content, not fixed offsets.

    There are two candidate columns: the keyword rules guess into the first,
    the second carries that guess through or replaces it with a hand-written
    category, and only the second is what the sheet's totals are built from.
    The hand label must win and the guess must never override it.
    """
    wb = Workbook()
    _drop_active_sheet(wb)
    ws = wb.create_sheet(spec.SHEET_TRANSACTIONS)
    ws.append([None, *TX_HEADER, None, None, "Income", "salary|bonus"])
    row = [None, _workbook_datetime(2025, 1, 1), "*1111", "OK", -100.0, "RUB", "Super", "5411"]
    ws.append([*[*row, "Lenta"][:9], "Groceries", "Groceries", "MTS", "mts|phone"])
    row2 = [None, _workbook_datetime(2025, 1, 2), "*1111", "OK", -50.0, "RUB", "Super", "5411"]
    ws.append([*[*row2, "Okey"][:9], "Cafes", "Lunch", "Cafes", "coffee|bar"])
    row3 = [None, _workbook_datetime(2025, 1, 3), "*1111", "OK", -20.0, "RUB", "Super", "5411"]
    ws.append([*[*row3, "Metro"][:9], "", "Transport"])
    idx = _tx_header_index(ws)
    assert idx is not None
    warnings: list[str] = []
    errors: list[WorkbookParseErrorRow] = []
    rows = _parse_transactions(ws, warnings, errors)
    assert errors == []
    assert [r.monori_category for r in rows] == ["Groceries", "Lunch", "Transport"]
    kws = _parse_keywords(ws, idx)
    assert kws == {"Income": "salary|bonus", "MTS": "mts|phone", "Cafes": "coffee|bar"}


def test_future_budgets_do_not_extend_reconciliation() -> None:
    """
    Budget cells in future months are planning, not activity; their stale.

    cached balances must not be reconciled into future-dated synthetic rows,.
    while the budgets themselves are still imported.
    """
    wb = Workbook()
    _drop_active_sheet(wb)
    _tx_sheet(
        wb,
        [
            TransactionFixture(
                _workbook_datetime(2025, 1, 15), -300.0, "Groceries", description="Lenta"
            )
        ],
    )
    _write_year(
        wb.create_sheet("2025"),
        months=[1, 2, 3],
        rows=[
            ("▼Daily", None),
            ("Groceries", {1: (1000, 300, 700), 2: (500, None, 0), 3: (None, None, 0)}),
        ],
        options=YearOptions(
            income={1: 5000},
            header_row=8,
        ),
    )
    parsed = parse_workbook(_save(wb))
    assert all(tx.date < "2025-02" for tx in parsed.transactions)
    assert {(budget.year, budget.month) for budget in parsed.budgets} == {(2025, 1), (2025, 2)}


def test_parse_keywords_falls_back_without_pipes() -> None:
    """
    No pipe anywhere -> content detection abstains and the known-header.

    positional fallback (immune to side-table pollution) must still hit.
    """
    ws = _tx_only_ws(
        [
            TransactionFixture(
                _workbook_datetime(2025, 1, 2),
                -10.0,
                "Cafes",
                keyword=("Cafes", "starbucks"),
            )
        ],
    )
    idx: dict[str, int] = {name: i for i, name in enumerate(TX_HEADER)}
    assert _parse_keywords(ws, idx) == {"Cafes": "starbucks"}


def test_label_col_picks_the_fullest_column_below_the_header() -> None:
    """
    The live grid never names its category column, so it is found by weight:.

    of the columns left of the first month block, the one carrying the most.
    labels — counted strictly below the header row, since the header itself and
    whatever title sits above it are not categories.
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.cell(row=4, column=1, value="Budget for year")
    ws.cell(row=5, column=1, value="Category")
    for r in range(6, 10):
        ws.cell(row=r, column=1, value=f"left {r}")
    for r in range(6, 11):
        ws.cell(row=r, column=2, value=f"middle {r}")
    assert _label_col(ws, 5, 4) == 2


def test_label_col_keeps_the_leftmost_of_a_tie_and_falls_back_to_the_first() -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    assert _label_col(ws, 5, 4) == 1
    for c in (1, 2):
        for r in range(6, 9):
            ws.cell(row=r, column=c, value=f"c{c} r{r}")
    assert _label_col(ws, 5, 4) == 1


def test_label_col_only_counts_the_rows_just_under_the_header() -> None:
    """
    A sheet carries hundreds of rows below its grid — notes, a second table,.

    leftovers. Only the band the categories live in decides the column.
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    for r in range(6, 9):
        ws.cell(row=r, column=1, value=f"category {r}")
    for r in range(80, 140):
        ws.cell(row=r, column=2, value=f"note {r}")
    assert _label_col(ws, 5, 4) == 1

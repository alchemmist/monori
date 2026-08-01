"""
Characterization tests aimed at the parser's least-exercised helpers. The
end-to-end tests pin whole imports; these pin the exact return of the small
functions those imports lean on, so a single flipped offset, dropped default or
renamed key is caught here instead of slipping through as a survived mutant.
"""

import datetime
from collections.abc import Sequence
from dataclasses import dataclass
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.workbook import spec
from app.workbook.models import WorkbookGroup as WorkbookGroupRow
from app.workbook.models import WorkbookParseError as WorkbookParseErrorRow
from app.workbook.parser import (
    TX_ALIASES,
    _category_col,
    _find_keyword_block,
    _find_layout,
    _kind_of,
    _known_max_col,
    _label_col,
    _parse_categories,
    _parse_keywords,
    _parse_transactions,
    _parse_year_sheet,
    _synthetic,
    _tx_header_index,
    parse_workbook,
)

TX_HEADER: list[str | None] = [
    TX_ALIASES[f][0]
    for f in ("date", "card", "status", "amount", "currency", "bank_category", "mcc", "description")
]


def _active(wb: Workbook) -> Worksheet:
    ws = wb.active
    assert ws is not None
    return ws


def _save(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- _parse_categories ----------------------------------------------------


def _sheet(title: str) -> tuple[Workbook, Worksheet]:
    wb = Workbook()
    ws = _active(wb)
    ws.title = title
    return wb, ws


def test_parse_categories_reads_group_table_and_rows_exactly() -> None:
    _, ws = _sheet(spec.SHEET_CATEGORIES)
    ws.append(["Daily", 2, "OUT"])
    ws.append(["▲Inflow", 1, "IN"])
    ws.append([1, "Daily", "Groceries", "lenta|okey"])
    ws.append([2, "▼Daily", "Cafes", ""])
    ws.append([1, "Inflow", "Salary", ""])
    warnings: list[str] = []
    groups, categories = _parse_categories(ws, warnings)
    assert [(group.name, group.sort, group.kind) for group in groups] == [
        ("Daily", 2, "expense"),
        ("Inflow", 1, "income"),
    ]
    assert [(c.group, c.group_kind, c.group_sort, c.name, c.keywords) for c in categories] == [
        ("Daily", None, 1, "Groceries", "lenta|okey"),
        ("Daily", "expense", 2, "Cafes", ""),
        ("Inflow", None, 1, "Salary", ""),
    ]
    assert warnings == []


def test_parse_categories_skips_header_and_blank_rows() -> None:
    _, ws = _sheet(spec.SHEET_CATEGORIES)
    ws.append(["Sort Order", "Category Group", "Category"])
    ws.append([None, None, None])
    ws.append([1, "Daily", "Groceries", ""])
    warnings: list[str] = []
    groups, categories = _parse_categories(ws, warnings)
    assert [c.name for c in categories] == ["Groceries"]
    assert warnings == ["Categories: group table missing, groups derived from category rows"]


def test_parse_categories_derives_groups_when_table_missing() -> None:
    _, ws = _sheet(spec.SHEET_CATEGORIES)
    ws.append([1, "Daily", "Groceries", ""])
    ws.append([1, "Daily", "Cafes", ""])
    ws.append([3, "▲Inflow", "Salary", ""])
    warnings: list[str] = []
    groups, categories = _parse_categories(ws, warnings)
    assert [(group.name, group.sort, group.kind) for group in groups] == [
        ("Daily", 1, "expense"),
        ("Inflow", 3, "income"),
    ]
    assert [c.name for c in categories] == ["Groceries", "Cafes", "Salary"]
    assert warnings == ["Categories: group table missing, groups derived from category rows"]


def test_parse_categories_warns_on_unrecognized_rows() -> None:
    _, ws = _sheet(spec.SHEET_CATEGORIES)
    ws.append([1, "Daily", "Groceries", ""])
    ws.append(["Note to self", None, None])
    warnings: list[str] = []
    _parse_categories(ws, warnings)
    assert "Categories: unrecognized row skipped: ['Note to self', '', '']" in warnings


# --- _find_layout ---------------------------------------------------------


def _grid_ws() -> Workbook:
    wb = Workbook()
    return wb


def test_find_layout_reads_header_on_the_first_row() -> None:
    ws = _active(_grid_ws())
    assert ws is not None
    ws.cell(row=1, column=1, value="ЯНВ 2025")
    ws.cell(row=1, column=2, value="Budgeted")
    ws.cell(row=1, column=3, value="Outflows")
    ws.cell(row=1, column=4, value="Balance")
    ws.cell(row=1, column=6, value="Budgeted")
    ws.cell(row=1, column=7, value="Outflows")
    ws.cell(row=1, column=8, value="Balance")
    layout = _find_layout(ws)
    assert layout is not None
    assert layout.header_row == 1
    assert layout.bases == [2, 6]
    assert layout.out_off == 1
    assert layout.bal_off == 2
    assert layout.start_month == 1


def test_find_layout_reads_the_balance_header_in_the_last_column() -> None:
    ws = _active(_grid_ws())
    assert ws is not None
    ws.cell(row=5, column=2, value="Budgeted")
    ws.cell(row=5, column=3, value="Outflows")
    ws.cell(row=5, column=6, value="Budgeted")
    ws.cell(row=5, column=7, value="Outflows")
    ws.cell(row=5, column=8, value="Balance")
    layout = _find_layout(ws)
    assert layout is not None
    assert layout.bal_off == 6


def test_find_layout_returns_none_when_balance_header_is_absent() -> None:
    ws = _active(_grid_ws())
    assert ws is not None
    ws.cell(row=5, column=2, value="Budgeted")
    ws.cell(row=5, column=3, value="Outflows")
    ws.cell(row=5, column=6, value="Budgeted")
    ws.cell(row=5, column=7, value="Outflows")
    assert _find_layout(ws) is None


def test_find_layout_finds_the_label_header_one_row_below_the_grid() -> None:
    ws = _active(_grid_ws())
    assert ws is not None
    for c, v in ((2, "Budgeted"), (3, "Outflows"), (4, "Balance")):
        ws.cell(row=5, column=c, value=v)
    for c, v in ((6, "Budgeted"), (7, "Outflows"), (8, "Balance")):
        ws.cell(row=5, column=c, value=v)
    # the label header sits on the row under the grid header, and a fuller
    # column to its right must not win the fallback
    ws.cell(row=6, column=1, value="Category")
    for r in range(6, 12):
        ws.cell(row=r, column=3, value=f"note {r}")
    layout = _find_layout(ws)
    assert layout is not None
    assert layout.label_col == 1


def test_find_layout_reads_the_start_month_from_the_second_row() -> None:
    ws = _active(_grid_ws())
    assert ws is not None
    for c, v in ((2, "Budgeted"), (3, "Outflows"), (4, "Balance")):
        ws.cell(row=5, column=c, value=v)
    for c, v in ((6, "Budgeted"), (7, "Outflows"), (8, "Balance")):
        ws.cell(row=5, column=c, value=v)
    ws.cell(row=2, column=2, value="ФЕВ 2025")  # FEB, only on row 2
    layout = _find_layout(ws)
    assert layout is not None
    assert layout.start_month == 2


def test_find_layout_reads_the_start_month_from_the_third_row() -> None:
    ws = _active(_grid_ws())
    assert ws is not None
    for c, v in ((2, "Budgeted"), (3, "Outflows"), (4, "Balance")):
        ws.cell(row=5, column=c, value=v)
    for c, v in ((6, "Budgeted"), (7, "Outflows"), (8, "Balance")):
        ws.cell(row=5, column=c, value=v)
    ws.cell(row=3, column=2, value="МАР 2025")  # MAR, only on row 3
    layout = _find_layout(ws)
    assert layout is not None
    assert layout.start_month == 3


# --- _label_col -----------------------------------------------------------


def test_label_col_counts_a_single_label_column() -> None:
    ws = _active(_grid_ws())
    assert ws is not None
    ws.cell(row=5, column=1, value="Category")
    ws.cell(row=6, column=2, value="only one label")
    assert _label_col(ws, 5, 4) == 2


def test_label_col_counts_the_row_immediately_below_the_header() -> None:
    ws = _active(_grid_ws())
    assert ws is not None
    ws.cell(row=6, column=2, value="just below header")
    assert _label_col(ws, 5, 4) == 2


# --- _kind_of -------------------------------------------------------------


def test_kind_of_defaults_to_expense_for_unknown_group() -> None:
    groups = [WorkbookGroupRow(name="Inflow", kind="income", sort=1)]
    assert _kind_of("Inflow", groups) == "income"
    assert _kind_of("Nowhere", groups) == "expense"


# --- _known_max_col -------------------------------------------------------


def test_known_max_col_defaults_to_minus_one_without_known_headers() -> None:
    assert _known_max_col({"Whatever": 0, "Else": 1}) == -1
    assert _known_max_col({"Date": 3, "Amount": 5, "Extra": 9}) == 5


# --- _synthetic -----------------------------------------------------------


def test_synthetic_carries_the_default_currency() -> None:
    row = _synthetic(2025, 3, -1000, "Groceries", "Groceries")
    assert row.currency == "RUB"
    assert row.marker == ""


# --- _parse_year_sheet ----------------------------------------------------


def test_parse_year_sheet_stops_months_at_december() -> None:
    ws = _active(_grid_ws())
    assert ws is not None
    ws.cell(row=1, column=2, value="ДЕК 2025")  # DEC
    for c, v in ((2, "Budgeted"), (3, "Outflows"), (4, "Balance")):
        ws.cell(row=5, column=c, value=v)
    for c, v in ((6, "Budgeted"), (7, "Outflows"), (8, "Balance")):
        ws.cell(row=5, column=c, value=v)
    layout = _find_layout(ws)
    assert layout is not None
    assert layout.start_month == 12
    parsed = _parse_year_sheet(ws, 2025, layout)
    assert parsed.months == [12]
    assert isinstance(parsed.cats, dict)


def test_parse_year_sheet_reads_russian_available_label() -> None:
    ws = _active(_grid_ws())
    assert ws is not None
    ws.cell(row=1, column=2, value="ЯНВ 2025")
    for c, v in ((2, "Budgeted"), (3, "Outflows"), (4, "Balance")):
        ws.cell(row=8, column=c, value=v)
    for c, v in ((6, "Budgeted"), (7, "Outflows"), (8, "Balance")):
        ws.cell(row=8, column=c, value=v)
    # summary block: Russian "Available" (Доступный) label on row 6, value row 5
    ws.cell(row=6, column=3, value="Доступный остаток")
    ws.cell(row=5, column=3, value=4200)
    layout = _find_layout(ws)
    assert layout is not None
    parsed = _parse_year_sheet(ws, 2025, layout)
    assert parsed.available == {1: 420000}


# --- _parse_transactions --------------------------------------------------


@dataclass(frozen=True)
class PlainTransactionFixture:
    date: datetime.datetime
    amount: float
    category: str
    status: str = "OK"
    currency: str = "RUB"
    description: str = ""
    card: str = "*1111"

    def append_to(self, ws: Worksheet) -> None:
        row = ws.max_row + 1
        ws.cell(row, 1, self.date)
        ws.cell(row, 2, self.card)
        ws.cell(row, 3, self.status)
        ws.cell(row, 4, self.amount)
        ws.cell(row, 5, self.currency)
        ws.cell(row, 8, self.description)
        ws.cell(row, 10, self.category)


def _tx_ws(
    header: Sequence[str | None], rows: Sequence[PlainTransactionFixture | None]
) -> Workbook:
    wb = Workbook()
    wb.remove(_active(wb))
    ws = wb.create_sheet(spec.SHEET_TRANSACTIONS)
    ws.append(header)
    for row in rows:
        if row is None:
            ws.append([])
        else:
            row.append_to(ws)
    return wb


def test_parse_transactions_uses_pay_amount_when_amount_column_is_absent() -> None:
    header = ["Дата операции", "Статус", "Сумма платежа", "Описание"]
    tx_wb, ws = _sheet(spec.SHEET_TRANSACTIONS)
    ws.append(header)
    ws.append([datetime.datetime(2025, 1, 5), "OK", -1500.0, "Shop"])
    ws = _active(tx_wb)
    assert ws is not None
    warnings: list[str] = []
    errors: list[WorkbookParseErrorRow] = []
    rows = _parse_transactions(ws, warnings, errors)
    assert errors == []
    assert [r.amount for r in rows] == [-150000]


def test_parse_transactions_empty_message_is_exact() -> None:
    from app.workbook.parser import WorkbookError

    wb = Workbook()
    wb.remove(_active(wb))
    wb.create_sheet(spec.SHEET_TRANSACTIONS)
    ro = load_workbook(BytesIO(_save(wb)), read_only=True, data_only=True)
    with pytest.raises(WorkbookError) as exc:
        _parse_transactions(ro[spec.SHEET_TRANSACTIONS], [], [])
    assert str(exc.value) == "Transactions sheet is empty"


def _plain_tx(
    date: datetime.datetime,
    amount: float,
    category: str,
    *,
    status: str = "OK",
    currency: str = "RUB",
    desc: str = "",
    card: str = "*1111",
) -> PlainTransactionFixture:
    return PlainTransactionFixture(date, amount, category, status, currency, desc, card)


def test_parse_transactions_counts_every_skipped_status_row() -> None:
    tx_wb = _tx_ws(
        TX_HEADER + [None, "Monori Category"],
        [
            _plain_tx(datetime.datetime(2025, 1, 1), -10.0, "A", status="FAILED", desc="a"),
            _plain_tx(datetime.datetime(2025, 1, 2), -20.0, "B", status="DECLINED", desc="b"),
            _plain_tx(datetime.datetime(2025, 1, 3), -30.0, "C", desc="c"),
        ],
    )
    ws = _active(tx_wb)
    assert ws is not None
    warnings: list[str] = []
    errors: list[WorkbookParseErrorRow] = []
    _parse_transactions(ws, warnings, errors)
    assert "Transactions: 2 non-OK rows skipped" in warnings


def test_parse_transactions_counts_every_foreign_currency_row() -> None:
    tx_wb = _tx_ws(
        TX_HEADER + [None, "Monori Category"],
        [
            _plain_tx(datetime.datetime(2025, 1, 1), -10.0, "A", currency="USD", desc="a"),
            _plain_tx(datetime.datetime(2025, 1, 2), -20.0, "B", currency="USD", desc="b"),
        ],
    )
    ws = _active(tx_wb)
    assert ws is not None
    warnings: list[str] = []
    errors: list[WorkbookParseErrorRow] = []
    _parse_transactions(ws, warnings, errors)
    assert "Transactions: 2 rows in USD — they need an account held in USD to land on" in warnings


def test_parse_transactions_counts_every_duplicate_row() -> None:
    tx_wb = _tx_ws(
        TX_HEADER + [None, "Monori Category"],
        [
            _plain_tx(datetime.datetime(2025, 1, 1), -10.0, "A", desc="same"),
            _plain_tx(datetime.datetime(2025, 1, 1), -10.0, "A", desc="same"),
            _plain_tx(datetime.datetime(2025, 1, 1), -10.0, "A", desc="same"),
        ],
    )
    ws = _active(tx_wb)
    assert ws is not None
    warnings: list[str] = []
    errors: list[WorkbookParseErrorRow] = []
    rows = _parse_transactions(ws, warnings, errors)
    assert len(rows) == 1
    assert (
        "Transactions: 2 rows identical in date, amount, description and card — kept once"
        in warnings
    )


def test_parse_transactions_keeps_reading_after_a_blank_row() -> None:
    tx_wb = _tx_ws(
        TX_HEADER + [None, "Monori Category"],
        [
            None,
            _plain_tx(datetime.datetime(2025, 1, 2), -20.0, "B", desc="after blank"),
        ],
    )
    ws = _active(tx_wb)
    assert ws is not None
    warnings: list[str] = []
    errors: list[WorkbookParseErrorRow] = []
    rows = _parse_transactions(ws, warnings, errors)
    assert [r.description for r in rows] == ["after blank"]


# --- _parse_keywords ------------------------------------------------------


def test_parse_keywords_keeps_two_character_keywords() -> None:
    header = TX_HEADER + [None, None, "Category", "Keywords"]
    wb = Workbook()
    wb.remove(_active(wb))
    ws = wb.create_sheet(spec.SHEET_TRANSACTIONS)
    ws.append(header)
    ws.append([None] * 10 + [None, None, "Taxi", "gо"])
    ws.append([None] * 10 + [None, None, "Food", "a|b"])
    idx = _tx_header_index(ws)
    assert idx is not None
    kws = _parse_keywords(ws, idx)
    assert kws == {"Taxi": "gо", "Food": "a|b"}


def test_parse_keywords_finds_the_block_by_content_not_a_fixed_offset() -> None:
    # the keyword table sits two columns past where the positional fallback
    # (known_max + 3) would look, so only content detection lands on it
    header = TX_HEADER + [None] * 6
    wb = Workbook()
    wb.remove(_active(wb))
    ws = wb.create_sheet(spec.SHEET_TRANSACTIONS)
    ws.append(header)
    ws.append([None] * 8 + [None, None, None, None, "Groceries", "lenta|okey"])
    ws.append([None] * 8 + [None, None, None, None, "Cafes", "coffee|bar"])
    idx = _tx_header_index(ws)
    assert idx is not None
    kws = _parse_keywords(ws, idx)
    assert kws == {"Groceries": "lenta|okey", "Cafes": "coffee|bar"}


# --- _category_col --------------------------------------------------------


def test_category_col_prefers_the_fuller_of_two_candidate_columns() -> None:
    header = TX_HEADER + [None, None, "Category", "Keywords"]
    wb = Workbook()
    wb.remove(_active(wb))
    ws = wb.create_sheet(spec.SHEET_TRANSACTIONS)
    ws.append(header)
    # first candidate column half-filled, second fully filled: the second wins
    ws.append([None] * 8 + ["auto", "Groceries", "Groceries", "lenta|okey"])
    ws.append([None] * 8 + [None, "Cafes", "Cafes", "coffee|bar"])
    idx = _tx_header_index(ws)
    assert idx is not None
    start = _known_max_col(idx) + 1
    col = _category_col(ws, idx)
    assert col == start + 1


def test_find_keyword_block_ignores_columns_without_pipes() -> None:
    header = TX_HEADER + [None, None, "Category", "Keywords"]
    wb = Workbook()
    wb.remove(_active(wb))
    ws = wb.create_sheet(spec.SHEET_TRANSACTIONS)
    ws.append(header)
    ws.append([None] * 8 + [None, None, "Groceries", "lenta|okey"])
    idx = _tx_header_index(ws)
    assert idx is not None
    base = _find_keyword_block(ws, idx)
    assert base == _known_max_col(idx) + 3


# --- end to end guards ----------------------------------------------------


def test_parse_workbook_reads_a_stated_category_sheet() -> None:
    wb = Workbook()
    wb.remove(_active(wb))
    cats = wb.create_sheet(spec.SHEET_CATEGORIES)
    for row in (
        ["Daily", 1, "OUT"],
        [1, "Daily", "Groceries", "lenta|okey"],
    ):
        cats.append(row)
    tx = wb.create_sheet(spec.SHEET_TRANSACTIONS)
    tx.append(TX_HEADER + [None, "Monori Category"])
    _plain_tx(datetime.datetime(2025, 1, 15), -300.0, "Groceries", desc="Lenta").append_to(tx)
    parsed = parse_workbook(_save(wb))
    assert [(group.name, group.sort, group.kind) for group in parsed.groups] == [
        ("Daily", 1, "expense")
    ]
    groceries = next(category for category in parsed.categories if category.name == "Groceries")
    assert groceries.keywords == "lenta|okey"

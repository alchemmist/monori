import os
import sqlite3
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.importer import tx_hash
from tests.conftest import Api

pytestmark = pytest.mark.integration

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _export(client: TestClient) -> Workbook:
    r = client.get("/api/export/xlsx")
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith(XLSX_MIME)
    assert "monori-export.xlsx" in r.headers["content-disposition"]
    return load_workbook(BytesIO(r.content))


def _setup(api: Api, client: TestClient) -> tuple[int, int]:
    g_out = api.group("Daily Expenses")
    g_in = api.group("Inflow", kind="income")
    cat = api.category("Groceries", g_out, keywords="lenta|okey")
    salary = api.category("Salary", g_in)
    acct = api.account("Card")
    api.tx("2026-01-05T10:00:00", -12550, accountId=acct, categoryId=cat, description="Lenta")
    api.tx("2026-01-10T09:00:00", 500000, accountId=acct, categoryId=salary, description="Pay")
    client.put("/api/budgets", json={"categoryId": cat, "year": 2026, "month": 1, "amount": 20000})
    return cat, acct


def test_export_sheet_structure(api: Api, client: TestClient) -> None:
    _setup(api, client)
    wb = _export(client)
    assert wb.sheetnames == ["Categories", "Transactions", "2026", "DashData"]


def test_export_categories_sheet(api: Api, client: TestClient) -> None:
    _setup(api, client)
    ws = _export(client)["Categories"]
    assert [c.value for c in ws[1]] == ["Sort Order", "Category Group", "Category", "Keywords"]
    rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    assert [1, "▼Daily Expenses", "Groceries", "lenta|okey"] in rows
    assert ["▼Daily Expenses", 1, "OUT"] in [r[:3] for r in rows]
    assert ["▲Inflow", 2, "IN"] in [r[:3] for r in rows]


def test_export_transactions_sheet(api: Api, client: TestClient) -> None:
    _setup(api, client)
    ws = _export(client)["Transactions"]
    headers = [c.value for c in ws[1]]
    assert headers[0] == "Дата операции"
    assert headers[-3:] == ["Monori Category", "Account", "Comment"]
    row = [c.value for c in ws[2]]
    assert row[0] == "05.01.2026 10:00:00"
    assert row[1] == "05.01.2026"
    assert row[4] == "OK"
    assert row[5] == -125.50
    assert row[7] == "-125.50 ₽"
    assert row[12] == "Lenta"
    assert row[13] == "Groceries"
    assert row[14] == "Card"


def test_export_uses_split_parts_for_rows_and_totals(api: Api, client: TestClient) -> None:
    expenses = api.group("Expenses")
    groceries = api.category("Groceries", expenses)
    household = api.category("Household", expenses)
    account = api.account("Card")
    tx = api.tx(
        "2026-01-05T10:00:00",
        -10000,
        accountId=account,
        description="Mixed receipt",
    )
    response = client.put(
        f"/api/transactions/{tx}/splits",
        json={
            "parts": [
                {"categoryId": groceries, "amount": -6000, "comment": "food"},
                {"categoryId": household, "amount": -4000, "comment": "soap"},
            ]
        },
    )
    assert response.status_code == 200

    wb = _export(client)
    rows = [[cell.value for cell in row] for row in wb["Transactions"].iter_rows(min_row=2)]
    assert [(row[5], row[13], row[15]) for row in rows] == [
        (-60, "Groceries", "food"),
        (-40, "Household", "soap"),
    ]
    dash = [cell.value for cell in wb["DashData"][2]]
    assert dash[2] == 100

    year = wb["2026"]
    activity = {
        year.cell(row=row, column=1).value: year.cell(row=row, column=3).value
        for row in range(1, year.max_row + 1)
    }
    assert activity["Groceries"] == 60
    assert activity["Household"] == 40


def test_export_year_sheet_grid(api: Api, client: TestClient) -> None:
    _setup(api, client)
    ws = _export(client)["2026"]
    assert ws.cell(row=1, column=2).value == "January"
    assert [ws.cell(row=2, column=c).value for c in (2, 3, 4)] == [
        "Budgeted",
        "Outflows",
        "Balance",
    ]
    labels = {ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)}
    assert "Month Summary" in labels
    assert "▼Daily Expenses" in labels
    assert "Groceries" in labels
    groceries_row = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Groceries"
    )
    assert ws.cell(row=groceries_row, column=2).value == 200.0
    assert ws.cell(row=groceries_row, column=3).value == 125.5
    assert ws.cell(row=groceries_row, column=4).value == 74.5
    assert ws.cell(row=3, column=2).value == 200.0
    assert ws.cell(row=3, column=3).value == 125.5


def test_export_dashdata_sheet(api: Api, client: TestClient) -> None:
    _setup(api, client)
    ws = _export(client)["DashData"]
    assert [c.value for c in ws[1]] == ["Month", "Income", "Expense", "Ratio", "CumNet"]
    row = [c.value for c in ws[2]]
    assert row[0] == "2026-01"
    assert row[1] == 5000.0
    assert row[2] == 125.5
    assert row[3] == round(125.5 / 5000.0, 2)
    assert row[4] == 4874.5
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=3)]
    header_idx = next(i for i, r in enumerate(rows) if r[0] == "Category")
    assert rows[header_idx][1] == 2026
    by_cat = {r[0]: r[1] for r in rows[header_idx + 1 :]}
    assert by_cat["Groceries"] == -125.5
    assert by_cat["Salary"] == 5000.0


def test_export_excludes_transfers_from_dashdata(api: Api, client: TestClient) -> None:
    _setup(api, client)
    a2 = api.account("Second")
    api.transfer(api.snapshot()["accounts"][0]["id"], a2, 10000, date="2026-01-15T12:00:00")
    ws = _export(client)["DashData"]
    row = [c.value for c in ws[2]]
    assert row[1] == 5000.0
    assert row[2] == 125.5


def test_export_transactions_static_columns(api: Api, client: TestClient) -> None:
    _setup(api, client)
    ws = _export(client)["Transactions"]
    row = [c.value for c in ws[2]]
    assert row[2] == "05.01.2026"
    assert row[3] is None or row[3] == ""
    assert row[6] == "RUB"
    assert row[8] == "RUB"
    assert row[9] is None or row[9] == ""
    assert ws.cell(row=2, column=6).number_format == "0.00"


def test_export_categories_layout(api: Api, client: TestClient) -> None:
    _setup(api, client)
    ws = _export(client)["Categories"]
    assert ws.freeze_panes == "A2"
    assert all(c.font.bold for c in ws[1])
    assert [c.value for c in ws[2]] == [1, "▼Daily Expenses", "Groceries", "lenta|okey"]
    assert [c.value for c in ws[3]] == [2, "▲Inflow", "Salary", None]
    assert ws.cell(row=4, column=1).value is None
    assert [c.value for c in ws[5]][:3] == ["Category Group", "Sort Order", "Type"]
    assert ws.cell(row=5, column=1).font.bold


def test_export_year_sheet_layout(api: Api, client: TestClient) -> None:
    _setup(api, client)
    ws = _export(client)["2026"]
    assert ws.freeze_panes == "B3"
    months = [ws.cell(row=1, column=2 + m * 3).value for m in range(12)]
    assert months == [
        "January",
        "February",
        "March",
        "April",
        "May",
        "June",
        "July",
        "August",
        "September",
        "October",
        "November",
        "December",
    ]
    merged = {str(r) for r in ws.merged_cells.ranges}
    assert "B1:D1" in merged
    assert "AI1:AK1" in merged
    assert [ws.cell(row=2, column=c).value for c in (35, 36, 37)] == [
        "Budgeted",
        "Outflows",
        "Balance",
    ]
    assert ws.cell(row=1, column=38).value == "Total"
    assert ws.cell(row=1, column=39).value == "Average"
    assert ws.column_dimensions["A"].width == 24
    assert ws.column_dimensions["B"].width == 11
    assert ws.cell(row=1, column=2).font.bold
    assert ws.cell(row=3, column=1).font.bold


def test_export_year_sheet_totals(api: Api, client: TestClient) -> None:
    _setup(api, client)
    ws = _export(client)["2026"]
    groceries_row = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Groceries"
    )
    assert ws.cell(row=groceries_row, column=38).value == 125.5
    assert ws.cell(row=groceries_row, column=39).value == 10.46
    assert ws.cell(row=groceries_row, column=5).value == 0.0
    assert ws.cell(row=groceries_row, column=6).value == 0.0
    assert ws.cell(row=groceries_row, column=7).value == 74.5
    assert ws.cell(row=3, column=38).value == 125.5
    assert ws.cell(row=3, column=39).value == 10.46
    salary_row = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Salary"
    )
    assert ws.cell(row=salary_row, column=3).value == 5000.0


def test_export_escapes_at_prefix(api: Api, client: TestClient) -> None:
    cat, acct = _setup(api, client)
    api.tx("2026-03-01T10:00:00", -100, accountId=acct, categoryId=cat, description="@cmd|test")
    ws = _export(client)["Transactions"]
    descriptions = {ws.cell(row=r, column=13).value for r in range(2, ws.max_row + 1)}
    assert "'@cmd|test" in descriptions


def test_export_dashdata_freeze_and_bold(api: Api, client: TestClient) -> None:
    _setup(api, client)
    ws = _export(client)["DashData"]
    assert ws.freeze_panes == "A2"
    assert all(c.font.bold for c in ws[1])


def test_export_amount_uses_account_currency_symbol(api: Api, client: TestClient) -> None:
    cat, _ = _setup(api, client)
    usd = api.account("Dollars", currency="USD")
    eur = api.account("Euros", currency="EUR")
    chf = api.account("Francs", currency="CHF")
    api.tx("2026-04-01T10:00:00", -30000, accountId=usd, categoryId=cat, description="Hotel")
    api.tx("2026-04-02T10:00:00", -20000, accountId=eur, categoryId=cat, description="Train")
    api.tx("2026-04-03T10:00:00", -10000, accountId=chf, categoryId=cat, description="Cheese")
    ws = _export(client)["Transactions"]
    amounts = {ws.cell(row=r, column=8).value for r in range(2, ws.max_row + 1)}
    assert "-125.50 ₽" in amounts
    assert "-300.00 $" in amounts
    assert "-200.00 €" in amounts
    assert "-100.00 CHF" in amounts


def test_export_dashdata_skips_uncategorized(api: Api, client: TestClient) -> None:
    _setup(api, client)
    api.tx("2026-01-25T10:00:00", -99900, description="Mystery")
    ws = _export(client)["DashData"]
    row = [c.value for c in ws[2]]
    assert row[1] == 5000.0
    assert row[2] == 125.5


def test_export_requires_auth(anon: TestClient) -> None:
    r = anon.get("/api/export/xlsx")
    assert r.status_code == 401


def test_export_empty_user(client: TestClient) -> None:
    wb = _export(client)
    assert wb.sheetnames == ["Categories", "Transactions", "DashData"]
    assert [c.value for c in wb["Categories"][1]] == [
        "Sort Order",
        "Category Group",
        "Category",
        "Keywords",
    ]
    assert wb["Transactions"].cell(row=1, column=1).value == "Дата операции"
    assert [c.value for c in wb["DashData"][1]] == ["Month", "Income", "Expense", "Ratio", "CumNet"]


def test_export_escapes_formula_prefixes(api: Api, client: TestClient) -> None:
    cat, acct = _setup(api, client)
    api.tx(
        "2026-02-01T10:00:00",
        -100,
        accountId=acct,
        categoryId=cat,
        description="=HYPERLINK(evil)",
        comment="+SUM(A1)",
    )
    ws = _export(client)["Transactions"]
    descriptions = {ws.cell(row=r, column=13).value for r in range(2, ws.max_row + 1)}
    comments = {ws.cell(row=r, column=16).value for r in range(2, ws.max_row + 1)}
    assert "'=HYPERLINK(evil)" in descriptions
    assert "'+SUM(A1)" in comments


def test_export_header_band_is_slate(api: Api, client: TestClient) -> None:
    _setup(api, client)
    wb = _export(client)
    for name in ("Categories", "Transactions", "DashData", "2026"):
        head = wb[name].cell(row=1, column=1)
        assert head.fill.fgColor.rgb == "FF3C464D"
        assert head.font.color.rgb == "FFFFFFFF"
        assert head.font.bold


def test_export_year_sheet_bands(api: Api, client: TestClient) -> None:
    _setup(api, client)
    ws = _export(client)["2026"]
    assert ws.cell(row=1, column=2).fill.fgColor.rgb == "FF3C464D"
    assert ws.cell(row=2, column=2).fill.fgColor.rgb == "FF3C464D"
    assert ws.cell(row=3, column=2).fill.fgColor.rgb == "FFEEF5E7"
    group_row = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "▼Daily Expenses"
    )
    assert ws.cell(row=group_row, column=1).fill.fgColor.rgb == "FFE6F4FB"


def test_export_summary_balance_is_colored(api: Api, client: TestClient) -> None:
    _setup(api, client)
    ws = _export(client)["2026"]
    balance = ws.cell(row=3, column=4)
    assert balance.value == 74.5
    assert balance.fill.fgColor.rgb == "FFEEF5E7"
    assert balance.font.color.rgb == "FF4F7A00"


def test_export_money_cells_have_grid_border(api: Api, client: TestClient) -> None:
    _setup(api, client)
    ws = _export(client)["2026"]
    groceries_row = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Groceries"
    )
    cell = ws.cell(row=groceries_row, column=2)
    assert cell.border.left.style == "thin"
    assert cell.border.bottom.style == "thin"


def test_export_positive_balance_is_green(api: Api, client: TestClient) -> None:
    _setup(api, client)
    ws = _export(client)["2026"]
    groceries_row = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Groceries"
    )
    balance = ws.cell(row=groceries_row, column=4)
    assert balance.value == 74.5
    assert balance.font.color.rgb == "FF4F7A00"


def test_export_negative_balance_is_red(api: Api, client: TestClient) -> None:
    g_out = api.group("Overspend")
    cat = api.category("Splurge", g_out)
    acct = api.account("Card")
    api.tx("2026-01-05T10:00:00", -20000, accountId=acct, categoryId=cat, description="Big")
    client.put("/api/budgets", json={"categoryId": cat, "year": 2026, "month": 1, "amount": 10000})
    ws = _export(client)["2026"]
    splurge_row = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Splurge"
    )
    balance = ws.cell(row=splurge_row, column=4)
    assert balance.value == -100.0
    assert balance.font.color.rgb == "FFC0392B"


def test_export_zero_balance_is_grey(api: Api, client: TestClient) -> None:
    g_out = api.group("OnBudget")
    cat = api.category("Exact", g_out)
    acct = api.account("Card")
    api.tx("2026-01-05T10:00:00", -10000, accountId=acct, categoryId=cat, description="Spend")
    client.put("/api/budgets", json={"categoryId": cat, "year": 2026, "month": 1, "amount": 10000})
    ws = _export(client)["2026"]
    exact_row = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Exact"
    )
    balance = ws.cell(row=exact_row, column=4)
    assert balance.value == 0.0
    assert balance.font.color.rgb == "FF434343"


def test_export_dashdata_refund_reduces_expense(api: Api, client: TestClient) -> None:
    """
    Direction enforcement keeps the API from filing an inflow into an expense
    category, but migrated workbooks and old synced statements carry such
    refund rows — DashData must still net them out of the category's spend.
    """
    cat, acct = _setup(api, client)
    date, amount, desc = "2026-01-20T10:00:00", 2550, "Refund"
    db = sqlite3.connect(os.environ["MONORI_DB"])
    db.execute(
        "INSERT INTO transactions"
        " (date, amount, description, account_id, category_id, hash, source)"
        " VALUES (?, ?, ?, ?, ?, ?, 'sync')",
        (date, amount, desc, acct, cat, tx_hash(acct, date, amount, desc)),
    )
    db.commit()
    db.close()
    ws = _export(client)["DashData"]
    row = [c.value for c in ws[2]]
    assert row[1] == 5000.0
    assert row[2] == 100.0

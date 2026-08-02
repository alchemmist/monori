"""
Export tests over a dataset with several transactions per category, more than.

one month and a month without income. The single-row fixtures elsewhere cannot.
tell an accumulation (`+=`) from an assignment (`=`), a three-wide month stride
from a four-wide one, or an empty ratio from a filled one — this one can.
"""

from dataclasses import dataclass
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from tests.conftest import AccountOptions, Api, TransactionOptions

pytestmark = pytest.mark.integration


def _export(client: TestClient) -> Workbook:
    r = client.get("/api/export/xlsx")
    assert r.status_code == 200, r.text
    return load_workbook(BytesIO(r.content))


def _rich(api: Api, client: TestClient) -> dict[str, int]:
    daily = api.group("Daily")
    inflow = api.group("Inflow", kind="income")
    groceries = api.category("Groceries", daily)
    cafes = api.category("Cafes", daily)
    salary = api.category("Salary", inflow)
    acct = api.account(
        "Card",
        AccountOptions(
            account_type="card",
            icon="wallet",
            color="#5b6472",
            bank_ref="card-1",
            currency="RUB",
            opening_balance=0,
        ),
    )

    api.tx(
        "2026-01-05T10:00:00",
        -10000,
        TransactionOptions(account_id=acct, category_id=groceries, description="G1"),
    )
    api.tx(
        "2026-01-06T10:00:00",
        -5000,
        TransactionOptions(account_id=acct, category_id=groceries, description="G2"),
    )
    api.tx(
        "2026-01-07T10:00:00",
        -3000,
        TransactionOptions(account_id=acct, category_id=cafes, description="C1"),
    )
    api.tx(
        "2026-01-10T10:00:00",
        500000,
        TransactionOptions(account_id=acct, category_id=salary, description="S1"),
    )
    api.tx(
        "2026-01-11T10:00:00",
        100000,
        TransactionOptions(account_id=acct, category_id=salary, description="S2"),
    )

    api.tx(
        "2026-02-05T10:00:00",
        -20000,
        TransactionOptions(account_id=acct, category_id=groceries, description="G3"),
    )

    api.tx(
        "2026-03-05T10:00:00",
        -4000,
        TransactionOptions(account_id=acct, category_id=cafes, description="C2"),
    )
    for month, cat, amount in (
        (1, groceries, 30000),
        (1, cafes, 20000),
        (2, groceries, 40000),
    ):
        client.put(
            "/api/budgets",
            json={"categoryId": cat, "year": 2026, "month": month, "amount": amount},
        )
    return {"groceries": groceries, "cafes": cafes, "salary": salary, "acct": acct}


def _row_of(ws: Worksheet, label: str) -> int:
    return next(r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == label)


def test_year_sheet_sums_repeated_category_rows(api: Api, client: TestClient) -> None:
    _rich(api, client)
    ws = _export(client)["2026"]
    row = _row_of(ws, "Groceries")

    assert ws.cell(row=row, column=3).value == 150.0


def test_month_summary_sums_across_categories(api: Api, client: TestClient) -> None:
    _rich(api, client)
    ws = _export(client)["2026"]

    assert ws.cell(row=3, column=2).value == 500.0


def test_year_sheet_uses_a_three_wide_month_stride(api: Api, client: TestClient) -> None:
    _rich(api, client)
    ws = _export(client)["2026"]
    row = _row_of(ws, "Groceries")

    assert ws.cell(row=row, column=5).value == 400.0
    assert ws.cell(row=row, column=6).value == 200.0


def test_year_sheet_money_cells_keep_the_number_font(api: Api, client: TestClient) -> None:
    _rich(api, client)
    ws = _export(client)["2026"]
    row = _row_of(ws, "Groceries")
    assert ws.cell(row=row, column=2).font.color.rgb == "FF434343"


@dataclass(frozen=True)
class DashRow:
    month: str
    income: float
    expenses: float
    ratio: float | None
    cumulative: float


def _dash_row(ws: Worksheet, month: str) -> DashRow:
    idx = next(r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value == month)
    values = [ws.cell(row=idx, column=column).value for column in range(1, 6)]
    row_month, income, expenses, ratio, cumulative = values
    assert isinstance(row_month, str)
    assert isinstance(income, int | float)
    assert isinstance(expenses, int | float)
    assert isinstance(ratio, int | float) or ratio is None
    assert isinstance(cumulative, int | float)
    return DashRow(
        row_month,
        float(income),
        float(expenses),
        None if ratio is None else float(ratio),
        float(cumulative),
    )


def test_dashdata_sums_income_across_rows(api: Api, client: TestClient) -> None:
    _rich(api, client)
    ws = _export(client)["DashData"]
    jan = _dash_row(ws, "2026-01")

    assert jan.income == 6000.0


def test_dashdata_carries_a_running_cumulative_net(api: Api, client: TestClient) -> None:
    _rich(api, client)
    ws = _export(client)["DashData"]

    assert _dash_row(ws, "2026-01").cumulative == 5820.0
    assert _dash_row(ws, "2026-02").cumulative == 5620.0


def test_dashdata_leaves_the_ratio_blank_without_income(api: Api, client: TestClient) -> None:
    _rich(api, client)
    ws = _export(client)["DashData"]
    march = _dash_row(ws, "2026-03")
    assert march.expenses == 40.0
    assert march.ratio is None


def test_transactions_sheet_is_frozen_below_the_header(api: Api, client: TestClient) -> None:
    _rich(api, client)
    ws = _export(client)["Transactions"]
    assert ws.freeze_panes == "A2"

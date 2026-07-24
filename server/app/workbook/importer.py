"""Parses a YNAB-style workbook (the format written by ``export.py``) back
into monori data. The sheet layout contract lives in ``spec.py`` and is shared
with the exporter so the two directions never drift.
"""

import datetime
import re
from io import BytesIO

from openpyxl import load_workbook

from ..importer import parse_amount_kop, parse_date, tx_hash
from . import spec

YEAR_SHEET_RE = re.compile(r"^\d{4}$")

TX_REQUIRED = ["Дата операции", "Status", "Сумма операции", "Категория", "MCC", "Description"]


class WorkbookError(Exception):
    pass


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _unquote(value: str) -> str:
    """Reverses the exporter's formula-escape and nothing else: a leading
    apostrophe is stripped only when it guards a formula prefix, so legitimate
    values that happen to start with an apostrophe survive the round-trip."""
    if value.startswith("'") and value[1:].startswith(("=", "+", "@")):
        return value[1:]
    return value


def _parse_dt_cell(value):
    if isinstance(value, datetime.datetime):
        return value
    if isinstance(value, datetime.date):
        return datetime.datetime(value.year, value.month, value.day)
    s = _cell_str(value)
    if not s:
        return None
    parsed = parse_date(s)
    if parsed:
        return parsed
    try:
        return datetime.datetime.fromisoformat(s)
    except ValueError:
        return None


def _parse_amount_cell(value):
    if isinstance(value, int | float):
        return spec.kop_from_rub(value)
    s = _cell_str(value)
    return parse_amount_kop(s) if s else None


def _parse_categories(ws, warnings):
    groups = []
    categories = []
    group_rows_seen = False
    for row in ws.iter_rows(min_row=1, values_only=True):
        cells = list(row) + [None] * (4 - len(row))
        c1, c2, c3, c4 = cells[:4]
        s1, s2, s3 = _cell_str(c1), _cell_str(c2), _cell_str(c3)
        if s1 in ("Sort Order", "Category Group") or (not s1 and not s2):
            continue
        if s3 in (spec.TYPE_IN, spec.TYPE_OUT) and isinstance(c2, int | float):
            name, kind = spec.strip_glyph(_unquote(s1))
            groups.append(
                {
                    "name": name,
                    "sort": int(c2),
                    "kind": "income" if s3 == spec.TYPE_IN else "expense",
                }
            )
            group_rows_seen = True
            continue
        if isinstance(c1, int | float) and s2 and s3:
            name, kind = spec.strip_glyph(_unquote(s2))
            categories.append(
                {
                    "group": name,
                    "group_kind": kind,
                    "group_sort": int(c1),
                    "name": _unquote(s3),
                    "keywords": _unquote(_cell_str(c4)),
                }
            )
            continue
        if s1 or s2 or s3:
            warnings.append(f"Categories: unrecognized row skipped: {[s1, s2, s3][:3]}")
    if not group_rows_seen:
        seen: dict[str, dict] = {}
        for cat in categories:
            if str(cat["group"]) not in seen:
                seen[str(cat["group"])] = {
                    "name": cat["group"],
                    "sort": cat["group_sort"],
                    "kind": cat["group_kind"] or "expense",
                }
        groups = list(seen.values())
        if groups:
            warnings.append("Categories: group table missing, groups derived from category rows")
    return groups, categories


def _header_index(ws):
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header is None:
        return None
    return {_cell_str(v): i for i, v in enumerate(header) if _cell_str(v)}


def _parse_transactions(ws, warnings):
    idx = _header_index(ws)
    if idx is None or any(h not in idx for h in TX_REQUIRED):
        missing = [h for h in TX_REQUIRED if idx is None or h not in idx]
        raise WorkbookError(f"Transactions sheet is missing required columns: {missing}")

    def col(row, name, default=""):
        i = idx.get(name)
        if i is None or i >= len(row):
            return default
        return row[i]

    rows = []
    skipped_status = 0
    errors = []
    for n, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or _cell_str(v) == "" for v in row):
            continue
        status = _cell_str(col(row, "Status"))
        if status and status.upper() not in ("OK", ""):
            skipped_status += 1
            continue
        dt = _parse_dt_cell(col(row, "Дата операции"))
        amount = _parse_amount_cell(col(row, "Сумма операции"))
        if dt is None or amount is None:
            errors.append({"row": n, "error": "unparseable date or amount"})
            continue
        card = _unquote(_cell_str(col(row, "Номер карты")))
        account = _unquote(_cell_str(col(row, "Account")))
        date_iso = dt.strftime("%Y-%m-%dT%H:%M:%S")
        description = _unquote(_cell_str(col(row, "Description")))
        rows.append(
            {
                "date": date_iso,
                "amount": amount,
                "description": description,
                "bank_category": _unquote(_cell_str(col(row, "Категория"))),
                "mcc": _unquote(_cell_str(col(row, "MCC"))),
                "comment": _unquote(_cell_str(col(row, "Comment"))),
                "monori_category": _unquote(_cell_str(col(row, "Monori Category"))),
                "marker": card or account,
                "hash": tx_hash(date_iso, amount, description),
            }
        )
    if skipped_status:
        warnings.append(f"Transactions: {skipped_status} non-OK rows skipped")
    return rows, errors


def _parse_year_sheet(ws, year, category_names, warnings):
    cells = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        label = _unquote(_cell_str(row[0] if row else None))
        if not label or label == "Month Summary":
            continue
        if label not in category_names:
            _, kind = spec.strip_glyph(label)
            if kind is None:
                warnings.append(f"{year}: unknown row label skipped: {label[:60]}")
            continue
        for m in range(1, 13):
            i = 1 + (m - 1) * 3
            value = row[i] if i < len(row) else None
            if value in (None, ""):
                continue
            kop = _parse_amount_cell(value)
            if kop:
                cells.append({"category": label, "year": year, "month": m, "amount": kop})
    return cells


def parse_workbook(data: bytes):
    """Returns {groups, categories, transactions, budgets, warnings, errors}."""
    try:
        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:
        raise WorkbookError(f"not a readable .xlsx workbook: {exc}") from exc
    try:
        warnings: list[str] = []
        if spec.SHEET_CATEGORIES not in wb.sheetnames:
            raise WorkbookError(f"missing required sheet: {spec.SHEET_CATEGORIES}")
        if spec.SHEET_TRANSACTIONS not in wb.sheetnames:
            raise WorkbookError(f"missing required sheet: {spec.SHEET_TRANSACTIONS}")
        groups, categories = _parse_categories(wb[spec.SHEET_CATEGORIES], warnings)
        transactions, errors = _parse_transactions(wb[spec.SHEET_TRANSACTIONS], warnings)
        category_names = {c["name"] for c in categories}
        budgets = []
        for name in wb.sheetnames:
            if YEAR_SHEET_RE.match(name):
                budgets.extend(_parse_year_sheet(wb[name], int(name), category_names, warnings))
        known = {
            spec.SHEET_CATEGORIES,
            spec.SHEET_TRANSACTIONS,
            spec.SHEET_DASHDATA,
        }
        for name in wb.sheetnames:
            if name not in known and not YEAR_SHEET_RE.match(name):
                warnings.append(f"unknown sheet ignored: {name}")
        return {
            "groups": groups,
            "categories": categories,
            "transactions": transactions,
            "budgets": budgets,
            "warnings": warnings,
            "errors": errors,
        }
    finally:
        wb.close()

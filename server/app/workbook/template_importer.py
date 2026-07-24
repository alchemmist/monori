"""
Parses the original live "Budget YNAB-Like" Google-Sheets template (the
spreadsheet monori grew from) into the same contract as
``importer.parse_workbook``.

The live template cannot be replayed literally: transactions exist only for
recent months, earlier history survives only as hand-written ``+N``
corrections inside formulas, and the transaction sheet contains duplicated
blocks with broken amounts. So instead of re-running the sheet's logic this
parser reproduces the sheet's *cached numbers*: real transactions are
imported as-is, and per category/month differences between the sheet's cached
aggregates and what monori would compute are closed with synthetic
"Migration" transactions. After migration the user sees the same budgeted /
outflows / balance / available figures the spreadsheet showed.
"""

import datetime
import re
from io import BytesIO

from openpyxl import load_workbook

from ..importer import tx_hash
from . import spec

YEAR_RE = re.compile(r"^(\d{4})(_archive)?$")

RU_HEADERS = {
    "date": "Дата операции",
    "card": "Номер карты",
    "status": "Статус",
    "amount": "Сумма операции",
    "currency": "Валюта операции",
    "bank_category": "Категория",
    "mcc": "MCC",
    "description": "Описание",
}

MONTH_TOKENS = {
    "ЯНВ": 1,
    "ФЕВ": 2,
    "МАР": 3,
    "АПР": 4,
    "МАЙ": 5,
    "МАЯ": 5,
    "ИЮН": 6,
    "ИЮЛ": 7,
    "АВГ": 8,
    "СЕН": 9,
    "ОКТ": 10,
    "НОЯ": 11,
    "ДЕК": 12,
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}

PAY_AMOUNT_HEADER = "Сумма платежа"
PAY_CURRENCY_HEADER = "Валюта платежа"

BUDGET_HEADERS = ("Бюджет", "Budgeted")
OUTFLOW_HEADERS = ("Расход", "Outflows")
BALANCE_HEADERS = ("Баланс", "Balance")
LABEL_HEADERS = ("Категория", "Категории", "Category", "Categories")
SKIP_LABELS = LABEL_HEADERS + ("Month Summary", "Total", "Итого")

INCOME_GROUP = "Inflow"
INCOME_CATEGORY = "Income"

ADJUST_TOLERANCE_KOP = 2
VERIFY_TOLERANCE_KOP = 5
MAX_VERIFY_WARNINGS = 12


class TemplateError(Exception):
    pass


def _s(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _kop(value):
    if isinstance(value, bool):
        return None
    if isinstance(value, str):
        cleaned = re.sub(r"[\s\u00a0\u202f]", "", value).replace(",", ".")
        if not cleaned:
            return None
        try:
            value = float(cleaned)
        except ValueError:
            return None
    if not isinstance(value, int | float):
        return None
    return spec.kop_from_rub(value)


def _last_day(year, month):
    end = datetime.date(year + 1, 1, 1) if month == 12 else datetime.date(year, month + 1, 1)
    return end - datetime.timedelta(days=1)


def _stamp(year, month):
    return _last_day(year, month).strftime("%Y-%m-%dT12:00:00")


def is_template_transactions_header(header_cells) -> bool:
    names = {_s(v) for v in header_cells}
    return RU_HEADERS["status"] in names and "Status" not in names


def looks_like_template(data: bytes) -> bool:
    try:
        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    except Exception:
        return False
    try:
        if spec.SHEET_TRANSACTIONS not in wb.sheetnames:
            return False
        ws = wb[spec.SHEET_TRANSACTIONS]
        if not hasattr(ws, "iter_rows"):
            return False
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        return header is not None and is_template_transactions_header(header)
    finally:
        wb.close()


def _month_token(value):
    token = _s(value).upper()[:3]
    return MONTH_TOKENS.get(token)


def _find_layout(ws):
    """
    Locates the month blocks of a year sheet. Returns None when the sheet has
    no recognizable Budgeted/Outflows/Balance header row.
    """
    for r in range(5, 11):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        bases = [i + 1 for i, v in enumerate(row) if _s(v) in BUDGET_HEADERS]
        if len(bases) < 2:
            continue
        out_off = bal_off = None
        for i, v in enumerate(row):
            col = i + 1
            if col <= bases[0]:
                continue
            if _s(v) in OUTFLOW_HEADERS and out_off is None:
                out_off = col - bases[0]
            if _s(v) in BALANCE_HEADERS and bal_off is None:
                bal_off = col - bases[0]
        if out_off is None or bal_off is None:
            continue
        label_col = None
        for rr in (r, r + 1):
            for c in range(1, bases[0]):
                if _s(ws.cell(rr, c).value) in LABEL_HEADERS:
                    label_col = c
                    break
            if label_col:
                break
        start_month = None
        for rr in (1, 2, 3):
            token = _month_token(ws.cell(rr, bases[0]).value)
            if token:
                start_month = token
                break
        return {
            "header_row": r,
            "bases": bases,
            "out_off": out_off,
            "bal_off": bal_off,
            "label_col": label_col or 3,
            "start_month": start_month or 1,
        }
    return None


def _sheet_sections(ws, layout):
    """
    Splits the category area into (group, [(row, category), ...]) sections.
    A row whose label starts with a kind glyph opens a group; in the old
    glyph-less layout the first labelled row after a fully blank gap does.
    """
    label_col = layout["label_col"]
    sections: list[dict] = []
    current: dict | None = None
    in_gap = True
    for r in range(layout["header_row"] + 1, ws.max_row + 1):
        label = _s(ws.cell(r, label_col).value)
        if not label or label in SKIP_LABELS:
            in_gap = in_gap or not label
            if label in SKIP_LABELS:
                in_gap = True
            continue
        name, kind = spec.strip_glyph(label)
        if kind is not None or (in_gap and current is None) or (in_gap and current is not None):
            current = {"name": name, "kind": kind or "expense", "rows": []}
            sections.append(current)
        elif current is None:
            current = {"name": name, "kind": "expense", "rows": []}
            sections.append(current)
        else:
            current["rows"].append((r, label))
        in_gap = False
    return sections


def _summary_value(ws, base, labels):
    for r in range(1, 7):
        text = _s(ws.cell(r, base + 2).value)
        if any(text.startswith(lb) for lb in labels):
            return _kop(ws.cell(r, base + 1).value)
    return None


def _parse_year_sheet(ws, year, layout):
    months = []
    for i, base in enumerate(layout["bases"]):
        m = layout["start_month"] + i
        if m > 12:
            break
        months.append((m, base))
    cats: dict[str, dict] = {}
    for section in _sheet_sections(ws, layout):
        for r, name in section["rows"]:
            entry = cats.setdefault(
                name,
                {"group": section["name"], "budgets": {}, "outflows": {}, "balances": {}},
            )
            for m, base in months:
                b = _kop(ws.cell(r, base).value)
                o = _kop(ws.cell(r, base + layout["out_off"]).value)
                bal = _kop(ws.cell(r, base + layout["bal_off"]).value)
                if b is not None:
                    entry["budgets"][m] = b
                if o is not None:
                    entry["outflows"][m] = o
                if bal is not None:
                    entry["balances"][m] = bal
    income = {}
    available = {}
    for m, base in months:
        inc = _summary_value(ws, base, ("Income for", "Поступления в"))
        if inc is not None:
            income[m] = inc
        for r in (5, 6):
            label = _s(ws.cell(r + 1, base + 1).value)
            if label.startswith(("Available", "Доступный")):
                av = _kop(ws.cell(r, base + 1).value)
                if av is not None:
                    available[m] = av
                break
    seed = None
    first_base = months[0][1] if months else None
    if first_base is not None:
        label = _s(ws.cell(1, first_base + 2).value)
        if label.startswith(("Not budgeted", "Не заложено")):
            seed = _kop(ws.cell(1, first_base + 1).value)
    return {
        "year": year,
        "months": [m for m, _ in months],
        "cats": cats,
        "income": income,
        "available": available,
        "seed": seed,
        "sections": _sheet_sections(ws, layout),
    }


def _tx_header_index(ws):
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header is None:
        return None
    return {_s(v): i for i, v in enumerate(header) if _s(v)}


def _parse_dt(value):
    if isinstance(value, datetime.datetime):
        return value
    if isinstance(value, datetime.date):
        return datetime.datetime(value.year, value.month, value.day)
    return None


def _parse_transactions(ws, warnings, errors):
    idx = _tx_header_index(ws)
    if idx is None:
        raise TemplateError("Transactions sheet is empty")
    missing = [h for h in RU_HEADERS.values() if h not in idx]
    if missing:
        raise TemplateError(f"Transactions sheet is missing required columns: {missing}")

    def col(row, key):
        i = idx[RU_HEADERS[key]]
        return row[i] if i < len(row) else None

    def opt(row, header):
        i = idx.get(header)
        return row[i] if i is not None and i < len(row) else None

    cat_col = max(idx.values()) + 2
    rows = []
    seen = set()
    skipped_status = 0
    non_rub = 0
    dupes = 0
    for n, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or _s(v) == "" for v in row):
            continue
        status = _s(col(row, "status")).upper()
        if status not in ("OK", ""):
            skipped_status += 1
            continue
        dt = _parse_dt(col(row, "date"))
        amount = _kop(col(row, "amount"))
        currency = _s(col(row, "currency"))
        if amount is None:
            amount = _kop(opt(row, PAY_AMOUNT_HEADER))
            if amount is not None:
                currency = _s(opt(row, PAY_CURRENCY_HEADER)) or currency
        description = _s(col(row, "description"))
        if dt is None or amount is None:
            if dt is None and amount is None and not description:
                continue
            errors.append({"row": n, "error": "unparseable date or amount"})
            continue
        if currency and currency != "RUB":
            non_rub += 1
        date_iso = dt.strftime("%Y-%m-%dT%H:%M:%S")
        h = tx_hash(date_iso, amount, description)
        if h in seen:
            dupes += 1
            continue
        seen.add(h)
        category = _s(row[cat_col]) if cat_col < len(row) else ""
        rows.append(
            {
                "date": date_iso,
                "amount": amount,
                "description": description,
                "bank_category": _s(col(row, "bank_category")),
                "mcc": _s(col(row, "mcc")),
                "comment": "",
                "monori_category": category,
                "marker": _s(col(row, "card")),
                "hash": h,
            }
        )
    if dupes:
        warnings.append(f"Transactions: {dupes} duplicated rows collapsed")
    if skipped_status:
        warnings.append(f"Transactions: {skipped_status} non-OK rows skipped")
    if non_rub:
        warnings.append(f"Transactions: {non_rub} non-RUB rows imported with their face value")
    return rows


def _parse_keywords(ws, idx):
    """
    The keyword table sits in the two columns right of the per-row category
    column: category name | pipe-separated keywords, starting at row 1.
    """
    base = max(idx.values()) + 3
    keywords: dict[str, str] = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        if base >= len(row):
            continue
        name = _s(row[base])
        kws = _s(row[base + 1]) if base + 1 < len(row) else ""
        if name and kws and ("|" in kws or len(kws) > 1):
            keywords.setdefault(name, kws)
    return keywords


def _synthetic(year, month, amount, category, description, marker=""):
    date_iso = _stamp(year, month)
    return {
        "date": date_iso,
        "amount": amount,
        "description": description,
        "bank_category": "",
        "mcc": "",
        "comment": "",
        "monori_category": category,
        "marker": marker,
        "hash": tx_hash(date_iso, amount, description),
    }


def _month_range(start, end):
    y, m = start
    while (y, m) <= end:
        yield (y, m)
        m += 1
        if m > 12:
            y, m = y + 1, 1


def parse_template_workbook(data: bytes):
    """
    Returns {groups, categories, transactions, budgets, warnings, errors}
    (the ``parse_workbook`` contract) built from a live YNAB-Like template.
    """
    try:
        wb = load_workbook(BytesIO(data), data_only=True)
    except Exception as exc:
        raise TemplateError(f"not a readable .xlsx workbook: {exc}") from exc
    try:
        return _parse(wb)
    finally:
        wb.close()


def _parse(wb):
    warnings: list[str] = []
    errors: list[dict] = []
    if spec.SHEET_TRANSACTIONS not in wb.sheetnames:
        raise TemplateError(f"missing required sheet: {spec.SHEET_TRANSACTIONS}")
    tx_ws = wb[spec.SHEET_TRANSACTIONS]
    tx_idx = _tx_header_index(tx_ws)
    transactions = _parse_transactions(tx_ws, warnings, errors)
    keywords = _parse_keywords(tx_ws, tx_idx)

    archive_years = {}
    live_years = {}
    plain_sheets = {}
    for name in wb.sheetnames:
        year_match = YEAR_RE.match(name)
        if not year_match:
            continue
        ws = wb[name]
        if not hasattr(ws, "iter_rows"):
            continue
        layout = _find_layout(ws)
        if layout is None:
            warnings.append(f"{name}: unrecognized year sheet layout, ignored")
            continue
        year = int(year_match.group(1))
        parsed = _parse_year_sheet(ws, year, layout)
        if year_match.group(2):
            archive_years[year] = parsed
        else:
            plain_sheets[year] = parsed
    for year, parsed in plain_sheets.items():
        if year not in archive_years:
            live_years[year] = parsed
    if not live_years:
        raise TemplateError("no live year sheets found")

    first_live = min(live_years)
    seam_year = first_live - 1
    seam_sheet = plain_sheets.get(seam_year)

    groups: list[dict] = []
    categories: list[dict] = []
    group_names = set()
    cat_names = {}

    def add_group(name, kind):
        if name in group_names:
            return
        group_names.add(name)
        groups.append({"name": name, "sort": len(groups), "kind": kind})

    def add_category(name, group):
        if name in cat_names:
            return
        cat_names[name] = group
        categories.append(
            {
                "group": group,
                "group_kind": None,
                "group_sort": 0,
                "name": name,
                "keywords": keywords.get(name, ""),
            }
        )

    add_group(INCOME_GROUP, "income")
    add_category(INCOME_CATEGORY, INCOME_GROUP)
    for year in sorted(live_years, reverse=True):
        for section in live_years[year]["sections"]:
            add_group(section["name"], section["kind"])
            for _, name in section["rows"]:
                add_category(name, section["name"])
    for year in sorted(archive_years, reverse=True):
        for section in archive_years[year]["sections"]:
            add_group(section["name"], section["kind"])
            for _, name in section["rows"]:
                add_category(name, section["name"])

    kinds = {}
    group_kind = {g["name"]: g["kind"] for g in groups}
    for cat in categories:
        kinds[cat["name"]] = group_kind.get(cat["group"], "expense")

    budgets = []
    for source in list(archive_years.values()) + list(live_years.values()):
        for name, entry in source["cats"].items():
            for m, amount in entry["budgets"].items():
                if amount:
                    budgets.append(
                        {"category": name, "year": source["year"], "month": m, "amount": amount}
                    )

    tx_sums: dict[tuple, int] = {}
    income_sums: dict[tuple, int] = {}
    for tx in transactions:
        cat = tx["monori_category"]
        if not cat:
            continue
        y, m = int(tx["date"][:4]), int(tx["date"][5:7])
        if kinds.get(cat) == "income":
            income_sums[(y, m)] = income_sums.get((y, m), 0) + tx["amount"]
        else:
            tx_sums[(cat, y, m)] = tx_sums.get((cat, y, m), 0) + tx["amount"]

    synthetic = []
    n_hist = n_adjust = 0
    for source in list(archive_years.values()) + list(live_years.values()):
        year = source["year"]
        live = year in live_years
        label = "Migration adjustment: income" if live else "Migration history: income"
        for m, target in source["income"].items():
            have = income_sums.get((year, m), 0)
            delta = target - have
            if abs(delta) > ADJUST_TOLERANCE_KOP:
                synthetic.append(_synthetic(year, m, delta, INCOME_CATEGORY, label))
                income_sums[(year, m)] = have + delta
                if live:
                    n_adjust += 1
                else:
                    n_hist += 1

    budget_map: dict[tuple, int] = {}
    for cell in budgets:
        key = (cell["category"], cell["year"], cell["month"])
        budget_map[key] = budget_map.get(key, 0) + cell["amount"]

    all_years = sorted(set(archive_years) | set(live_years))
    first_sheet = archive_years.get(all_years[0]) or live_years[all_years[0]]
    start = (all_years[0], min(first_sheet["months"]))
    end = (all_years[-1], 12)
    expense_cats = [c["name"] for c in categories if kinds[c["name"]] != "income"]

    seam_targets = {}
    if seam_sheet is not None:
        last_m = max(seam_sheet["months"])
        for name, entry in seam_sheet["cats"].items():
            bal = entry["balances"].get(last_m)
            if bal is not None:
                seam_targets[name] = bal
    seam_seed = live_years[first_live]["seed"]

    balances: dict[str, int] = {}
    avail = 0
    prev_overspent = 0
    avail_residuals = []
    n_seam = 0
    for y, m in _month_range(start, end):
        income = income_sums.get((y, m), 0)
        budgeted_total = sum(budget_map.get((name, y, m), 0) for name in cat_names)
        avail = avail + prev_overspent + income - budgeted_total
        live = y in live_years
        source = live_years.get(y) or archive_years.get(y)
        sheet_cats = source["cats"] if source is not None else {}
        at_seam = seam_sheet is not None and (y, m) == (seam_year, 12)
        overspent = 0
        for name in expense_cats:
            carry = max(balances.get(name, 0), 0)
            have = tx_sums.get((name, y, m), 0)
            projected = carry + budget_map.get((name, y, m), 0) + have
            entry = sheet_cats.get(name)
            target = None
            if at_seam:
                if name in seam_targets:
                    target = seam_targets[name]
                elif name not in live_years[first_live]["cats"]:
                    target = 0
            elif entry is not None:
                target = entry["balances"].get(m)
                if target is None and m in entry["outflows"]:
                    target = projected - have + entry["outflows"][m]
            elif (
                not live
                and balances.get(name, 0) != 0
                and source is not None
                and m == max(source["months"])
            ):
                target = 0
            delta = 0 if target is None else target - projected
            if abs(delta) > ADJUST_TOLERANCE_KOP:
                if at_seam:
                    desc = f"Migration carry: {name}"
                    n_seam += 1
                elif live:
                    desc = f"Migration adjustment: {name}"
                    n_adjust += 1
                else:
                    desc = f"Migration history: {name}"
                    n_hist += 1
                synthetic.append(_synthetic(y, m, delta, name, desc))
                tx_sums[(name, y, m)] = have + delta
                projected += delta
            balances[name] = projected
            overspent += min(projected, 0)
        if at_seam and seam_seed is not None:
            delta = seam_seed - avail
            if abs(delta) > ADJUST_TOLERANCE_KOP:
                synthetic.append(
                    _synthetic(y, m, delta, INCOME_CATEGORY, "Migration: available seed")
                )
                income_sums[(y, m)] = income_sums.get((y, m), 0) + delta
                avail += delta
                n_seam += 1
        prev_overspent = overspent
        if live and source is not None:
            target_avail = source["available"].get(m)
            if target_avail is not None and abs(target_avail - avail) > VERIFY_TOLERANCE_KOP:
                avail_residuals.append((y, m, target_avail - avail))

    if n_hist:
        warnings.append(f"history: {n_hist} synthetic transactions rebuilt from archive sheets")
    if n_adjust:
        warnings.append(
            f"reconciliation: {n_adjust} adjustment transactions align live months with the sheet"
        )
    if n_seam:
        warnings.append(f"seam: {n_seam} carry corrections at {seam_year}-12")
    for y, m, diff in avail_residuals[:MAX_VERIFY_WARNINGS]:
        warnings.append(f"verify: available {y}-{m:02d} differs by {diff / 100:.2f}")
    if len(avail_residuals) > MAX_VERIFY_WARNINGS:
        warnings.append(
            f"verify: {len(avail_residuals) - MAX_VERIFY_WARNINGS} more available mismatches"
        )

    groups_out = [{"name": g["name"], "sort": i, "kind": g["kind"]} for i, g in enumerate(groups)]
    return {
        "groups": groups_out,
        "categories": categories,
        "transactions": transactions + synthetic,
        "budgets": budgets,
        "warnings": warnings,
        "errors": errors,
    }

"""
Reads a budget workbook — ours or the live "Budget YNAB-Like" Google-Sheets.

spreadsheet monori grew from — into {groups, categories, transactions, budgets}.

There is one pipeline, not one per file we have seen. What a workbook happens to
carry is read off its own content: which columns the transaction sheet names,
whether the category structure is spelled out on a sheet of its own or only
implied by the sections of the year grids, and whether a month has rows in it or
just the totals the sheet cached. So a workbook is never classified; it is
measured, and every stage does the most it can with what is actually there.

The last part matters most. A hand-kept spreadsheet holds real rows only for
recent months and keeps its earlier history as cached aggregates. Every source
row is copied unchanged; when a grid total differs, a separate synthetic
"Migration" correction closes the gap. That preserves hand-maintained formula
adjustments while making budgeted / outflows / balance / available figures
survive the move.
"""

import datetime
import re
from collections.abc import Iterable, Mapping
from dataclasses import dataclass, field
from io import BytesIO
from typing import Literal

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from monori.server.app.importer import parse_amount_kop, parse_date
from monori.server.app.workbook import spec
from monori.server.app.workbook.models import (
    ParsedWorkbook,
    WorkbookTransaction,
)
from monori.server.app.workbook.models import (
    WorkbookBudget as WorkbookBudgetRow,
)
from monori.server.app.workbook.models import (
    WorkbookCategory as WorkbookCategoryRow,
)
from monori.server.app.workbook.models import (
    WorkbookGroup as WorkbookGroupRow,
)
from monori.server.app.workbook.models import (
    WorkbookParseError as WorkbookParseErrorRow,
)
from monori.server.app.workbook.models import (
    WorkbookTransaction as WorkbookTransactionRow,
)


@dataclass(slots=True)
class YearCategoryRow:
    """Represent YearCategoryRow."""

    group: str
    budgets: dict[int, int] = field(default_factory=dict)
    outflows: dict[int, int] = field(default_factory=dict)
    balances: dict[int, int] = field(default_factory=dict)


@dataclass(slots=True)
class YearSheetRow:
    """Represent YearSheetRow."""

    year: int
    months: list[int]
    cats: dict[tuple[str, str], YearCategoryRow]
    income: dict[int, int]
    available: dict[int, int]
    seeds: dict[int, int]
    seed: int | None
    sections: list["YearSection"]


@dataclass(slots=True)
class YearSection:
    """Represent YearSection."""

    name: str
    kind: str
    rows: list[tuple[int, str]]


type CategoryKey = tuple[str, str]


def _year_entry(source: YearSheetRow, key: CategoryKey) -> YearCategoryRow | None:
    return source.cats.get(key)


@dataclass(slots=True)
class LayoutRow:
    """Represent LayoutRow."""

    header_row: int
    bases: list[int]
    out_off: int
    bal_off: int
    label_col: int
    start_month: int


YEAR_RE = re.compile(r"^(\d{4})(_archive)?$")


TX_ALIASES = {
    "date": ("Дата операции", "Date", "Operation date"),
    "card": ("Номер карты", "Card"),
    "account": ("Account",),
    "status": ("Статус", "Status"),
    "amount": ("Operation amount", "Сумма операции", "Amount"),
    "currency": ("Валюта операции", "Transaction currency"),
    "pay_amount": ("Сумма платежа", "Payment amount"),
    "pay_currency": ("Валюта платежа", "Payment currency"),
    "bank_category": ("Категория", "Category"),
    "mcc": ("MCC",),
    "description": ("Описание", "Description", "Transaction description"),
    "category": ("Monori Category",),
    "category_group": ("Monori Category Group",),
    "comment": ("Comment",),
}


TX_REQUIRED = ("date", "amount")

MONTH_ABBREVS = {
    "ЯНВ": 1,
    "ФЕВ": 2,
    "МАР": 3,
    "АПР": 4,
    "МАЙ": 5,
    "МАЯ": 5,
    "ИЮН": 6,
    "ИЮЛ": 7,
    "АВГ": 8,
    "СЕН": 9,
    "ОКТ": 10,
    "НОЯ": 11,
    "ДЕК": 12,
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}


KNOWN_TX_HEADERS = (
    *tuple(name for names in TX_ALIASES.values() for name in names if name != "Category"),
    "Дата платежа",
    "Кэшбэк",
    "Бонусы (включая кэшбэк)",
    "Округление на инвесткопилку",
    "Сумма операции с округлением",
)


SIDE_TABLE_SCAN_ROWS = 500

BUDGET_HEADERS = ("Бюджет", "Budgeted")
OUTFLOW_HEADERS = ("Расход", "Outflows")
BALANCE_HEADERS = ("Баланс", "Balance")
LABEL_HEADERS = ("Категория", "Категории", "Category", "Categories")
SKIP_LABELS = (*LABEL_HEADERS, "Month Summary", "Total", "Итого")

DEFAULT_CURRENCY = "RUB"
MIN_LAYOUT_BASES = 2
MONTHS_IN_YEAR = 12
INCOME_GROUP = "Inflow"
INCOME_CATEGORY = "Income"
OTHER_GROUP = "Other"
OPENING_DESCRIPTION = "Opening balance"

ADJUST_TOLERANCE_KOP = 2
VERIFY_TOLERANCE_KOP = 5


class WorkbookError(Exception):
    """Represent WorkbookError."""


def _group_kind(value: str | None) -> Literal["income", "expense"] | None:
    if value is None:
        return None
    return "income" if value == "income" else "expense"


def _s(cell: Cell | MergedCell | None) -> str:
    value = None if cell is None else cell.value
    if value is None:
        return ""
    return str(value).strip()


SUM_FORMULA_RE = re.compile(r"^=[-+]?\d+(\.\d+)?([-+]\d+(\.\d+)?)*$")


def _kop(cell: Cell | MergedCell | None) -> int | None:
    value = None if cell is None else cell.value
    if isinstance(value, bool):
        return None
    if isinstance(value, str):
        cleaned = re.sub(r"[\s\u00a0\u202f]", "", value).replace(",", ".")
        if not cleaned:
            return None

        if SUM_FORMULA_RE.match(cleaned):
            terms = re.findall(r"[-+]?\d+(?:\.\d+)?", cleaned[1:])
            return spec.kop_from_rub(sum(float(t) for t in terms))
        try:
            value = float(cleaned)
        except ValueError:
            return None
    if not isinstance(value, int | float):
        return None
    return spec.kop_from_rub(value)


def _last_day(year: int, month: int) -> datetime.date:
    end = (
        datetime.date(year + 1, 1, 1)
        if month == MONTHS_IN_YEAR
        else datetime.date(year, month + 1, 1)
    )
    return end - datetime.timedelta(days=1)


def _stamp(year: int, month: int) -> str:
    return _last_day(year, month).strftime("%Y-%m-%dT12:00:00")


def _month_num(cell: Cell | MergedCell | None) -> int | None:
    abbr = _s(cell).upper()[:3]
    return MONTH_ABBREVS.get(abbr)


def _layout_bases(ws: Worksheet, row_number: int) -> list[int]:
    return [
        column
        for column in range(1, ws.max_column + 1)
        if _s(ws.cell(row_number, column)) in BUDGET_HEADERS
    ]


def _layout_offsets(ws: Worksheet, row_number: int, first_base: int) -> tuple[int, int] | None:
    out_off = bal_off = None
    for column in range(first_base + 1, ws.max_column + 1):
        value = _s(ws.cell(row_number, column))
        if value in OUTFLOW_HEADERS and out_off is None:
            out_off = column - first_base
        if value in BALANCE_HEADERS and bal_off is None:
            bal_off = column - first_base
    return (out_off, bal_off) if out_off is not None and bal_off is not None else None


def _layout_label_column(ws: Worksheet, header_row: int, first_base: int) -> int:
    for row_number in (header_row, header_row + 1):
        for column in range(1, first_base):
            if _s(ws.cell(row_number, column)) in LABEL_HEADERS:
                return column
    return _label_col(ws, header_row, first_base)


def _layout_start_month(ws: Worksheet, first_base: int) -> int:
    for row_number in (1, 2, 3):
        month = _month_num(ws.cell(row_number, first_base))
        if month is not None:
            return month
    return 1


def _find_layout(ws: Worksheet) -> LayoutRow | None:
    """
    Locates the month blocks of a year sheet by looking for the row that repeats.

    a Budgeted/Outflows/Balance header per month — which is the same grid in a.
    workbook we wrote and in the hand-kept spreadsheet, only sitting at a
    different row and labelled in a different language. Returns None when no
    such row exists, which is how a sheet says it is not a year grid at all.
    """
    for header_row in range(1, 11):
        bases = _layout_bases(ws, header_row)
        if len(bases) < MIN_LAYOUT_BASES:
            continue
        offsets = _layout_offsets(ws, header_row, bases[0])
        if offsets is None:
            continue
        out_off, bal_off = offsets
        return LayoutRow(
            header_row=header_row,
            bases=bases,
            out_off=out_off,
            bal_off=bal_off,
            label_col=_layout_label_column(ws, header_row, bases[0]),
            start_month=_layout_start_month(ws, bases[0]),
        )
    return None


def _label_col(ws: Worksheet, header_row: int, first_base: int) -> int:
    """
    Handle The category column when the grid never names it: of the columns left of the.

    first month block, the one carrying the most labels below the header.
    """
    best, best_count = 1, 0
    for c in range(1, first_base):
        count = sum(
            1
            for r in range(header_row + 1, min(ws.max_row, header_row + 60) + 1)
            if _s(ws.cell(r, c))
        )
        if count > best_count:
            best, best_count = c, count
    return best


def _kind_of(group_name: str, groups: Iterable[WorkbookGroupRow]) -> str:
    return next((g.kind for g in groups if g.name == group_name), "expense")


def _parse_categories(
    ws: Worksheet,
    warnings: list[str],
) -> tuple[list[WorkbookGroupRow], list[WorkbookCategoryRow]]:
    """
    Handle Reads a category sheet that states the structure outright: category rows.

    (`sort | group | category | keywords`) and, when present, a group table.
    (`group | sort | IN/OUT`). Groups fall back to the ones the category rows
    name so a sheet missing that table still imports.
    """
    groups: list[WorkbookGroupRow] = []
    categories: list[WorkbookCategoryRow] = []
    group_rows_seen = False
    for row in ws.iter_rows(min_row=1):
        cells = list(row) + [None] * (4 - len(row))
        c1, c2, c3, c4 = cells[:4]
        s1, s2, s3 = _s(c1), _s(c2), _s(c3)
        if s1 in ("Sort Order", "Category Group") or (not s1 and not s2):
            continue
        c2_value = None if c2 is None else c2.value
        c1_value = None if c1 is None else c1.value
        if s3 in (spec.TYPE_IN, spec.TYPE_OUT) and isinstance(c2_value, int | float):
            name, _ = spec.strip_glyph(_unquote(s1))
            groups.append(
                WorkbookGroupRow(
                    name=name,
                    sort=int(c2_value),
                    kind="income" if s3 == spec.TYPE_IN else "expense",
                ),
            )
            group_rows_seen = True
            continue
        if isinstance(c1_value, int | float) and s2 and s3:
            name, kind = spec.strip_glyph(_unquote(s2))
            categories.append(
                WorkbookCategoryRow(
                    group=name,
                    group_kind=_group_kind(kind),
                    group_sort=int(c1_value),
                    name=_unquote(s3),
                    keywords=_unquote(_s(c4)),
                ),
            )
            continue
        if s1 or s2 or s3:
            warnings.append(f"Categories: unrecognized row skipped: {[s1, s2, s3][:3]}")
    if not group_rows_seen:
        seen: dict[str, WorkbookGroupRow] = {}
        for cat in categories:
            if cat.group not in seen:
                seen[cat.group] = WorkbookGroupRow(
                    name=cat.group,
                    sort=cat.group_sort,
                    kind=cat.group_kind or "expense",
                )
        groups = list(seen.values())
        if groups:
            warnings.append("Categories: group table missing, groups derived from category rows")
    return groups, categories


def _sheet_sections(ws: Worksheet, layout: LayoutRow) -> list[YearSection]:
    """
    Handle Splits the category area into (group, [(row, category), ...]) sections.

    A row whose label starts with a kind glyph opens a group; in the old
    glyph-less layout the first labelled row after a fully blank gap does.
    """
    label_col = layout.label_col
    sections: list[YearSection] = []
    current: YearSection | None = None
    in_gap = True
    for r in range(layout.header_row + 1, ws.max_row + 1):
        label = _s(ws.cell(r, label_col))
        if not label or label in SKIP_LABELS:
            in_gap = in_gap or not label
            if label in SKIP_LABELS:
                in_gap = True
            continue
        name, kind = spec.strip_glyph(label)
        if kind is not None or (in_gap and current is None) or (in_gap and current is not None):
            current = YearSection(name=name, kind=kind or "expense", rows=[])
            sections.append(current)
        elif current is None:
            current = YearSection(name=name, kind="expense", rows=[])
            sections.append(current)
        else:
            current.rows.append((r, label))
        in_gap = False
    return sections


def _summary_value(ws: Worksheet, base: int, labels: tuple[str, ...]) -> int | None:
    for r in range(1, 7):
        text = _s(ws.cell(r, base + 2))
        if any(text.startswith(lb) for lb in labels):
            return _kop(ws.cell(r, base + 1))
    return None


def _year_months(layout: LayoutRow) -> list[tuple[int, int]]:
    return [
        (layout.start_month + i, base)
        for i, base in enumerate(layout.bases)
        if layout.start_month + i <= MONTHS_IN_YEAR
    ]


def _year_categories(
    ws: Worksheet,
    layout: LayoutRow,
    months: list[tuple[int, int]],
    sections: list[YearSection],
) -> dict[tuple[str, str], YearCategoryRow]:
    cats: dict[tuple[str, str], YearCategoryRow] = {}
    for section in sections:
        for r, name in section.rows:
            entry = cats.setdefault(
                (section.name, name),
                YearCategoryRow(group=section.name),
            )
            for m, base in months:
                b = _kop(ws.cell(r, base))
                o = _kop(ws.cell(r, base + layout.out_off))
                bal = _kop(ws.cell(r, base + layout.bal_off))
                if b is not None:
                    entry.budgets[m] = b
                if o is not None:
                    entry.outflows[m] = o
                if bal is not None:
                    entry.balances[m] = bal
    return cats


def _year_summaries(
    ws: Worksheet,
    months: list[tuple[int, int]],
) -> tuple[dict[int, int], dict[int, int]]:
    income: dict[int, int] = {}
    available: dict[int, int] = {}
    for m, base in months:
        inc = _summary_value(ws, base, ("Income for", "Поступления в"))
        if inc is not None:
            income[m] = inc
        for r in (5, 6):
            label = _s(ws.cell(r + 1, base + 1))
            if label.startswith(("Available", "Доступный")):
                av = _kop(ws.cell(r, base + 1))
                if av is not None:
                    available[m] = av
                break
    return income, available


def _year_seeds(ws: Worksheet, months: list[tuple[int, int]]) -> dict[int, int]:
    seeds: dict[int, int] = {}
    for m, base in months:
        carried = _summary_value(ws, base, ("Not budgeted", "Не заложено"))
        if carried is not None:
            seeds[m] = carried
    return seeds


def _parse_year_sheet(ws: Worksheet, year: int, layout: LayoutRow) -> YearSheetRow:
    months = _year_months(layout)
    sections = _sheet_sections(ws, layout)
    cats = _year_categories(ws, layout, months, sections)
    income, available = _year_summaries(ws, months)
    seeds = _year_seeds(ws, months)
    return YearSheetRow(
        year=year,
        months=[m for m, _ in months],
        cats=cats,
        income=income,
        available=available,
        seeds=seeds,
        seed=seeds.get(months[0][0]) if months else None,
        sections=sections,
    )


def _tx_header_index(ws: Worksheet) -> dict[str, int] | None:
    header = next(ws.iter_rows(min_row=1, max_row=1), None)
    if header is None:
        return None
    index: dict[str, int] = {}
    for i, cell in enumerate(header):
        name = _s(cell)
        if name:
            index.setdefault(name, i)
    return index


def _parse_dt(cell: Cell | MergedCell | None) -> datetime.datetime | None:
    value = None if cell is None else cell.value
    if isinstance(value, datetime.datetime):
        return value
    if isinstance(value, datetime.date):
        return datetime.datetime(value.year, value.month, value.day, tzinfo=datetime.UTC).replace(
            tzinfo=None
        )
    text = _s(cell)
    if not text:
        return None
    parsed = parse_date(text)
    if parsed:
        return parsed.replace(tzinfo=None)
    try:
        return datetime.datetime.fromisoformat(text)
    except ValueError:
        return None


def _unquote(value: str) -> str:
    """
    Reverses our exporter's formula-escape and nothing else: a leading.

    apostrophe is stripped only when it guards a formula prefix, so a value that.
    legitimately starts with one survives the round-trip.
    """
    if value.startswith("'") and value[1:].startswith(("=", "+", "@")):
        return value[1:]
    return value


def _amount(cell: Cell | MergedCell | None) -> int | None:
    """Kopecks from a cell that may be a number, a formatted string, or blank."""
    kop = _kop(cell)
    if kop is not None:
        return kop
    text = _s(cell)
    return parse_amount_kop(text) if text else None


def _tx_columns(idx: Mapping[str, int]) -> dict[str, int | None]:
    """Which column, if any, holds each field this reader knows how to use."""
    return {
        field: next((idx[name] for name in names if name in idx), None)
        for field, names in TX_ALIASES.items()
    }


@dataclass(slots=True)
class _TransactionReader:
    columns: dict[str, int | None]
    category_column: int

    def cell(
        self,
        row: tuple[Cell | MergedCell, ...],
        field: str,
    ) -> Cell | MergedCell | None:
        column = self.columns[field]
        return row[column] if column is not None and column < len(row) else None

    def text(self, row: tuple[Cell | MergedCell, ...], field: str) -> str:
        return _unquote(_s(self.cell(row, field)))

    def category(self, row: tuple[Cell | MergedCell, ...]) -> str:
        return _unquote(_s(row[self.category_column])) if self.category_column < len(row) else ""


def _transaction_reader(ws: Worksheet) -> _TransactionReader:
    index = _tx_header_index(ws)
    if index is None:
        msg = "Transactions sheet is empty"
        raise WorkbookError(msg)
    columns = _tx_columns(index)
    _require_transaction_columns(columns)
    category_column = columns["category"]
    return _TransactionReader(
        columns,
        category_column if category_column is not None else _category_col(ws, index),
    )


def _require_transaction_columns(columns: Mapping[str, int | None]) -> None:
    missing = [TX_ALIASES[field][0] for field in TX_REQUIRED if columns[field] is None]
    if columns["pay_amount"] is not None and TX_ALIASES["amount"][0] in missing:
        missing.remove(TX_ALIASES["amount"][0])
    if missing:
        msg = f"Transactions sheet is missing required columns: {missing}"
        raise WorkbookError(msg)


@dataclass(slots=True)
class _TransactionParseState:
    errors: list[WorkbookParseErrorRow]
    rows: list[WorkbookTransactionRow] = field(default_factory=list)
    seen: set[tuple[str, int, str, str, str]] = field(default_factory=set)
    skipped_status: int = 0
    foreign: dict[str, int] = field(default_factory=dict)
    duplicates: int = 0

    def consume(
        self,
        reader: _TransactionReader,
        row: tuple[Cell | MergedCell, ...],
        number: int,
    ) -> None:
        if _blank_transaction_row(row) or self._skipped_status(reader, row):
            return
        transaction = self._transaction(reader, row, number)
        if transaction is None:
            return
        if transaction.currency != DEFAULT_CURRENCY:
            self.foreign[transaction.currency] = self.foreign.get(transaction.currency, 0) + 1
        key = _transaction_key(transaction)
        if key in self.seen:
            self.duplicates += 1
            return
        self.seen.add(key)
        self.rows.append(transaction)

    def _skipped_status(
        self,
        reader: _TransactionReader,
        row: tuple[Cell | MergedCell, ...],
    ) -> bool:
        if _s(reader.cell(row, "status")).upper() in ("OK", ""):
            return False
        self.skipped_status += 1
        return True

    def _transaction(
        self,
        reader: _TransactionReader,
        row: tuple[Cell | MergedCell, ...],
        number: int,
    ) -> WorkbookTransactionRow | None:
        date = _parse_dt(reader.cell(row, "date"))
        amount, currency = _transaction_amount(reader, row)
        description = reader.text(row, "description")
        if date is None or amount is None:
            if date is not None or amount is not None or description:
                self.errors.append(
                    WorkbookParseErrorRow(row=number, error="unparseable date or amount")
                )
            return None
        return WorkbookTransactionRow(
            date=date.strftime("%Y-%m-%dT%H:%M:%S"),
            amount=amount,
            description=description,
            currency=(currency or DEFAULT_CURRENCY).upper(),
            bank_category=reader.text(row, "bank_category"),
            mcc=reader.text(row, "mcc"),
            comment=reader.text(row, "comment"),
            monori_category_group=reader.text(row, "category_group"),
            monori_category=reader.category(row),
            marker=reader.text(row, "card") or reader.text(row, "account"),
        )


def _blank_transaction_row(row: tuple[Cell | MergedCell, ...]) -> bool:
    return all(_s(cell) == "" for cell in row)


def _transaction_amount(
    reader: _TransactionReader,
    row: tuple[Cell | MergedCell, ...],
) -> tuple[int | None, str]:
    amount = _amount(reader.cell(row, "pay_amount"))
    currency = _s(reader.cell(row, "pay_currency"))
    if amount is None:
        amount = _amount(reader.cell(row, "amount"))
        currency = _s(reader.cell(row, "currency"))
    return amount, currency or _s(reader.cell(row, "currency"))


def _transaction_key(transaction: WorkbookTransactionRow) -> tuple[str, int, str, str, str]:
    return (
        transaction.date,
        transaction.amount,
        transaction.description,
        transaction.marker,
        transaction.currency,
    )


def _parse_transactions(
    ws: Worksheet,
    warnings: list[str],
    errors: list[WorkbookParseErrorRow],
) -> list[WorkbookTransactionRow]:
    reader = _transaction_reader(ws)
    state = _TransactionParseState(errors)
    for n, row in enumerate(ws.iter_rows(min_row=2), start=2):
        state.consume(reader, row, n)
    if state.duplicates:
        warnings.append(
            f"Transactions: {state.duplicates} rows identical in date, amount, description and card"
            " — kept once",
        )
    if state.skipped_status:
        warnings.append(f"Transactions: {state.skipped_status} non-OK rows skipped")
    for code, count in sorted(state.foreign.items()):
        warnings.append(
            f"Transactions: {count} rows in {code} — they need an account held in {code} "
            "to land on",
        )
    return state.rows


def _known_max_col(idx: Mapping[str, int]) -> int:
    return max((i for h, i in idx.items() if h in KNOWN_TX_HEADERS), default=-1)


def _find_keyword_block(ws: Worksheet, idx: Mapping[str, int]) -> int | None:
    """
    Locates the `category name | pipe-separated keywords` side table by.

    content: the column pair (right of the known bank headers) with the most.
    rows whose second cell contains a pipe. Purely positional lookup broke on
    the live file — the table starts at row 1, so its own cells pollute the
    header index and shift any fixed offset.
    """
    start = _known_max_col(idx) + 1
    scores: dict[int, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, SIDE_TABLE_SCAN_ROWS)):
        for base in range(start, len(row) - 1):
            if _s(row[base]) and "|" in _s(row[base + 1]):
                scores[base] = scores.get(base, 0) + 1
    if not scores:
        return None
    return max(scores, key=lambda b: (scores[b], -b))


def _category_col(ws: Worksheet, idx: Mapping[str, int]) -> int:
    """
    Handle The per-row category lives right of the known bank headers and left of the.

    keyword table — but the live template puts *two* columns there: the keyword.
    rules compute a guess in the first, and the second either carries that guess
    through or replaces it with what the user typed by hand. Only the second one
    is what the sheet's own totals are built from, so it is the truth: a hand
    label wins outright, and the automatic guess only survives where the user let
    it. Taking the first populated column instead left 56% of the rows here
    uncategorized.

    Picking the fullest column finds it without hardcoding an offset, and still
    works for our own exporter, which writes a single column.
    """
    start = _known_max_col(idx) + 1
    stop = _find_keyword_block(ws, idx)
    stop_col = ws.max_column if stop is None else stop
    filled: dict[int, int] = {}
    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, SIDE_TABLE_SCAN_ROWS)):
        for c in range(start, min(stop_col, len(row))):
            if _s(row[c]):
                filled[c] = filled.get(c, 0) + 1
    if not filled:
        return _known_max_col(idx) + 2
    return max(filled, key=lambda c: (filled[c], c))


def _parse_keywords(ws: Worksheet, idx: Mapping[str, int]) -> dict[str, str]:
    """
    Handle Reads the keyword side table (see _find_keyword_block): category name |.

    pipe-separated keywords, starting at row 1.
    """
    base = _find_keyword_block(ws, idx)
    base_col = _known_max_col(idx) + 3 if base is None else base
    keywords: dict[str, str] = {}
    for row in ws.iter_rows(min_row=1):
        if base_col >= len(row):
            continue
        name = _s(row[base_col])
        kws = _s(row[base_col + 1]) if base_col + 1 < len(row) else ""
        if name and kws and ("|" in kws or len(kws) > 1):
            keywords.setdefault(name, kws)
    return keywords


def _synthetic(
    period: tuple[int, int],
    amount: int,
    category: str,
    description: str,
    group: str = "",
) -> WorkbookTransactionRow:
    date_iso = _stamp(*period)
    return WorkbookTransactionRow(
        date=date_iso,
        amount=amount,
        description=description,
        currency=DEFAULT_CURRENCY,
        monori_category=category,
        monori_category_group=group,
    )


def account_slot(tx: WorkbookTransaction) -> str:
    """
    Which account a row must land on. A card marker alone is not enough: the.

    same marker can carry rows in more than one currency (interest on a foreign.
    balance arrives with no card number at all), and an amount only means
    anything on an account held in that currency. Marker and currency together
    are the unit the user maps.
    """
    return f"{tx.currency or DEFAULT_CURRENCY}:{tx.marker}"


def _activity_span(
    transactions: Iterable[WorkbookTransactionRow],
    sources: Iterable[YearSheetRow],
) -> tuple[tuple[int, int] | None, tuple[int, int] | None]:
    """
    First and last (year, month) showing real activity: a transaction, a nonzero.

    cached outflow or income. Budgets deliberately do not count — planning months.
    ahead is normal, a budget alone creates no transactions in the sheet, and
    the cached balances of those future months are pure carry residue;
    reconciling against them fabricates future-dated synthetic rows. A sheet
    whose earlier months are empty scaffolding is the same case read from the
    other end: nothing happened there, so nothing is owed there either.
    """
    seen = []

    seen.extend(
        [(int(tx.date[:4]), int(tx.date[5:7])) for tx in transactions if tx.monori_category]
    )
    for source in sources:
        y = source.year
        seen.extend((y, m) for m, v in source.income.items() if v)
        for entry in source.cats.values():
            seen.extend((y, m) for m, v in entry.outflows.items() if v)
    return (min(seen), max(seen)) if seen else (None, None)


def _month_range(start: tuple[int, int], end: tuple[int, int]) -> Iterable[tuple[int, int]]:
    y, m = start
    while (y, m) <= end:
        yield (y, m)
        m += 1
        if m > MONTHS_IN_YEAR:
            y, m = y + 1, 1


def parse_workbook(data: bytes) -> ParsedWorkbook:
    """
    Handle Returns {groups, categories, transactions, budgets, warnings, errors} for any.

    budget workbook — see the module docstring for how the shape is discovered.
    """
    try:
        wb = load_workbook(BytesIO(data), data_only=True)
    except Exception as exc:
        msg = f"not a readable .xlsx workbook: {exc}"
        raise WorkbookError(msg) from exc
    try:
        return _parse(wb)
    finally:
        wb.close()


def _parse(wb: Workbook) -> ParsedWorkbook:
    warnings: list[str] = []
    errors: list[WorkbookParseErrorRow] = []
    sheets = _parse_sheets(wb, warnings, errors)
    catalog = _build_catalog(wb, sheets, warnings)
    for transaction in sheets.transactions:
        name = transaction.monori_category
        if (
            name
            and not transaction.monori_category_group
            and len(catalog.category_groups.get(name, set())) > 1
        ):
            errors.append(
                WorkbookParseErrorRow(
                    row=0,
                    error=f"ambiguous legacy category {name!r}; add Monori Category Group",
                )
            )
    if not sheets.has_years:
        return _result(
            _ParsedRows(catalog.groups, catalog.categories, sheets.transactions, []),
            warnings,
            errors,
        )
    income_category = catalog.income_category()
    budgets = _collect_budgets(sheets, catalog)
    reconciliation = _Reconciliation(sheets, catalog, sheets.transactions, budgets, income_category)
    synthetic, reconciliation_warnings = reconciliation.run()
    warnings.extend(reconciliation_warnings)

    return _result(
        _ParsedRows(catalog.groups, catalog.categories, sheets.transactions + synthetic, budgets),
        warnings,
        errors,
    )


@dataclass(slots=True)
class _SheetData:
    transactions: list[WorkbookTransactionRow]
    keywords: dict[str, str]
    archive_years: dict[int, YearSheetRow]
    live_years: dict[int, YearSheetRow]
    plain_sheets: dict[int, YearSheetRow]
    seam_year: int
    seam_sheet: YearSheetRow | None

    @property
    def has_years(self) -> bool:
        return bool(self.live_years or self.archive_years)

    @property
    def sources(self) -> list[YearSheetRow]:
        return list(self.archive_years.values()) + list(self.live_years.values())


def _parse_sheets(
    wb: Workbook,
    warnings: list[str],
    errors: list[WorkbookParseErrorRow],
) -> _SheetData:
    tx_ws = _transaction_sheet(wb)
    tx_idx = _tx_header_index(tx_ws)
    transactions = _parse_transactions(tx_ws, warnings, errors)
    if tx_idx is None:
        msg = "Transactions sheet is missing required columns"
        raise WorkbookError(msg)
    archive_years, plain_sheets = _parse_year_sheets(wb, warnings)
    live_years = {
        year: parsed for year, parsed in plain_sheets.items() if year not in archive_years
    }
    _warn_unknown_sheets(wb, warnings)
    first_live = min(live_years) if live_years else None
    seam_year = -1 if first_live is None else first_live - 1
    return _SheetData(
        transactions,
        _parse_keywords(tx_ws, tx_idx),
        archive_years,
        live_years,
        plain_sheets,
        seam_year,
        plain_sheets.get(seam_year),
    )


def _transaction_sheet(wb: Workbook) -> Worksheet:
    if spec.SHEET_TRANSACTIONS not in wb.sheetnames:
        msg = f"missing required sheet: {spec.SHEET_TRANSACTIONS}"
        raise WorkbookError(msg)
    return wb[spec.SHEET_TRANSACTIONS]


def _parse_year_sheets(
    wb: Workbook,
    warnings: list[str],
) -> tuple[dict[int, YearSheetRow], dict[int, YearSheetRow]]:
    archive_years: dict[int, YearSheetRow] = {}
    plain_sheets: dict[int, YearSheetRow] = {}
    for name in wb.sheetnames:
        year_match = YEAR_RE.match(name)
        if not year_match:
            continue
        ws = wb[name]
        if not hasattr(ws, "iter_rows"):
            continue
        layout = _find_layout(ws)
        if layout is None:
            warnings.append(f"{name}: unrecognized year sheet layout, ignored")
            continue
        parsed = _parse_year_sheet(ws, int(year_match.group(1)), layout)
        destination = archive_years if year_match.group(2) else plain_sheets
        destination[parsed.year] = parsed
    return archive_years, plain_sheets


def _warn_unknown_sheets(wb: Workbook, warnings: list[str]) -> None:
    known_sheets = {spec.SHEET_CATEGORIES, spec.SHEET_TRANSACTIONS, spec.SHEET_DASHDATA}
    warnings.extend(
        f"unknown sheet ignored: {name}"
        for name in wb.sheetnames
        if name not in known_sheets and not YEAR_RE.match(name)
    )


@dataclass(slots=True)
class _CategoryCatalog:
    keywords: dict[str, str]
    groups: list[WorkbookGroupRow] = field(default_factory=list)
    categories: list[WorkbookCategoryRow] = field(default_factory=list)
    group_names: set[str] = field(default_factory=set)
    category_groups: dict[str, set[str]] = field(default_factory=dict)

    def add_group(
        self,
        name: str,
        kind: Literal["income", "expense"],
        sort: int | None = None,
    ) -> None:
        if name not in self.group_names:
            self.group_names.add(name)
            self.groups.append(
                WorkbookGroupRow(
                    name=name, sort=len(self.groups) if sort is None else sort, kind=kind
                ),
            )

    def add_category(
        self,
        name: str,
        group: str,
        keywords_text: str | None = None,
        group_kind: Literal["income", "expense"] | None = None,
        group_sort: int = 0,
    ) -> None:
        groups = self.category_groups.setdefault(name, set())
        if group not in groups:
            groups.add(group)
            self.categories.append(
                WorkbookCategoryRow(
                    group=group,
                    group_kind=group_kind,
                    group_sort=group_sort,
                    name=name,
                    keywords=keywords_text
                    if keywords_text is not None
                    else self.keywords.get(name, ""),
                ),
            )

    def income_category(self) -> str:
        income_category = next(
            (
                category.name
                for category in self.categories
                if category.name == INCOME_CATEGORY
                or _kind_of(category.group, self.groups) == "income"
            ),
            None,
        )
        if income_category is None:
            self.add_group(INCOME_GROUP, "income")
            self.add_category(INCOME_CATEGORY, INCOME_GROUP)
            return INCOME_CATEGORY
        return income_category


def _build_catalog(
    wb: Workbook,
    sheets: _SheetData,
    warnings: list[str],
) -> _CategoryCatalog:
    catalog = _CategoryCatalog(sheets.keywords)
    stated = _add_stated_categories(wb, catalog, warnings)
    _add_grid_categories(sheets, catalog, stated=stated, warnings=warnings)
    _add_transaction_categories(sheets.transactions, catalog)
    return catalog


def _add_stated_categories(
    wb: Workbook,
    catalog: _CategoryCatalog,
    warnings: list[str],
) -> bool:
    stated_warnings: list[str] = []
    stated_groups, stated_categories = (
        _parse_categories(wb[spec.SHEET_CATEGORIES], stated_warnings)
        if spec.SHEET_CATEGORIES in wb.sheetnames
        else ([], [])
    )
    if not stated_categories:
        _warn_missing_categories(stated_warnings, warnings)
        return False
    warnings.extend(stated_warnings)
    for group in stated_groups:
        catalog.add_group(group.name, group.kind, group.sort)
    for category in stated_categories:
        catalog.add_group(category.group, category.group_kind or "expense", category.group_sort)
        catalog.add_category(
            category.name,
            category.group,
            category.keywords,
            category.group_kind,
            category.group_sort,
        )
    return True


def _warn_missing_categories(
    stated_warnings: list[str],
    warnings: list[str],
) -> None:
    if stated_warnings:
        warnings.append(
            f"Categories: no category rows recognized ({len(stated_warnings)} rows skipped),"
            " structure taken from the year grids",
        )


def _add_grid_categories(
    sheets: _SheetData,
    catalog: _CategoryCatalog,
    *,
    stated: bool,
    warnings: list[str],
) -> None:
    all_sections = [
        section
        for years in (sheets.live_years, sheets.archive_years)
        for year in years
        for section in years[year].sections
    ]
    if (
        not stated
        and all_sections
        and not any(section.kind == "income" for section in all_sections)
    ):
        catalog.add_group(INCOME_GROUP, "income")
        catalog.add_category(INCOME_CATEGORY, INCOME_GROUP)
    for years in (sheets.live_years, sheets.archive_years):
        for year in sorted(years, reverse=True):
            _add_year_sections(year, years[year], catalog, stated=stated, warnings=warnings)


def _add_year_sections(
    year: int,
    source: YearSheetRow,
    catalog: _CategoryCatalog,
    *,
    stated: bool,
    warnings: list[str],
) -> None:
    for section in source.sections:
        if stated:
            _warn_unknown_section_rows(year, section.rows, catalog, warnings)
            continue
        catalog.add_group(section.name, _group_kind(section.kind) or "expense")
        for _, name in section.rows:
            catalog.add_category(name, section.name)


def _warn_unknown_section_rows(
    year: int,
    rows: Iterable[tuple[int, str]],
    catalog: _CategoryCatalog,
    warnings: list[str],
) -> None:
    for _, name in rows:
        if name not in catalog.category_groups:
            warnings.append(f"{year}: unknown row label skipped: {name[:60]}")


def _add_transaction_categories(
    transactions: Iterable[WorkbookTransactionRow],
    catalog: _CategoryCatalog,
) -> None:
    named_only: dict[str, list[int]] = {}
    for transaction in transactions:
        name = transaction.monori_category
        if name and name not in catalog.category_groups:
            named_only.setdefault(name, []).append(transaction.amount)
    income_group = next((group.name for group in catalog.groups if group.kind == "income"), None)
    expense_group = next((group.name for group in catalog.groups if group.kind != "income"), None)
    for name, amounts in named_only.items():
        income_group, expense_group = _add_transaction_category(
            name,
            amounts,
            catalog,
            income_group,
            expense_group,
        )


def _add_transaction_category(
    name: str,
    amounts: list[int],
    catalog: _CategoryCatalog,
    income_group: str | None,
    expense_group: str | None,
) -> tuple[str | None, str | None]:
    if all(amount >= 0 for amount in amounts):
        if income_group is None:
            catalog.add_group(INCOME_GROUP, "income")
            income_group = INCOME_GROUP
        catalog.add_category(name, income_group)
        return income_group, expense_group
    if expense_group is None:
        catalog.add_group(OTHER_GROUP, "expense")
        expense_group = OTHER_GROUP
    catalog.add_category(name, expense_group)
    return income_group, expense_group


def _collect_budgets(sheets: _SheetData, catalog: _CategoryCatalog) -> list[WorkbookBudgetRow]:
    budgets: list[WorkbookBudgetRow] = []
    for source in sheets.sources:
        for (group, name), entry in source.cats.items():
            if name not in catalog.category_groups:
                continue
            for month, amount in entry.budgets.items():
                if amount:
                    budgets.append(
                        WorkbookBudgetRow(
                            category=name,
                            year=source.year,
                            month=month,
                            amount=amount,
                            group=group,
                        ),
                    )
    return budgets


@dataclass(slots=True)
class _Reconciliation:
    sheets: _SheetData
    catalog: _CategoryCatalog
    transactions: list[WorkbookTransactionRow]
    budgets: list[WorkbookBudgetRow]
    income_category: str
    kinds: dict[CategoryKey, str] = field(init=False)
    tx_sums: dict[tuple[str, str, int, int], int] = field(init=False)
    income_sums: dict[tuple[int, int], int] = field(init=False)
    budget_map: dict[tuple[str, str, int, int], int] = field(init=False)
    months_with_rows: set[tuple[int, int]] = field(init=False)
    synthetic: list[WorkbookTransactionRow] = field(default_factory=list)
    n_hist: int = 0
    n_adjust: int = 0
    n_seam: int = 0
    avail_residuals: list[tuple[int, int, int]] = field(default_factory=list)
    opened: tuple[int, int, int] | None = None

    def __post_init__(self) -> None:
        self.kinds = _category_kinds(self.catalog.groups, self.catalog.categories)
        self.tx_sums, self.income_sums = _transaction_totals(self.transactions, self.kinds)
        self.budget_map = _budget_totals(self.budgets)
        self.months_with_rows = {
            (int(tx.date[:4]), int(tx.date[5:7])) for tx in self.transactions if tx.monori_category
        }

    def run(self) -> tuple[list[WorkbookTransactionRow], list[str]]:
        self._adjust_income_targets()
        self._reconcile_balances()
        return self.synthetic, self._warnings()

    def _adjust_income_targets(self) -> None:
        for source in self.sheets.sources:
            for month, target in source.income.items():
                have = self.income_sums.get((source.year, month), 0)
                delta = target - have
                if abs(delta) > ADJUST_TOLERANCE_KOP:
                    self.synthetic.append(
                        _synthetic(
                            (source.year, month), delta, self.income_category, self.income_category
                        ),
                    )
                    self.income_sums[(source.year, month)] = have + delta
                    self._count_adjustment(source.year, month)

    def _reconcile_balances(self) -> None:
        start, end, first_active = self._range()
        opening = self._opening_balance(first_active)
        balances: dict[CategoryKey, int] = {}
        available = 0
        overspent = 0
        for year, month in _month_range(start, end):
            available = self._apply_opening(year, month, first_active, opening, available)
            available += overspent + self.income_sums.get((year, month), 0)
            available -= self._budgeted_total(year, month)
            source = self._year_sheet(year)
            overspent = self._reconcile_month(year, month, source, balances)
            available = self._apply_seam_seed(year, month, available)
            self._record_available_residual(year, month, source, available)

    def _range(self) -> tuple[tuple[int, int], tuple[int, int], tuple[int, int] | None]:
        years = sorted(set(self.sheets.archive_years) | set(self.sheets.live_years))
        first_sheet = self.sheets.archive_years.get(years[0]) or self.sheets.live_years[years[0]]
        start = years[0], min(first_sheet.months)
        end = years[-1], 12
        first_active, last_active = _activity_span(self.transactions, self.sheets.sources)
        if last_active is not None and last_active < end:
            end = max(last_active, start)
            if self.sheets.seam_sheet is not None:
                end = max(end, (self.sheets.seam_year, 12))
        return start, end, first_active

    def _opening_balance(self, first_active: tuple[int, int] | None) -> int | None:
        if first_active is None:
            return None
        seeds = {
            (year, month): value
            for years in (self.sheets.archive_years, self.sheets.live_years)
            for year, source in years.items()
            for month, value in source.seeds.items()
        }
        return seeds.get(first_active)

    def _apply_opening(
        self,
        year: int,
        month: int,
        first_active: tuple[int, int] | None,
        opening: int | None,
        available: int,
    ) -> int:
        if (year, month) != first_active or opening is None:
            return available
        delta = opening - available
        if abs(delta) <= ADJUST_TOLERANCE_KOP:
            return available
        previous_year, previous_month = (year - 1, 12) if month == 1 else (year, month - 1)
        self.synthetic.append(
            _synthetic(
                (previous_year, previous_month), delta, self.income_category, OPENING_DESCRIPTION
            ),
        )
        self.income_sums[(previous_year, previous_month)] = (
            self.income_sums.get((previous_year, previous_month), 0) + delta
        )
        self.opened = opening, year, month
        return available + delta

    def _budgeted_total(self, year: int, month: int) -> int:
        return sum(
            self.budget_map.get((category.group, category.name, year, month), 0)
            for category in self.catalog.categories
            if self.kinds[(category.group, category.name)] != "income"
        )

    def _year_sheet(self, year: int) -> YearSheetRow:
        source = self.sheets.live_years.get(year) or self.sheets.archive_years.get(year)
        if source is None:
            msg = f"missing year sheet: {year}"
            raise WorkbookError(msg)
        return source

    def _reconcile_month(
        self,
        year: int,
        month: int,
        source: YearSheetRow,
        balances: dict[CategoryKey, int],
    ) -> int:
        overspent = 0
        for category in self.catalog.categories:
            key = (category.group, category.name)
            if self.kinds[key] == "income":
                continue
            projected = self._reconcile_category(year, month, source, key, balances)
            overspent += min(projected, 0)
        return overspent

    def _reconcile_category(
        self,
        year: int,
        month: int,
        source: YearSheetRow,
        key: CategoryKey,
        balances: dict[CategoryKey, int],
    ) -> int:
        group, name = key
        have = self.tx_sums.get((group, name, year, month), 0)
        projected = (
            max(balances.get(key, 0), 0) + self.budget_map.get((group, name, year, month), 0) + have
        )
        attempt = _CategoryBalance(key, projected, have, balances)
        desired = self._desired_balance(year, month, source, attempt)
        delta = 0 if desired is None else desired - projected
        if abs(delta) > ADJUST_TOLERANCE_KOP:
            self._record_category_adjustment(year, month, key, delta, have)
            projected += delta
        balances[key] = projected
        return projected

    def _desired_balance(
        self,
        year: int,
        month: int,
        source: YearSheetRow,
        attempt: "_CategoryBalance",
    ) -> int | None:
        if self._at_seam(year, month):
            return self._seam_balance(attempt.key)
        entry = _year_entry(source, attempt.key)
        if entry is not None:
            balance = entry.balances.get(month)
            return (
                attempt.projected - attempt.have + entry.outflows[month]
                if balance is None and month in entry.outflows
                else balance
            )
        if self._is_archived_final_balance(year, month, source, attempt):
            return 0
        return None

    def _seam_balance(self, key: CategoryKey) -> int | None:
        seam_sheet = self.sheets.seam_sheet
        if seam_sheet is None:
            return None
        last_month = max(seam_sheet.months)
        entry = _year_entry(seam_sheet, key)
        if entry is not None and (balance := entry.balances.get(last_month)) is not None:
            return balance
        first_live = min(self.sheets.live_years) if self.sheets.live_years else None
        if first_live is not None and key not in self.sheets.live_years[first_live].cats:
            return 0
        return None

    def _is_archived_final_balance(
        self,
        year: int,
        month: int,
        source: YearSheetRow,
        attempt: "_CategoryBalance",
    ) -> bool:
        return (
            year not in self.sheets.live_years
            and attempt.balances.get(attempt.key, 0) != 0
            and month == max(source.months)
        )

    def _record_category_adjustment(
        self,
        year: int,
        month: int,
        key: CategoryKey,
        delta: int,
        have: int,
    ) -> None:
        if self._at_seam(year, month):
            self.n_seam += 1
        else:
            self._count_adjustment(year, month)
        group, name = key
        self.synthetic.append(_synthetic((year, month), delta, name, name, group=group))
        self.tx_sums[(group, name, year, month)] = have + delta

    def _apply_seam_seed(self, year: int, month: int, available: int) -> int:
        if not self._at_seam(year, month):
            return available
        first_live = min(self.sheets.live_years) if self.sheets.live_years else None
        seed = None if first_live is None else self.sheets.live_years[first_live].seed
        if seed is None or abs(seed - available) <= ADJUST_TOLERANCE_KOP:
            return available
        delta = seed - available
        self.synthetic.append(
            _synthetic((year, month), delta, self.income_category, self.income_category)
        )
        self.income_sums[(year, month)] = self.income_sums.get((year, month), 0) + delta
        self.n_seam += 1
        return available + delta

    def _record_available_residual(
        self,
        year: int,
        month: int,
        source: YearSheetRow,
        available: int,
    ) -> None:
        if year in self.sheets.live_years:
            target = source.available.get(month)
            if target is not None and abs(target - available) > VERIFY_TOLERANCE_KOP:
                self.avail_residuals.append((year, month, target - available))

    def _at_seam(self, year: int, month: int) -> bool:
        return self.sheets.seam_sheet is not None and (year, month) == (self.sheets.seam_year, 12)

    def _count_adjustment(self, year: int, month: int) -> None:
        if (year, month) in self.months_with_rows:
            self.n_adjust += 1
        else:
            self.n_hist += 1

    def _warnings(self) -> list[str]:
        warnings: list[str] = []
        self._append_adjustment_warnings(warnings)
        self._append_opening_warning(warnings)
        self._append_available_warning(warnings)
        return warnings

    def _append_adjustment_warnings(self, warnings: list[str]) -> None:
        if self.n_hist:
            warnings.append(
                f"history: {self.n_hist} transactions stand in for months the sheet keeps only as "
                "monthly"
                " totals, with no rows of their own — one per category per month, so those months"
                " still add up",
            )
        if self.n_adjust:
            warnings.append(
                f"reconciliation: {self.n_adjust} adjustment transactions align live months with "
                "the sheet",
            )
        if self.n_seam:
            warnings.append(f"seam: {self.n_seam} carry corrections at {self.sheets.seam_year}-12")

    def _append_opening_warning(self, warnings: list[str]) -> None:
        if self.opened is not None:
            amount, year, month = self.opened
            warnings.append(
                f"opening balance: {amount / 100:,.2f} was already there when the sheet's first "
                "month"
                f" with rows ({year}-{month:02d}) began — imported as one transaction, since no "
                "row in the"
                " sheet accounts for it",
            )

    def _append_available_warning(self, warnings: list[str]) -> None:
        if self.avail_residuals:
            (first_year, first_month, first_delta), (last_year, last_month, last_delta) = (
                self.avail_residuals[0],
                self.avail_residuals[-1],
            )
            warnings.append(
                f"verify: the sheet's own Available differs from the one rebuilt from its rows in "
                f"{len(self.avail_residuals)} months, from {first_delta / 100:,.2f} "
                f"({first_year}-{first_month:02d}) to {last_delta / 100:,.2f} "
                f"({last_year}-{last_month:02d}) — every row, budget and carried balance is "
                "imported"
                " as the sheet has it, so a gap that only grows in months with overspending is the"
                " sheet's own header formula, not a mis-read row",
            )


@dataclass(slots=True)
class _CategoryBalance:
    key: CategoryKey
    projected: int
    have: int
    balances: dict[CategoryKey, int]


def _category_kinds(
    groups: Iterable[WorkbookGroupRow],
    categories: Iterable[WorkbookCategoryRow],
) -> dict[CategoryKey, str]:
    group_kinds = {group.name: group.kind for group in groups}
    return {
        (category.group, category.name): group_kinds.get(category.group, "expense")
        for category in categories
    }


def _transaction_totals(
    transactions: Iterable[WorkbookTransactionRow],
    kinds: Mapping[CategoryKey, str],
) -> tuple[dict[tuple[str, str, int, int], int], dict[tuple[int, int], int]]:
    tx_sums: dict[tuple[str, str, int, int], int] = {}
    income_sums: dict[tuple[int, int], int] = {}
    for transaction in transactions:
        name = transaction.monori_category
        if name:
            candidates = [key for key in kinds if key[1] == name]
            key = (transaction.monori_category_group, name)
            if not transaction.monori_category_group and len(candidates) == 1:
                key = candidates[0]
            year, month = int(transaction.date[:4]), int(transaction.date[5:7])
            if kinds.get(key) == "income":
                income_sums[(year, month)] = income_sums.get((year, month), 0) + transaction.amount
            else:
                total_key = key[0], key[1], year, month
                tx_sums[total_key] = tx_sums.get(total_key, 0) + transaction.amount
    return tx_sums, income_sums


def _budget_totals(
    budgets: Iterable[WorkbookBudgetRow],
) -> dict[tuple[str, str, int, int], int]:
    totals: dict[tuple[str, str, int, int], int] = {}
    for budget in budgets:
        key = budget.group, budget.category, budget.year, budget.month
        totals[key] = totals.get(key, 0) + budget.amount
    return totals


@dataclass(slots=True)
class _ParsedRows:
    groups: Iterable[WorkbookGroupRow]
    categories: Iterable[WorkbookCategoryRow]
    transactions: Iterable[WorkbookTransactionRow]
    budgets: Iterable[WorkbookBudgetRow]


def _result(
    rows: _ParsedRows,
    warnings: list[str],
    errors: Iterable[WorkbookParseErrorRow],
) -> ParsedWorkbook:
    return ParsedWorkbook(
        groups=list(rows.groups),
        categories=list(rows.categories),
        transactions=list(rows.transactions),
        budgets=list(rows.budgets),
        warnings=warnings,
        errors=list(errors),
    )

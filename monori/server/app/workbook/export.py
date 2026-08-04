"""
Builds the YNAB-style export workbook from a snapshot dict.

The workbook layout (sheet names, glyphs, column orders) is defined in
``spec.py`` and shared with the future spreadsheet importer so the two
directions never drift. The visual language — a slate header band, mint
summary rows, blue group bands, grid borders and a red/green balance — is
lifted from the reference template so the export reads as a designed document
rather than a raw dump.
"""

import datetime
from collections import defaultdict
from dataclasses import dataclass
from io import BytesIO

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from monori.server.app.deps import (
    CategoryResponse,
    GroupResponse,
    SnapshotResponse,
    TransactionResponse,
)
from monori.server.app.workbook import spec


@dataclass(frozen=True, slots=True)
class EffectiveTransaction:
    """Represent EffectiveTransaction."""

    id: int | str
    date: str
    account_id: int
    amount: int
    bank_category: str
    mcc: str
    description: str
    category_id: int | None
    comment: str
    transfer_id: str | None


BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center")

HEADER_FONT = Font(bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="FF3C464D")
SUMMARY_FONT = Font(bold=True, color="FF434343")
SUMMARY_FILL = PatternFill("solid", fgColor="FFEEF5E7")
GROUP_FILL = PatternFill("solid", fgColor="FFE6F4FB")
NUMBER_FONT = Font(color="FF434343")
POSITIVE_FONT = Font(color="FF4F7A00")
NEGATIVE_FONT = Font(color="FFC0392B")
_GRID_SIDE = Side(style="thin", color="FFD9D9D9")
GRID_BORDER = Border(left=_GRID_SIDE, right=_GRID_SIDE, top=_GRID_SIDE, bottom=_GRID_SIDE)


def _parse_dt(value: str) -> datetime.datetime:
    try:
        return datetime.datetime.fromisoformat(value)
    except ValueError:
        return datetime.datetime.fromisoformat(value[:10])


def _text(value: str) -> str:
    if value.startswith(("=", "+", "@")):
        return "'" + value
    return value


def _style_header(ws: Worksheet, row: int) -> None:
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def _fill_band(ws: Worksheet, row: int, last_col: int, fill: PatternFill, font: Font) -> None:
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font


def _balance_font(kop: int) -> Font:
    if kop > 0:
        return POSITIVE_FONT
    if kop < 0:
        return NEGATIVE_FONT
    return NUMBER_FONT


def _money_cell(ws: Worksheet, row: int, col: int, kop: int) -> Cell:
    cell = ws.cell(row=row, column=col, value=spec.kop_to_rub(kop))
    cell.number_format = spec.MONEY_FORMAT
    cell.font = NUMBER_FONT
    cell.border = GRID_BORDER
    return cell


def _categories_sheet(ws: Worksheet, snap: SnapshotResponse) -> None:
    ws.title = spec.SHEET_CATEGORIES
    ws.append(spec.CATEGORY_HEADERS)
    _style_header(ws, 1)
    by_group = defaultdict(list)
    for cat in snap.categories:
        by_group[cat.group_id].append(cat)
    for group in snap.groups:
        display = spec.group_display(group.name, group.kind)
        for cat in by_group[group.id]:
            ws.append([group.sort, _text(display), _text(cat.name), _text(cat.keywords)])
    ws.append([])
    ws.append(spec.GROUP_HEADERS)
    _style_header(ws, ws.max_row)
    for group in snap.groups:
        ws.append(
            [
                _text(spec.group_display(group.name, group.kind)),
                group.sort,
                spec.group_type(group.kind),
            ],
        )
    ws.freeze_panes = "A2"


def _effective_transaction(tx: TransactionResponse) -> EffectiveTransaction:
    return EffectiveTransaction(
        id=tx.id,
        date=tx.date,
        account_id=tx.account_id,
        amount=tx.amount,
        bank_category=tx.bank_category,
        mcc=tx.mcc,
        description=tx.description,
        category_id=tx.category_id,
        comment=tx.comment,
        transfer_id=tx.transfer_id,
    )


def _effective_transactions(snap: SnapshotResponse) -> list[EffectiveTransaction]:
    """Yield categorized split parts instead of their uncategorized container."""
    effective: list[EffectiveTransaction] = []
    for tx in snap.transactions:
        if not tx.splits:
            effective.append(_effective_transaction(tx))
            continue
        effective.extend(
            [
                EffectiveTransaction(
                    id=f"{tx.id}:{part.id}",
                    date=tx.date,
                    account_id=tx.account_id,
                    amount=part.amount,
                    bank_category=tx.bank_category,
                    mcc=tx.mcc,
                    description=tx.description,
                    category_id=part.category_id,
                    comment=part.comment,
                    transfer_id=tx.transfer_id,
                )
                for part in tx.splits
            ],
        )
    return effective


def _transactions_sheet(
    ws: Worksheet,
    snap: SnapshotResponse,
    cat_names: dict[int, str],
    acct_names: dict[int, str],
    acct_currency: dict[int, str],
) -> None:
    ws.append(spec.TRANSACTION_HEADERS)
    _style_header(ws, 1)
    row = 2
    for tx in _effective_transactions(snap):
        dt = _parse_dt(tx.date)
        currency = acct_currency.get(tx.account_id, "RUB")
        rub = spec.kop_to_rub(tx.amount)
        ws.cell(row=row, column=1, value=dt.strftime("%d.%m.%Y %H:%M:%S"))
        ws.cell(row=row, column=2, value=dt.strftime("%d.%m.%Y"))
        ws.cell(row=row, column=3, value=dt.strftime("%d.%m.%Y"))
        ws.cell(row=row, column=4, value="")
        ws.cell(row=row, column=5, value="OK")
        _money_cell(ws, row, 6, tx.amount)
        ws.cell(row=row, column=7, value=currency)
        ws.cell(row=row, column=8, value=spec.amount_display(rub, currency))
        ws.cell(row=row, column=9, value=currency)
        ws.cell(row=row, column=10, value="")
        ws.cell(row=row, column=11, value=_text(tx.bank_category))
        ws.cell(row=row, column=12, value=_text(tx.mcc))
        ws.cell(row=row, column=13, value=_text(tx.description))
        category_id = tx.category_id
        category_name = cat_names.get(category_id, "") if category_id is not None else ""
        ws.cell(row=row, column=14, value=_text(category_name))
        ws.cell(row=row, column=15, value=_text(acct_names.get(tx.account_id, "")))
        ws.cell(row=row, column=16, value=_text(tx.comment))
        row += 1
    ws.freeze_panes = "A2"


def _month_activity(snap: SnapshotResponse) -> dict[tuple[int, int, int], int]:
    activity: dict[tuple[int, int, int], int] = defaultdict(int)
    for tx in _effective_transactions(snap):
        if tx.category_id is None:
            continue
        dt = _parse_dt(tx.date)
        activity[(tx.category_id, dt.year, dt.month)] += tx.amount
    return activity


def _budget_index(snap: SnapshotResponse) -> dict[tuple[int, int, int], int]:
    budgets: dict[tuple[int, int, int], int] = {}
    for cell in snap.budgets:
        budgets[(cell.category_id, cell.year, cell.month)] = cell.amount
    return budgets


def _year_sheet(
    ws: Worksheet,
    year: int,
    snap: SnapshotResponse,
    activity: dict[tuple[int, int, int], int],
    budgets: dict[tuple[int, int, int], int],
) -> None:
    total_col = _year_headers(ws)
    data = _YearSheetData(year, snap, activity, budgets, total_col)
    hero = _write_year_groups(ws, data)
    _write_year_summary(ws, data, hero)
    _style_year_sheet(ws, total_col)


@dataclass(frozen=True, slots=True)
class _YearSheetData:
    year: int
    snapshot: SnapshotResponse
    activity: dict[tuple[int, int, int], int]
    budgets: dict[tuple[int, int, int], int]
    total_col: int


@dataclass(frozen=True, slots=True)
class _YearGroup:
    group: GroupResponse
    categories: list[CategoryResponse]


@dataclass(frozen=True, slots=True)
class _YearCategory:
    category: CategoryResponse
    expense: bool


def _year_headers(ws: Worksheet) -> int:
    for m in range(12):
        col = 2 + m * 3
        head = ws.cell(row=1, column=col, value=spec.MONTHS[m])
        head.alignment = CENTER
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
        for i, label in enumerate(spec.MONTH_COLS):
            ws.cell(row=2, column=col + i, value=label)
    total_col = 2 + 12 * 3
    ws.cell(row=1, column=total_col, value="Total")
    ws.cell(row=1, column=total_col + 1, value="Average")
    return total_col


def _write_year_groups(ws: Worksheet, data: _YearSheetData) -> dict[int, list[int]]:
    by_group: defaultdict[int, list[CategoryResponse]] = defaultdict(list)
    for category in data.snapshot.categories:
        by_group[category.group_id].append(category)
    hero = {month: [0, 0, 0] for month in range(1, 13)}
    row = 4
    for group in data.snapshot.groups:
        categories = by_group[group.id]
        if categories:
            row = _write_year_group(ws, row, data, hero, _YearGroup(group, categories))
    return hero


def _write_year_group(
    ws: Worksheet,
    row: int,
    data: _YearSheetData,
    hero: dict[int, list[int]],
    section: _YearGroup,
) -> int:
    group = section.group
    ws.cell(row=row, column=1, value=_text(spec.group_display(group.name, group.kind)))
    _fill_band(ws, row, data.total_col + 1, GROUP_FILL, BOLD)
    row += 1
    for category in section.categories:
        _write_year_category(ws, row, data, hero, _YearCategory(category, group.kind == "expense"))
        row += 1
    return row + 1


def _write_year_category(
    ws: Worksheet,
    row: int,
    data: _YearSheetData,
    hero: dict[int, list[int]],
    item: _YearCategory,
) -> None:
    category = item.category
    ws.cell(row=row, column=1, value=_text(category.name))
    balance = total_out = 0
    for month in range(1, 13):
        budgeted = data.budgets.get((category.id, data.year, month), 0)
        signed = data.activity.get((category.id, data.year, month), 0)
        out = -signed if item.expense else signed
        balance += budgeted - out
        col = 2 + (month - 1) * 3
        _money_cell(ws, row, col, budgeted)
        _money_cell(ws, row, col + 1, out)
        _money_cell(ws, row, col + 2, balance).font = _balance_font(balance)
        total_out += out
        if item.expense:
            hero[month][0] += budgeted
            hero[month][1] += out
    _money_cell(ws, row, data.total_col, total_out)
    _money_cell(ws, row, data.total_col + 1, round(total_out / 12))


def _write_year_summary(ws: Worksheet, data: _YearSheetData, hero: dict[int, list[int]]) -> None:
    hero_balance = 0
    ws.cell(row=3, column=1, value="Month Summary")
    hero_total = 0
    summary_balances: list[tuple[int, int]] = []
    for month in range(1, 13):
        budgeted, out, _ = hero[month]
        hero_balance += budgeted - out
        col = 2 + (month - 1) * 3
        _money_cell(ws, 3, col, budgeted)
        _money_cell(ws, 3, col + 1, out)
        _money_cell(ws, 3, col + 2, hero_balance)
        summary_balances.append((col + 2, hero_balance))
        hero_total += out
    _money_cell(ws, 3, data.total_col, hero_total)
    _money_cell(ws, 3, data.total_col + 1, round(hero_total / 12))
    _fill_band(ws, 3, data.total_col + 1, SUMMARY_FILL, SUMMARY_FONT)
    for col, bal in summary_balances:
        ws.cell(row=3, column=col).font = _balance_font(bal)


def _style_year_sheet(ws: Worksheet, total_col: int) -> None:
    _fill_band(ws, 1, total_col + 1, HEADER_FILL, HEADER_FONT)
    ws.cell(row=1, column=2).alignment = CENTER
    _fill_band(ws, 2, total_col + 1, HEADER_FILL, HEADER_FONT)

    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 24
    for c in range(2, total_col + 2):
        ws.column_dimensions[get_column_letter(c)].width = 11


def _dashdata_sheet(
    ws: Worksheet,
    snap: SnapshotResponse,
    activity: dict[tuple[int, int, int], int],
) -> None:
    ws.append(spec.DASH_HEADERS)
    _style_header(ws, 1)
    _write_monthly_dashdata(ws, snap)
    _write_category_year_totals(ws, snap, activity)
    ws.freeze_panes = "A2"


def _write_monthly_dashdata(ws: Worksheet, snap: SnapshotResponse) -> None:
    monthly: defaultdict[tuple[int, int], list[int]] = defaultdict(lambda: [0, 0])
    kinds = {g.id: g.kind for g in snap.groups}
    cat_kind = {c.id: kinds[c.group_id] for c in snap.categories}
    for tx in _effective_transactions(snap):
        if tx.transfer_id:
            continue
        category_id = tx.category_id
        if category_id is None:
            continue
        kind = cat_kind.get(category_id)
        if kind is None:
            continue
        dt = _parse_dt(tx.date)
        key = (dt.year, dt.month)
        if kind == "income":
            monthly[key][0] += tx.amount
        else:
            monthly[key][1] -= tx.amount
    cum_net = 0
    row = 2
    for year, month in sorted(monthly):
        income, expense = monthly[(year, month)]
        cum_net += income - expense
        ws.cell(row=row, column=1, value=f"{year:04d}-{month:02d}")
        _money_cell(ws, row, 2, income)
        _money_cell(ws, row, 3, expense)
        ratio = round(expense / income, 2) if income else ""
        ws.cell(row=row, column=4, value=ratio)
        _money_cell(ws, row, 5, cum_net)
        row += 1


def _write_category_year_totals(
    ws: Worksheet,
    snap: SnapshotResponse,
    activity: dict[tuple[int, int, int], int],
) -> None:
    years = sorted({y for (_, y, _m) in activity})
    if years:
        row = ws.max_row + 2
        ws.cell(row=row, column=1, value="Category").font = BOLD
        for i, year in enumerate(years):
            ws.cell(row=row, column=2 + i, value=year).font = BOLD
        for cat in snap.categories:
            row += 1
            ws.cell(row=row, column=1, value=_text(cat.name))
            for i, year in enumerate(years):
                total = sum(activity.get((cat.id, year, m), 0) for m in range(1, 13))
                _money_cell(ws, row, 2 + i, total)


def build_workbook(snap: SnapshotResponse) -> Workbook:
    """Handle build workbook."""
    cat_names = {c.id: c.name for c in snap.categories}
    acct_names = {a.id: a.name for a in snap.accounts}
    acct_currency = {a.id: a.currency for a in snap.accounts}
    activity = _month_activity(snap)
    budgets = _budget_index(snap)

    wb = Workbook()
    active = wb.active
    if active is None:
        msg = "workbook has no active worksheet"
        raise RuntimeError(msg)
    _categories_sheet(active, snap)
    _transactions_sheet(
        wb.create_sheet(spec.SHEET_TRANSACTIONS),
        snap,
        cat_names,
        acct_names,
        acct_currency,
    )
    years = sorted({y for (_, y, _m) in activity} | {b.year for b in snap.budgets})
    for year in years:
        _year_sheet(wb.create_sheet(str(year)), year, snap, activity, budgets)
    _dashdata_sheet(wb.create_sheet(spec.SHEET_DASHDATA), snap, activity)
    return wb


def workbook_bytes(snap: SnapshotResponse) -> bytes:
    """Handle workbook bytes."""
    buf = BytesIO()
    build_workbook(snap).save(buf)
    return buf.getvalue()
